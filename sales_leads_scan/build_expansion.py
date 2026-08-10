# -*- coding: utf-8 -*-
"""客户扩列 Excel：天眼查相似企业扩列(Top30带联系方式) + 专利信号线索 + 全部119候选"""
import csv, json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DIR = "/Users/zhangang/Desktop/excel share/excel-share-system/sales_leads_scan"

with open(f"{DIR}/_candidates.json", encoding="utf-8") as fh:
    cands = json.load(fh)

# 联系方式合并
contact = {}
for f in ("tyc_contact1.csv", "tyc_contact2.csv"):
    with open(f"{DIR}/{f}", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            n = (r.get("_query_company") or r.get("name") or "").strip()
            if n:
                contact[n] = {"phone": (r.get("phoneNumber") or "").strip(),
                              "email": (r.get("email") or "").strip(),
                              "staff": (r.get("staffNumRange") or "").strip(),
                              "scope": (r.get("businessScope") or "")[:80],
                              "city": (r.get("city") or "").strip()}

wb = Workbook()
hf = PatternFill("solid", fgColor="1F4E79")
hfont = Font(color="FFFFFF", bold=True)
gf = {"A": PatternFill("solid", fgColor="FFE699"), "B": PatternFill("solid", fgColor="DDEBF7")}

def style_sheet(ws, headers, rows, widths, grade_col=2):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == grade_col and v in gf:
                cell.fill = gf[v]; cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

# Sheet1: 扩列 Top30
top = cands[:30]
rows1 = []
for r in top:
    c = contact.get(r["name"], {})
    grade = "A" if r["score"] >= 90 else "B"
    rows1.append([r["score"], grade, r["name"], r["niche"], r["base"], r["est"], r["cap"],
                  c.get("phone", "（未取，可二次补充）"), c.get("staff", ""), c.get("scope", "")])
ws1 = wb.active
ws1.title = "天眼查扩列Top30"
style_sheet(ws1, ["评分", "等级", "企业名称", "细分赛道", "省份", "成立", "注册资本", "电话", "人员规模", "经营范围(截)"],
            rows1, [6, 6, 32, 13, 8, 8, 14, 22, 10, 60])

# Sheet2: 专利信号线索
rows2 = [
    [92, "A", "广东乐远化学材料科技有限公司", "东莞｜2012｜500万",
     "公开无卤阻燃有机硅导热灌封胶基料专利（CN122427643A，2026-07-22）",
     "灌封胶基料高填充粉体混合——真空行星搅拌机的典型工况；小企业决策快，专利公开=正在产品化",
     "天眼查补联系方式后电访"],
    [88, "A", "成都相变科技有限公司", "成都｜2025｜222万",
     "公开电磁屏蔽相变凝胶制备方法专利（CN122377383A，2026-07-15），摘要明确写『高速剪切或行星搅拌分散』",
     "2025 年新成立的研发型公司，专利到量产必经中试——5L/10L/20L 实验机型精准匹配",
     "高校/孵化园渠道或电访"],
    [82, "B", "惠州东铭新能源材料股份有限公司", "惠州｜股份公司",
     "公开高延展性导热硅胶材料专利（CN122404999A，2026-07-17），另有聚氨酯泡棉等多条材料专利",
     "导热硅胶+泡棉产品线持续扩张，导热硅胶制备正是双行星搅拌+压料场景",
     "天眼查补联系方式后电访"],
    [60, "C", "深圳市飞荣达科技股份有限公司", "深圳｜上市公司300602",
     "公开石墨烯/CNT 柔性导热垫片专利（CN122404995A，2026-07-17）",
     "需求真实但上市公司供应商体系成熟，建档跟踪其扩产/新基地动态",
     "长期跟踪"],
    [55, "C", "深圳市信维通信股份有限公司", "深圳｜上市公司",
     "液态金属导热膏、复合吸波材料等多项材料专利（2026-03 公开）",
     "通信大厂自研导热/吸波材料，观察其材料产线外包或扩产机会",
     "长期跟踪"],
]
ws2 = wb.create_sheet("专利信号线索")
style_sheet(ws2, ["评分", "等级", "企业名称", "概况", "专利信号", "推荐理由", "建议动作"],
            rows2, [6, 6, 30, 20, 52, 52, 20])

# Sheet3: 全部候选
rows3 = [[r["score"], r["name"], r["niche"], r["base"], r["est"], r["cap"], r["match"]] for r in cands]
ws3 = wb.create_sheet("全部候选119家")
style_sheet(ws3, ["评分", "企业名称", "细分赛道", "省份", "成立", "注册资本", "匹配方式"],
            rows3, [6, 34, 14, 8, 8, 16, 14], grade_col=1)

# Sheet4: 筛选逻辑
ws4 = wb.create_sheet("扩列逻辑")
logic = [
    ["项", "说明"],
    ["种子画像", "来自 ERP 销售台账 123 条记录的 50+ 家真实成交客户：胶粘剂/密封胶、电子材料（导热/灌封/浆料）、锂电材料、工业涂料、药业膏体"],
    ["天眼查查询", "增值服务-企业高级搜索，6 组关键词（密封胶/灌封胶/导热/电子浆料/负极材料/工业涂料）× 化工制造行业代码 GB/T 4754-26，各取 20 条"],
    ["清洗", "剔除：现有成交客户、设备/机械/贸易同行（名称正则）、非存续状态"],
    ["评分", "地域（江浙沪皖 30｜其他 10）+ 成立年限（2018 后 25｜2010 后 15｜更早 5）+ 注册资本（≥1000 万 25｜≥500 万 15）+ 匹配方式（名称 20｜经营范围 12｜其他 6）"],
    ["联系方式", "Top10 已调天眼查企业基本信息接口补电话；其余可按需二次补充"],
    ["来源说明", "天眼查第三方企业数据库，联系方式以企业年报/公示登记为准，首次触达建议先核实"],
]
for r, row in enumerate(logic, 1):
    for c, v in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if r == 1: cell.fill = hf; cell.font = hfont
ws4.column_dimensions["A"].width = 14
ws4.column_dimensions["B"].width = 110

out = f"{DIR}/客户扩列_2026-07-23.xlsx"
wb.save(out)
print("saved:", out)
