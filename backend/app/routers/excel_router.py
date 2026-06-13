"""Excel 导入导出"""
import json
import re
import tempfile
from datetime import datetime, date
from io import BytesIO
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from openpyxl import load_workbook, Workbook

from ..database import get_db
from .. import models, schemas
from ..deps import (
    get_current_user, require_not_viewer,
    user_can_view_project, user_can_edit_project,
)
from ..sheet_templates import SHEET_TEMPLATES, is_known_sheet, map_excel_to_template

router = APIRouter(prefix="/api", tags=["Excel 导入导出"])


# ============== 类型推断 ==============
DATE_PATTERN = re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$")
NUMBER_PATTERN = re.compile(r"^-?\d+(\.\d+)?$")


def _infer_field_type(samples: list[Any]) -> str:
    """推断字段类型。只保留 text / number / date 三种（不做 select，全部可自由输入）。"""
    non_null = [s for s in samples if s is not None and s != ""]
    if not non_null:
        return "text"
    # 全是 datetime / date
    if all(isinstance(s, (datetime, date)) for s in non_null):
        return "date"
    # 全是 number
    if all(isinstance(s, (int, float)) and not isinstance(s, bool) for s in non_null):
        return "number"
    strs = [str(s).strip() for s in non_null]
    if all(DATE_PATTERN.match(s) for s in strs):
        return "date"
    if all(NUMBER_PATTERN.match(s) for s in strs):
        return "number"
    # 其它一律 text（可输入编辑）
    return "text"


def _normalize_value(v: Any, ftype: str) -> Any:
    """根据字段类型规范化值，用于入库。"""
    if v is None or v == "":
        return None
    if ftype == "number":
        try:
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return v
            return float(v) if "." in str(v) else int(v)
        except Exception:
            return str(v)
    if ftype == "date":
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, date):
            return v.strftime("%Y-%m-%d")
        s = str(v).replace("/", "-")
        if DATE_PATTERN.match(s):
            return s
        return s
    if ftype in ("select", "multi_select"):
        return str(v).strip()
    return str(v).strip()


def _fmt_preamble_cell(v: Any, is_date: bool = False, datemode: int = 0) -> str:
    """格式化 preamble 单元格：日期序列号转日期、整数浮点去 .0"""
    if v is None or v == "":
        return ""
    if is_date:
        try:
            from xlrd.xldate import xldate_as_datetime
            dt = xldate_as_datetime(v, datemode)
            if dt.hour or dt.minute:
                return dt.strftime("%Y-%m-%d %H:%M")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            pass
    if isinstance(v, datetime):
        if v.hour or v.minute:
            return v.strftime("%Y-%m-%d %H:%M")
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


# 固定格式：第 5 行（index=4）是数据列名，第 6 行（index=5）起是数据
# 第 1-4 行是公司标题 / 项目信息区，导入时保留但不作为数据
HEADER_ROW_INDEX = 3   # 0-indexed；第 4 行是数据表列头，前 3 行是公司标题/项目信息


def _fill_merged_data_cells(
    rows: list[list[Any]],
    merged_ranges: list[tuple[int, int, int, int]],
    header_row_idx: int,
) -> None:
    """填充"数据区"的合并单元格：
    Excel 模板里常用合并表达"分类"（如"钣金装配"列跨多行），
    导入时 openpyxl/xlrd 只在左上角返回值，其余给 None/''。
    这里把合并范围内的所有单元格都填上左上角的值，方便每行独立可读。

    - merged_ranges 统一为闭区间 [rlo, rhi] × [clo, chi]（0-indexed）
    - 只处理 rlo > header_row_idx 的合并（项目头 0..header_row_idx-1
      和表头行 header_row_idx 都不应该填充，否则会把"品牌"等表头
      值塞到原本空的标头位置，造成后续 fallback 列识别错乱）
    - 仅当目标位置为 None / '' / '-' 时才填，避免覆盖手填值
    """
    for (rlo, rhi, clo, chi) in merged_ranges:
        if rlo <= header_row_idx:
            continue  # 项目头 + 表头行的合并不动
        if rlo >= len(rows) or clo >= len(rows[rlo]):
            continue
        top_left = rows[rlo][clo]
        if top_left in (None, '', '-'):
            continue
        for r in range(rlo, min(rhi + 1, len(rows))):
            row_len = len(rows[r])
            for c in range(clo, min(chi + 1, row_len)):
                if r == rlo and c == clo:
                    continue
                if rows[r][c] in (None, '', '-'):
                    rows[r][c] = top_left


# 已知数据表表头里会出现的列名（用于自动定位"列名表头行"）。
# 用模板所有字段名 + 旧格式列名 + 行号，覆盖面足够宽。
_HEADER_KEYWORDS: set[str] = set()
for _fields in SHEET_TEMPLATES.values():
    _HEADER_KEYWORDS.update(_fields)
_HEADER_KEYWORDS.update({
    '序号', '#', 'No', 'No.',
    # 旧格式（钣金装配 改名前）列名，兼容老文件
    '钣金/钳工', '钣金发出日期', '钣金完成日期',
    '封板/抛光', '封板发出日期', '封板完成日期',
})


def _detect_header_row(rows: list, max_scan: int = 12) -> int:
    """自动定位"列名表头行"：在前 max_scan 行里，挑与已知表头列名
    匹配最多的那一行。

    这样同时支持两种 Excel 排版：
    - 老格式：前几行是公司标题 / 项目信息，第 4 行（HEADER_ROW_INDEX）才是列头
    - 精简格式：第 1 行直接就是列头，没有项目头区

    匹配数 >= 3 才认定；否则回退 HEADER_ROW_INDEX（再不够就第 1 行）。
    """
    best_idx, best_score = -1, 0
    for i in range(min(max_scan, len(rows))):
        row = rows[i] or []
        score = 0
        for cell in row:
            name = str(cell).strip() if cell is not None else ''
            if name in _HEADER_KEYWORDS:
                score += 1
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx >= 0 and best_score >= 3:
        return best_idx
    # 回退：老逻辑（假设前 3 行是项目头）
    if len(rows) > HEADER_ROW_INDEX:
        return HEADER_ROW_INDEX
    return 0


def _find_data_header_row(rows: list) -> int:
    """兼容旧调用名 —— 现在走自动检测。"""
    return _detect_header_row(rows)


# 表格自带 # 行号列，这些字段名导入时会跳过（避免视觉冗余）
ROWNUM_FIELD_NAMES = {"序号", "#", "no", "no.", "序", "行号", "index"}

# 导入时若 header 为空，会 fallback 成"列{n}"。这种格式 + 整列空 = 空白尾列
_FALLBACK_HEADER_RE = re.compile(r'^列\d+$')


def is_empty_filler_column(hname: str, col_values: list) -> bool:
    """判断是否是 Excel 模板的"空白填充列"。

    场景：Excel 文件本身宽度大（如 35 列）但用户只填了前 16 列，
    后面 19 列全是空。导入时这些列 header 是 ""，被 fallback 成
    "列17/列18..."，且数据也都是空。需要跳过避免表格过宽撑爆窗口。

    判定：header 是空 / "列N" 自动命名，且整列数据为空或 '-'。
    """
    name = (hname or '').strip()
    if name and not _FALLBACK_HEADER_RE.match(name):
        return False  # 用户给了真实列名，即使数据为空也保留
    for v in col_values:
        if v not in (None, '', '-'):
            return False
    return True


def is_rownum_field_name(name: str) -> bool:
    """判断字段名是否明显是"行号"性质（表格 # 列已经有了，重复）"""
    if not name:
        return False
    return name.strip().lower() in ROWNUM_FIELD_NAMES


def _is_rownum_column(values: list[Any]) -> bool:
    """判断一列的所有值是否是连续整数 1, 2, 3, ...（允许空）。"""
    non_null = [v for v in values if v not in (None, "")]
    if not non_null:
        return False
    seq = []
    for v in non_null:
        try:
            if isinstance(v, bool):
                return False
            if isinstance(v, (int, float)):
                if isinstance(v, float) and not v.is_integer():
                    return False
                seq.append(int(v))
            else:
                s = str(v).strip()
                if not s.isdigit():
                    return False
                seq.append(int(s))
        except Exception:
            return False
    # 全是 1,2,3,...,n
    return seq == list(range(1, len(seq) + 1))


# ============== 导入 ==============
@router.post("/projects/{pid}/import-excel", response_model=schemas.Msg)
async def import_excel(
    pid: int, file: UploadFile = File(...),
    current: models.User = Depends(require_not_viewer),
    db: AsyncSession = Depends(get_db),
):
    """上传 Excel：每个 sheet → 一个 datasheet。自动识别表头 + 字段类型。"""
    res = await db.execute(
        select(models.Project).where(models.Project.id == pid, models.Project.is_deleted == False)
    )
    p = res.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "项目不存在")
    if not await user_can_edit_project(db, current, p):
        raise HTTPException(403, "无权导入")
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise HTTPException(400, "仅支持 .xlsx/.xlsm/.xls")
    content = await file.read()
    tmp = Path(tempfile.gettempdir()) / f"_imp_{pid}_{file.filename}"
    tmp.write_bytes(content)

    # 每个 sheet 的"表头合并延续列"（用户场景：如 F、G 合并显示"品牌"，
    # G 列本身不算单独的字段，跳过它）
    sheet_header_merge_skip: dict[str, set[int]] = {}

    try:
        sheets_meta: list[tuple[str, list[str], list[list[Any]]]] = []
        if suffix == ".xls":
            import xlrd
            from xlrd.xldate import xldate_as_datetime
            # 老 .xls 多为 cp936/GBK；先尝试 cp936，失败则用默认
            try:
                wb = xlrd.open_workbook(str(tmp), formatting_info=True, encoding_override="cp936")
            except Exception:
                wb = xlrd.open_workbook(str(tmp), formatting_info=True)
            for si in range(wb.nsheets):
                ws = wb.sheet_by_index(si)
                if ws.nrows < 1:
                    continue
                # 1) 先把整张 sheet 读到 full_rows（含空行，保持 0-indexed 对齐）
                full_rows: list[list[Any]] = []
                for r in range(ws.nrows):
                    row = []
                    for c in range(ws.ncols):
                        v = ws.cell_value(r, c)
                        ct = ws.cell_type(r, c)
                        if ct == 3:
                            try:
                                v = xldate_as_datetime(v, wb.datemode)
                            except Exception:
                                pass
                        row.append(v)
                    full_rows.append(row)
                # 2) 先自动定位"列名表头行"（兼容有/无项目头两种排版）
                header_idx = _detect_header_row(full_rows)
                # 3) 数据区合并填充：xlrd 的 (rlo, rhi, clo, chi) 中 rhi/chi 是 exclusive，
                #    转成 inclusive 0-indexed 给统一函数处理；用检测到的表头行做基准
                merged = [(rlo, rhi - 1, clo, chi - 1)
                          for (rlo, rhi, clo, chi) in ws.merged_cells]
                _fill_merged_data_cells(full_rows, merged, header_idx)
                # 3b) 收集"表头行的合并延续列" —— 这些列虽然在 Excel 里
                # 有视觉占位，但语义上属于左侧合并大列，不算单独的字段
                header_merge_skip: set[int] = set()
                for (rlo, rhi, clo, chi) in ws.merged_cells:
                    # xlrd: rhi/chi exclusive；表头行 = header_idx
                    if rlo <= header_idx < rhi:
                        for c in range(clo + 1, chi):
                            header_merge_skip.add(c)
                sheet_header_merge_skip[ws.name] = header_merge_skip
                # 4) 取 preamble / headers / rows
                preamble = []
                for ri in range(header_idx):
                    line = []
                    for c in range(ws.ncols):
                        v = full_rows[ri][c]
                        ct = ws.cell_type(ri, c)
                        line.append(_fmt_preamble_cell(v, is_date=(ct == 3), datemode=wb.datemode))
                    preamble.append(line)
                headers = [str(full_rows[header_idx][c]).strip() or f"列{c+1}"
                           for c in range(ws.ncols)]
                rows = []
                for r in range(header_idx + 1, len(full_rows)):
                    row = full_rows[r]
                    if any(v not in (None, "") for v in row):
                        rows.append(row)
                sheets_meta.append((ws.name, headers, rows, preamble))
        else:
            wb = load_workbook(tmp, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                # 先按当前内容自动定位表头行（兼容有/无项目头两种排版）
                _preview = list(ws.iter_rows(values_only=True))
                if not _preview:
                    continue
                header_idx = _detect_header_row(_preview)
                # 数据区合并：在 ws 上 unmerge + 把左上角值填充到所有单元格
                # 之后正常 iter_rows 就能拿到填好的数据
                # 先收集表头合并延续列（unmerge 之前）
                header_merge_skip_xlsx: set[int] = set()
                for merge in list(ws.merged_cells.ranges):
                    min_row_0 = merge.min_row - 1
                    max_row_0 = merge.max_row - 1
                    if min_row_0 <= header_idx <= max_row_0:
                        for c in range(merge.min_col, merge.max_col + 1):  # 1-indexed
                            # 转 0-indexed，跳过左上角（min_col）
                            if c > merge.min_col:
                                header_merge_skip_xlsx.add(c - 1)
                sheet_header_merge_skip[sn] = header_merge_skip_xlsx

                for merge in list(ws.merged_cells.ranges):
                    # openpyxl: min_row / max_row 都是 1-indexed inclusive
                    # 数据区 = 表头行的下一行起；表头行 + 项目头都不动
                    if (merge.min_row - 1) <= header_idx:
                        continue  # 项目头 + 表头区合并不动
                    top_left = ws.cell(row=merge.min_row, column=merge.min_col).value
                    if top_left in (None, "", "-"):
                        continue
                    min_r, max_r = merge.min_row, merge.max_row
                    min_c, max_c = merge.min_col, merge.max_col
                    ws.unmerge_cells(str(merge))
                    for r in range(min_r, max_r + 1):
                        for c in range(min_c, max_c + 1):
                            if r == min_r and c == min_c:
                                continue
                            cur = ws.cell(row=r, column=c).value
                            if cur in (None, "", "-"):
                                ws.cell(row=r, column=c).value = top_left

                rows_iter = list(ws.iter_rows(values_only=True))
                if not rows_iter:
                    continue
                # 保留 header_idx 之前所有行作为 preamble（前 4 行）
                preamble = []
                for r in rows_iter[:header_idx]:
                    preamble.append([_fmt_preamble_cell(c) for c in r])
                hdr_row = rows_iter[header_idx]
                headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(hdr_row)]
                rows = []
                for r in rows_iter[header_idx + 1:]:
                    row = list(r)
                    if any(v not in (None, "") for v in row):
                        rows.append(row)
                sheets_meta.append((sn, headers, rows, preamble))
    except Exception as e:
        raise HTTPException(400, f"解析失败：{e}")

    if not sheets_meta:
        raise HTTPException(400, "Excel 中没有可识别的数据")

    # ===== 模板对齐：已知 sheet 类型按模板重排字段（钣金装配 / 标准件清单 / 外协外购 / 原料下料单）=====
    aligned_meta = []
    for sname, headers, rows, preamble in sheets_meta:
        if is_known_sheet(sname):
            # 传入"表头合并延续列"：如 F、G 合并显示"品牌"，G 不算单独字段
            skip = sheet_header_merge_skip.get(sname, set())
            res = map_excel_to_template(sname, headers, rows,
                                         header_merge_skip_cols=skip)
            if res:
                tpl_headers, tpl_rows = res
                aligned_meta.append((sname, tpl_headers, tpl_rows, preamble))
                continue
        aligned_meta.append((sname, headers, rows, preamble))
    sheets_meta = aligned_meta

    # ===== 全量替换：先删本项目现有数据表（含字段、记录、字段权限） =====
    # 🆕 v3：白名单保护「电工采购单」第 5 表——它由电工部上传生成、非用户导入文件的一部分，
    # 重导四表 Excel 不应误删它（最小侵入：删除范围排除该表）。
    from sqlalchemy import delete as _del
    from ..sheet_templates import ELEC_PO_SHEET_NAME
    PROTECT = (ELEC_PO_SHEET_NAME,)
    # 待删 datasheet id（排除受保护表）
    ds_ids_res = await db.execute(
        select(models.Datasheet.id).where(
            models.Datasheet.project_id == pid,
            models.Datasheet.name.notin_(PROTECT),
        )
    )
    ds_ids_to_del = [r[0] for r in ds_ids_res.all()]
    field_ids = []
    if ds_ids_to_del:
        fres = await db.execute(
            select(models.Field.id).where(models.Field.datasheet_id.in_(ds_ids_to_del))
        )
        field_ids = [r[0] for r in fres.all()]
    # 数 records / datasheets 用于回显
    rres = await db.execute(
        select(func.count(models.Record.id))
        .join(models.Datasheet, models.Record.datasheet_id == models.Datasheet.id)
        .where(models.Datasheet.project_id == pid)
    )
    old_record_count = rres.scalar() or 0
    old_ds_count = len(ds_ids_to_del)
    # 显式删整链：field_permissions / fields / records / datasheets（仅非保护表）
    if field_ids:
        await db.execute(_del(models.FieldPermission).where(models.FieldPermission.field_id.in_(field_ids)))
        await db.execute(_del(models.Field).where(models.Field.id.in_(field_ids)))
    if ds_ids_to_del:
        await db.execute(_del(models.Record).where(models.Record.datasheet_id.in_(ds_ids_to_del)))
        await db.execute(_del(models.Datasheet).where(models.Datasheet.id.in_(ds_ids_to_del)))
    # 受保护的电工采购单排到末尾（重导的四表占 0..N，它保持最后一个 tab）
    from sqlalchemy import update as _upd
    await db.execute(
        _upd(models.Datasheet).where(
            models.Datasheet.project_id == pid,
            models.Datasheet.name.in_(PROTECT),
        ).values(sort_order=100)
    )
    await db.flush()
    # =====================================================================

    # 入库：每个 sheet 一个 datasheet
    total_records = 0
    # 导入的表从 0 开始排（全量替换四表）；受保护的电工采购单已被置 sort_order=100 留在末尾
    base_order = 0
    from datetime import datetime as _dt, timezone as _tz
    for idx, (sname, headers, rows, preamble) in enumerate(sheets_meta):
        d = models.Datasheet(
            project_id=pid, name=sname, sort_order=base_order + idx,
            header_lines=json.dumps(preamble, ensure_ascii=False) if preamble else None,
            imported_at=_dt.now(_tz.utc),  # 🆕 v3 P-16：导入标记（D1 四表校验依据）
        )
        db.add(d)
        await db.flush()  # 拿到 d.id

        # 推断各列类型 + 识别需要跳过的列
        # - "序号"列：避免和表格 # 列重复
        # - 空白尾列：Excel 模板宽但实际未用到的尾列（"列N" + 数据全空），
        #            否则表格被撑得很宽，窗口卡得很难缩放
        col_count = len(headers)
        fields: list[models.Field] = []
        skipped_cols: set[int] = set()
        for ci, hname in enumerate(headers):
            col_samples = [r[ci] if ci < len(r) else None for r in rows[:50]]
            # 1) 序号列：内容是连续整数 1,2,3...
            if is_rownum_field_name(hname) and _is_rownum_column(col_samples):
                skipped_cols.add(ci)
                continue
            # 2) 空白填充列：用全量数据判断（不只是前 50 行样本，避免误判）
            col_full = [r[ci] if ci < len(r) else None for r in rows]
            if is_empty_filler_column(hname, col_full):
                skipped_cols.add(ci)
                continue
            ftype = _infer_field_type(col_samples)
            f = models.Field(
                datasheet_id=d.id, name=hname, type=ftype, sort_order=ci,
            )
            db.add(f)
            fields.append(f)
        await db.flush()

        # 行入库（跳过被忽略的列）
        for ri, row in enumerate(rows):
            values: dict[str, Any] = {}
            # 按 field 的源列序号去 row 里取值
            field_idx = 0
            for ci in range(col_count):
                if ci in skipped_cols:
                    continue
                f = fields[field_idx]
                field_idx += 1
                if ci < len(row):
                    nv = _normalize_value(row[ci], f.type)
                    if nv is not None:
                        values[str(f.id)] = nv
            r = models.Record(
                datasheet_id=d.id, sort_order=ri, values=values,
                created_by=current.id, updated_by=current.id,
            )
            db.add(r)
            total_records += 1

    await db.commit()
    return schemas.Msg(
        message=(
            f"导入完成：清除旧数据 {old_ds_count} 个数据表 / {old_record_count} 行 → "
            f"新建 {len(sheets_meta)} 个数据表 / {total_records} 行"
        )
    )


# ============== 导出公共工具：preamble 公式列实时计算 ==============
# 与前端 DatasheetGrid.vue 的 DATE_DERIVED_COLS / DATE_KEYS 保持一致
_PREAMBLE_DERIVED = {"货期", "已过时间", "已经过时间", "倒计时", "剩余天数", "剩余"}
_PREAMBLE_ORDER_KEYS = {"下单日期", "下单时间", "下单"}
_PREAMBLE_DELIVER_KEYS = {"交货日期", "交付日期", "交期", "交货时间"}
_DATE_HEAD_RE = re.compile(r"^(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})")


def _parse_loose_date(s: Any) -> Optional[date]:
    """松散解析日期字符串：识别 YYYY-MM-DD / YYYY/MM/DD / YYYY.MM.DD / YYYY年MM月DD日"""
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    text = str(s).strip()
    m = _DATE_HEAD_RE.match(text)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def _find_header_idx(header_row: list[Any], candidates: set[str]) -> int:
    for i, h in enumerate(header_row):
        k = str(h or "").strip()
        if k in candidates:
            return i
    return -1


def _compute_preamble_value(header_row: list[Any], value_row: list[Any], col_idx: int) -> Any:
    """对值行的某列：是公式列就实时计算，否则原样返回。

    与 DatasheetGrid.preambleCell 行为对齐：
      货期         = 交货日期 - 下单日期
      已过时间     = TODAY() - 下单日期
      倒计时       = 交货日期 - TODAY()
    """
    raw = value_row[col_idx] if col_idx < len(value_row) else ""
    header = str(header_row[col_idx] if col_idx < len(header_row) else "").strip()
    if header not in _PREAMBLE_DERIVED:
        return raw

    order_idx = _find_header_idx(header_row, _PREAMBLE_ORDER_KEYS)
    deliver_idx = _find_header_idx(header_row, _PREAMBLE_DELIVER_KEYS)
    order_date = _parse_loose_date(value_row[order_idx]) if 0 <= order_idx < len(value_row) else None
    deliver_date = _parse_loose_date(value_row[deliver_idx]) if 0 <= deliver_idx < len(value_row) else None
    today = date.today()

    if header == "货期":
        if order_date and deliver_date:
            return (deliver_date - order_date).days
        return raw
    if header in ("已过时间", "已经过时间"):
        if order_date:
            return (today - order_date).days
        return raw
    if header in ("倒计时", "剩余天数", "剩余"):
        if deliver_date:
            return (deliver_date - today).days
        return raw
    return raw


def _write_preamble_to_sheet(ws, header_lines_json: Optional[str]) -> None:
    """把 datasheet.header_lines（JSON 字符串）写到工作表顶部，
    其中值行（idx=2）的公式列按 TODAY 实时计算。"""
    if not header_lines_json:
        return
    try:
        lines = json.loads(header_lines_json)
    except Exception:
        return
    if not lines:
        return
    header_row = lines[1] if len(lines) > 1 else []
    for idx, line in enumerate(lines):
        if idx == 2 and header_row:
            row = [
                _compute_preamble_value(header_row, line, i)
                for i in range(len(line))
            ]
            ws.append(row)
        else:
            ws.append(list(line))


# ============== 导出 ==============
@router.get("/datasheets/{did}/export")
async def export_datasheet(
    did: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出单个数据表为 .xlsx"""
    from ..deps import ensure_can_export
    ensure_can_export(current)  # 🆕 M16 导出闸（默认关）
    res = await db.execute(select(models.Datasheet).where(models.Datasheet.id == did))
    d = res.scalar_one_or_none()
    if not d:
        raise HTTPException(404, "数据表不存在")
    pres = await db.execute(select(models.Project).where(models.Project.id == d.project_id))
    p = pres.scalar_one_or_none()
    if not p or not await user_can_view_project(db, current, p):
        raise HTTPException(403, "无权下载")

    fres = await db.execute(
        select(models.Field).where(models.Field.datasheet_id == did)
        .order_by(models.Field.sort_order, models.Field.id)
    )
    fields = fres.scalars().all()
    rres = await db.execute(
        select(models.Record).where(models.Record.datasheet_id == did)
        .order_by(models.Record.sort_order, models.Record.id)
    )
    records = rres.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = d.name[:31] or "Sheet1"
    # 1. 项目信息行（preamble）—— 公式列按 TODAY 实时算，与网页一致
    _write_preamble_to_sheet(ws, d.header_lines)
    # 2. 数据表头
    ws.append([f.name for f in fields])
    # 3. 数据
    for r in records:
        row = []
        for f in fields:
            v = (r.values or {}).get(str(f.id))
            if isinstance(v, list):
                v = "、".join(str(x) for x in v)
            row.append(v)
        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"{p.code}_{d.name}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.get("/projects/{pid}/export")
async def export_project(
    pid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出整个项目所有数据表为一个 .xlsx 多 sheet 文件"""
    from ..deps import ensure_can_export
    ensure_can_export(current)  # 🆕 M16 导出闸（默认关）
    res = await db.execute(
        select(models.Project).where(models.Project.id == pid, models.Project.is_deleted == False)
    )
    p = res.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "项目不存在")
    if not await user_can_view_project(db, current, p):
        raise HTTPException(403, "无权下载")

    dres = await db.execute(
        select(models.Datasheet).where(models.Datasheet.project_id == pid)
        .order_by(models.Datasheet.sort_order, models.Datasheet.id)
    )
    sheets = dres.scalars().all()
    if not sheets:
        raise HTTPException(404, "项目没有数据表")

    wb = Workbook()
    wb.remove(wb.active)
    for d in sheets:
        ws = wb.create_sheet(title=d.name[:31] or "Sheet")
        fres = await db.execute(
            select(models.Field).where(models.Field.datasheet_id == d.id)
            .order_by(models.Field.sort_order, models.Field.id)
        )
        fields = fres.scalars().all()
        rres = await db.execute(
            select(models.Record).where(models.Record.datasheet_id == d.id)
            .order_by(models.Record.sort_order, models.Record.id)
        )
        records = rres.scalars().all()
        # 1. 项目信息行（preamble）—— 公式列按 TODAY 实时算
        _write_preamble_to_sheet(ws, d.header_lines)
        # 2. 数据表头
        ws.append([f.name for f in fields])
        # 3. 数据
        for r in records:
            row = []
            for f in fields:
                v = (r.values or {}).get(str(f.id))
                if isinstance(v, list):
                    v = "、".join(str(x) for x in v)
                row.append(v)
            ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"{p.code}_{p.name}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )
