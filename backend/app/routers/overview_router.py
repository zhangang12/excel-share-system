"""项目一览：跨项目的总览表"""
import json
import tempfile
from datetime import datetime, date
from pathlib import Path
from typing import List, Any, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func

from ..database import get_db
from .. import models, schemas
from .ws_router import broadcast_cell_changed
from ..deps import (
    get_current_user, require_admin, require_admin_or_manager, require_not_viewer,
    user_can_view_project, user_can_edit_project,
)
from ..sheet_templates import OVERVIEW_FIELDS, OVERVIEW_HEADER_ALIAS
from .projects_router import OVERVIEW_KEY_PREFIX, HEADER_KEY_PREFIX

# 一览模板中可写入 meta 的字段标签集合
OVERVIEW_META_LABELS = {
    f['label'] for f in OVERVIEW_FIELDS if f['source'] == 'meta'
}
NAME_ALIASES = {'项目名称', '设备名称', '名称'}
CODE_ALIASES = {'项目编号', '项目代码', '编号'}
STATUS_ALIASES = {'状态'}

router = APIRouter(prefix="/api/overview", tags=["项目一览"])


def _field_to_out(f: models.OverviewField) -> schemas.OverviewFieldOut:
    cfg = None
    if f.config:
        try: cfg = json.loads(f.config)
        except Exception: cfg = None
    return schemas.OverviewFieldOut(
        id=f.id, name=f.name, type=f.type, sort_order=f.sort_order, config=cfg,
    )


def _project_row(p: models.Project) -> schemas.OverviewProjectRow:
    manager_name = None
    if p.manager:
        manager_name = p.manager.full_name or p.manager.username
    return schemas.OverviewProjectRow(
        id=p.id, code=p.code, name=p.name, status=p.status,
        description=p.description,
        manager_id=p.manager_id, manager_name=manager_name,
        extra=p.extra or {}, updated_at=p.updated_at,
    )


# ==================== 字段管理 ====================
@router.get("/fields", response_model=List[schemas.OverviewFieldOut])
async def list_fields(
    _: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(
        select(models.OverviewField).order_by(models.OverviewField.sort_order, models.OverviewField.id)
    )
    return [_field_to_out(f) for f in res.scalars().all()]


@router.post("/fields", response_model=schemas.OverviewFieldOut)
async def create_field(
    data: schemas.OverviewFieldCreate,
    _: models.User = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    if data.type not in schemas.FIELD_TYPES:
        raise HTTPException(400, f"字段类型必须是 {schemas.FIELD_TYPES}")
    res = await db.execute(select(models.OverviewField).where(models.OverviewField.name == data.name))
    if res.scalar_one_or_none():
        raise HTTPException(400, "字段名已存在")
    sres = await db.execute(select(func.max(models.OverviewField.sort_order)))
    max_order = sres.scalar() or -1
    f = models.OverviewField(
        name=data.name, type=data.type, sort_order=max_order + 1,
        config=json.dumps(data.config, ensure_ascii=False) if data.config else None,
    )
    db.add(f); await db.commit(); await db.refresh(f)
    return _field_to_out(f)


@router.put("/fields/{fid}", response_model=schemas.OverviewFieldOut)
async def update_field(
    fid: int, data: schemas.OverviewFieldUpdate,
    _: models.User = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(models.OverviewField).where(models.OverviewField.id == fid))
    f = res.scalar_one_or_none()
    if not f: raise HTTPException(404, "字段不存在")
    if data.name is not None: f.name = data.name
    if data.type is not None:
        if data.type not in schemas.FIELD_TYPES:
            raise HTTPException(400, f"字段类型必须是 {schemas.FIELD_TYPES}")
        f.type = data.type
    if data.sort_order is not None: f.sort_order = data.sort_order
    if data.config is not None:
        f.config = json.dumps(data.config, ensure_ascii=False) if data.config else None
    await db.commit(); await db.refresh(f)
    return _field_to_out(f)


@router.delete("/fields/{fid}", response_model=schemas.Msg)
async def delete_field(
    fid: int,
    _: models.User = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(select(models.OverviewField).where(models.OverviewField.id == fid))
    f = res.scalar_one_or_none()
    if not f: raise HTTPException(404, "字段不存在")
    # 清掉所有 project.extra 里这个字段的值
    fid_str = str(fid)
    pres = await db.execute(select(models.Project))
    for p in pres.scalars().all():
        if p.extra and fid_str in p.extra:
            extra = dict(p.extra); extra.pop(fid_str, None)
            p.extra = extra
    await db.delete(f); await db.commit()
    return schemas.Msg(message="已删除")


# ==================== 一览数据 ====================
@router.get("", response_model=schemas.OverviewBundle)
async def get_overview(
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    import time, logging
    log = logging.getLogger("overview")
    t0 = time.perf_counter()

    fres = await db.execute(
        select(models.OverviewField).order_by(models.OverviewField.sort_order, models.OverviewField.id)
    )
    fields = [_field_to_out(f) for f in fres.scalars().all()]
    t1 = time.perf_counter()

    pres = await db.execute(
        select(models.Project)
        .where(models.Project.is_deleted == False)
        .order_by(models.Project.code)
    )
    all_projects = pres.scalars().all()
    t2 = time.perf_counter()

    # admin / manager 免权限校验
    is_super = current.role and current.role.code in ("admin", "manager")
    if is_super:
        visible_rows = [_project_row(p) for p in all_projects]
    else:
        # 一次性查所有成员关系，避免 N+1
        mres = await db.execute(
            select(models.ProjectMember.project_id).where(
                models.ProjectMember.user_id == current.id
            )
        )
        member_pids = {row[0] for row in mres.all()}
        visible_rows = [
            _project_row(p) for p in all_projects if p.id in member_pids
        ]

    t3 = time.perf_counter()
    log.info("overview: fields=%d total=%d visible=%d  | t_fields=%.2fms t_proj=%.2fms t_filter=%.2fms",
             len(fields), len(all_projects), len(visible_rows),
             (t1-t0)*1000, (t2-t1)*1000, (t3-t2)*1000)
    return schemas.OverviewBundle(fields=fields, rows=visible_rows)


@router.get("/export")
async def export_overview(
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """导出项目一览为 .xlsx（按当前用户可见范围）"""
    from io import BytesIO
    from urllib.parse import quote
    from openpyxl import Workbook

    fres = await db.execute(
        select(models.OverviewField).order_by(models.OverviewField.sort_order, models.OverviewField.id)
    )
    fields = fres.scalars().all()

    pres = await db.execute(
        select(models.Project)
        .where(models.Project.is_deleted == False)
        .order_by(models.Project.code)
    )
    all_projects = pres.scalars().all()

    is_super = current.role and current.role.code in ("admin", "manager")
    if is_super:
        visible = list(all_projects)
    else:
        mres = await db.execute(
            select(models.ProjectMember.project_id).where(
                models.ProjectMember.user_id == current.id
            )
        )
        member_pids = {row[0] for row in mres.all()}
        visible = [p for p in all_projects if p.id in member_pids]

    wb = Workbook()
    ws = wb.active
    ws.title = "项目目录"
    ws.append(["项目编号", "项目名称", "状态", "项目经理"] + [f.name for f in fields])
    for p in visible:
        manager_name = ""
        if p.manager:
            manager_name = p.manager.full_name or p.manager.username
        row = [p.code, p.name, p.status, manager_name]
        extra = p.extra or {}
        for f in fields:
            v = extra.get(str(f.id))
            if isinstance(v, list):
                v = "、".join(str(x) for x in v)
            row.append(v)
        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = "项目目录.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.put("/projects/{pid}/cell", response_model=schemas.OverviewProjectRow)
async def update_overview_cell(
    pid: int, data: schemas.OverviewCellUpdate,
    current: models.User = Depends(require_not_viewer),
    db: AsyncSession = Depends(get_db),
):
    res = await db.execute(
        select(models.Project).where(models.Project.id == pid, models.Project.is_deleted == False)
    )
    p = res.scalar_one_or_none()
    if not p: raise HTTPException(404, "项目不存在")
    if not await user_can_edit_project(db, current, p):
        raise HTTPException(403, "无权修改")
    # 校验 field 存在
    fres = await db.execute(select(models.OverviewField).where(models.OverviewField.id == data.field_id))
    if not fres.scalar_one_or_none():
        raise HTTPException(404, "字段不存在")
    # 字段级权限：admin / manager 跳过
    if current.role.code not in ("admin", "manager"):
        fp_res = await db.execute(select(models.OverviewFieldPermission).where(
            models.OverviewFieldPermission.field_id == data.field_id,
            models.OverviewFieldPermission.role_id == current.role_id,
        ))
        fp = fp_res.scalar_one_or_none()
        if fp is not None and (not fp.can_view or not fp.can_edit):
            raise HTTPException(403, "无权编辑该字段")
    extra = dict(p.extra or {})
    if data.value in (None, ""):
        extra.pop(str(data.field_id), None)
    else:
        extra[str(data.field_id)] = data.value
    p.extra = extra
    await db.commit(); await db.refresh(p)
    await broadcast_cell_changed(
        scope="overview", scope_id=None,
        record_id=None, project_id=p.id,
        field_id=data.field_id, value=data.value, by_user_id=current.id,
    )
    return _project_row(p)


# ==================== 导入 ====================
def _infer_type(samples: list[Any]) -> str:
    from .excel_router import _infer_field_type
    return _infer_field_type(samples)


def _normalize(v, ftype):
    from .excel_router import _normalize_value
    return _normalize_value(v, ftype)


@router.post("/import", response_model=schemas.Msg)
async def import_overview(
    file: UploadFile = File(...),
    current: models.User = Depends(require_admin_or_manager),
    db: AsyncSession = Depends(get_db),
):
    """上传公司汇总 Excel：
       - 必须包含"项目编号"列（或类似 编号/项目代码）
       - 找到/新建项目；其它列存到 extra
       - 自动创建缺失的 OverviewField
    """
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise HTTPException(400, "仅支持 .xlsx/.xlsm/.xls")
    content = await file.read()
    tmp = Path(tempfile.gettempdir()) / f"_ovw_{file.filename}"
    tmp.write_bytes(content)

    # 解析（沿用 excel_router 的逻辑）
    sheets: list[tuple[str, list[str], list[list[Any]]]] = []
    # ===== 全量替换：清空所有项目 extra 中的 __o__ 前缀 key（仅一览数据）
    # 不动 __h__ 前缀（项目详情头表数据）和 OverviewField 旧自定义列
    pall = await db.execute(select(models.Project))
    for _p in pall.scalars().all():
        if _p.extra:
            cleaned = {k: v for k, v in _p.extra.items()
                       if not (isinstance(k, str) and k.startswith(OVERVIEW_KEY_PREFIX))}
            _p.extra = cleaned
    await db.flush()
    # ============================================

    try:
        if suffix == ".xls":
            import xlrd
            try:
                wb = xlrd.open_workbook(str(tmp), formatting_info=False, encoding_override="cp936")
            except Exception:
                wb = xlrd.open_workbook(str(tmp), formatting_info=False)
            for si in range(wb.nsheets):
                ws = wb.sheet_by_index(si)
                if ws.nrows < 2: continue
                # 在前 10 行中找含"项目编号"的行作为表头
                header_row = 0
                for r in range(min(10, ws.nrows)):
                    vals = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    if any('项目编号' in v or '编号' in v or '项目代码' in v for v in vals):
                        header_row = r; break
                headers = [str(ws.cell_value(header_row, c)).strip() or f'列{c+1}' for c in range(ws.ncols)]
                rows = []
                for r in range(header_row + 1, ws.nrows):
                    row = []
                    for c in range(ws.ncols):
                        v = ws.cell_value(r, c)
                        ct = ws.cell_type(r, c)
                        if ct == 3:
                            try:
                                from xlrd.xldate import xldate_as_datetime
                                v = xldate_as_datetime(v, wb.datemode)
                            except Exception: pass
                        row.append(v)
                    if any(v not in (None, '') for v in row):
                        rows.append(row)
                sheets.append((ws.name, headers, rows))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(tmp, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                all_rows = list(ws.iter_rows(values_only=True))
                if len(all_rows) < 2: continue
                # 找表头
                header_idx = 0
                for i in range(min(10, len(all_rows))):
                    vs = [str(v or '') for v in all_rows[i]]
                    if any('项目编号' in v or '项目代码' in v for v in vs):
                        header_idx = i; break
                headers = [str(h).strip() if h is not None else f'列{i+1}'
                           for i, h in enumerate(all_rows[header_idx])]
                rows = []
                for r in all_rows[header_idx + 1:]:
                    row = list(r)
                    if any(v not in (None, '') for v in row):
                        rows.append(row)
                sheets.append((sn, headers, rows))
    except Exception as e:
        raise HTTPException(400, f"解析失败：{e}")

    total_create = 0
    total_update = 0
    total_filled_cells = 0
    touched_pids: set[int] = set()

    for sname, headers, rows in sheets:
        # 找关键列：项目编号 / 项目名称 / 状态
        def _find_col(aliases: set[str]) -> Optional[int]:
            # 精确匹配优先
            for i, h in enumerate(headers):
                if h in aliases:
                    return i
            # 子串包含次之
            for i, h in enumerate(headers):
                for a in aliases:
                    if a in h:
                        return i
            return None

        code_col = _find_col(CODE_ALIASES)
        name_col = _find_col(NAME_ALIASES)
        status_col = _find_col(STATUS_ALIASES)
        if code_col is None:
            continue

        # 把 Excel 表头中"属于 OVERVIEW_FIELDS meta 列"的列做映射：
        # col_index -> meta label（如 col 4 → "签订日期"）
        col_to_meta: dict[int, str] = {}
        for ci, hname in enumerate(headers):
            h = (hname or '').strip()
            if not h or h.startswith('列'):
                continue
            if ci in (code_col, name_col, status_col):
                continue
            if h in OVERVIEW_META_LABELS:
                col_to_meta[ci] = h

        # 按 sheet 名推断默认状态
        sheet_status = '进行中'
        if any(k in sname for k in ('完成', '已完')):
            sheet_status = '已完成'

        import re as _re
        for row in rows:
            code = ''
            if code_col is not None and code_col < len(row) and row[code_col] is not None:
                code = str(row[code_col]).strip()
            if not code:
                continue
            # 排除明显不是项目编号的"说明"/"标题"等中文行
            if not _re.match(r'^[\w\-]+$', code) or len(code) > 32:
                continue

            pname = ''
            if name_col is not None and name_col < len(row) and row[name_col] is not None:
                pname = str(row[name_col]).strip()

            # 行级状态：优先 status_col，否则 sheet_status
            row_status = sheet_status
            if status_col is not None and status_col < len(row) and row[status_col] is not None:
                rv = str(row[status_col]).strip()
                if rv in ('进行中', '已完成', '已归档'):
                    row_status = rv

            res = await db.execute(select(models.Project).where(models.Project.code == code))
            p = res.scalar_one_or_none()
            if not p:
                p = models.Project(
                    code=code, name=pname or code,
                    status=row_status,
                    extra={},
                )
                db.add(p)
                await db.flush()
                touched_pids.add(p.id)
                total_create += 1
            else:
                # 首次在本次导入中遇到 → 清空 __o__* + 一览映射到的 __h__* key
                # （覆盖语义；保留无映射的 __h__ key，如项目详情专有的"数量/制表日期"）
                if p.id not in touched_pids:
                    h_keys_to_clear = {
                        f"{HEADER_KEY_PREFIX}{alias}"
                        for alias in OVERVIEW_HEADER_ALIAS.values()
                    }
                    cleaned = {
                        k: v for k, v in (p.extra or {}).items()
                        if not (isinstance(k, str) and (
                            k.startswith(OVERVIEW_KEY_PREFIX) or k in h_keys_to_clear
                        ))
                    }
                    p.extra = cleaned
                    touched_pids.add(p.id)
                    total_update += 1
                if pname:
                    p.name = pname
                if row_status:
                    p.status = row_status

            # 按 col_to_meta 填 __o__<label>，同时同步到 __h__<alias>
            extra = dict(p.extra or {})
            for ci, meta_label in col_to_meta.items():
                if ci >= len(row):
                    continue
                v = row[ci]
                storage_key = f"{OVERVIEW_KEY_PREFIX}{meta_label}"
                h_alias = OVERVIEW_HEADER_ALIAS.get(meta_label)
                h_storage_key = f"{HEADER_KEY_PREFIX}{h_alias}" if h_alias else None

                if v is None or v == '':
                    extra.pop(storage_key, None)
                    if h_storage_key:
                        extra.pop(h_storage_key, None)
                else:
                    # 日期类型规范化为 YYYY-MM-DD 字符串
                    if isinstance(v, datetime):
                        normalized = v.strftime('%Y-%m-%d')
                    elif isinstance(v, date):
                        normalized = v.strftime('%Y-%m-%d')
                    else:
                        normalized = str(v).strip()
                    extra[storage_key] = normalized
                    if h_storage_key:
                        extra[h_storage_key] = normalized  # 同步到项目详情
                    total_filled_cells += 1
            p.extra = extra

    await db.commit()
    return schemas.Msg(
        message=(
            f"导入完成：新建项目 {total_create} 个，更新 {total_update} 个，"
            f"填入 {total_filled_cells} 个一览字段值"
        )
    )
