"""🆕 人事部一期（设计稿见《人事部功能设计.md》）：员工花名册 + 部门月度工资总额。

- 权限：整个模块 hr 角色 +（admin/manager 由 require_roles 自动放行）；tab 级再走二级菜单权限。
- 花名册：CRUD + Excel 导入（部门按名称匹配 OA departments，不自动新建）；
  状态改「离职」且关联了登录账号 → 提醒 admin 停用账号。
- 工资总额：一月一部门一行，只到部门汇总（不碰个人工资）；
  盈利改善规划 §五-1「含分摊人工」毛利口径的数据源。
- 到期提醒（合同 30 天 / 转正 7 天）在 overdue.scan_hr_reminders，每日扫、7 天窗口去重。
"""
from datetime import date
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, and_

from ..database import get_db
from .. import models, schemas
from ..deps import require_roles
from ..notify import push_message
from ..utils import write_audit

router = APIRouter(prefix="/api/hr", tags=["人事部"])

_HR = ("hr",)   # admin/manager 由 require_roles 自动放行
_STATUSES = ("试用", "在职", "离职")


# ==================== schemas（模块内自用） ====================
class EmployeeIn(BaseModel):
    name: str = Field(min_length=1, max_length=64)
    department_id: Optional[int] = None
    position: Optional[str] = None
    hire_date: Optional[str] = None
    regular_date: Optional[str] = None
    contract_end: Optional[str] = None
    status: str = "在职"
    leave_date: Optional[str] = None
    id_card: Optional[str] = None
    phone: Optional[str] = None
    emergency_contact: Optional[str] = None
    emergency_contact_phone: Optional[str] = None
    user_id: Optional[int] = None
    note: Optional[str] = None


class EmployeeOut(EmployeeIn):
    id: int
    emp_no: Optional[str] = None
    department_name: Optional[str] = None
    user_name: Optional[str] = None


async def _next_emp_no(db: AsyncSession) -> str:
    """🆕 #202 下一个企业工号：现有数字工号 max+1，5 位补零；离职工号照样占位不复用。"""
    rows = (await db.execute(select(models.Employee.emp_no).where(
        models.Employee.emp_no.isnot(None)))).scalars().all()
    mx = max((int(x) for x in rows if x and x.isdigit()), default=0)
    return f"{mx + 1:05d}"


class RosterOut(BaseModel):
    rows: List[EmployeeOut]
    stats: dict


class PayrollRowIn(BaseModel):
    department_id: int
    total_amount: float = 0
    note: Optional[str] = None


class PayrollSaveIn(BaseModel):
    rows: List[PayrollRowIn]


# 🆕 考勤(按月每人) —— 人事录入
class AttendanceRowIn(BaseModel):
    employee_id: int
    should_days: float = 0      # 应出勤天数
    actual_days: float = 0      # 实出勤天数
    leave_days: float = 0       # 请假天数
    overtime_hours: float = 0   # 加班工时
    late_count: int = 0             # 🆕 #239 迟到次数
    early_leave_count: int = 0      # 🆕 #239 早退次数
    missing_card_count: int = 0     # 🆕 #239 缺卡次数
    note: Optional[str] = None


class AttendanceSaveIn(BaseModel):
    rows: List[AttendanceRowIn]


# 🆕 个人工资(按月每人) —— 人事录入,敏感(仅 hr+管理层)
class SalaryRowIn(BaseModel):
    employee_id: int
    base: float = 0             # 基本工资
    merit: float = 0            # 绩效/奖金
    overtime_pay: float = 0     # 加班费
    allowance: float = 0        # 补贴
    social_deduct: float = 0    # 社保公积金扣款
    personal_tax: float = 0     # 🆕 #248 个税
    other_deduct: float = 0     # 其他扣款
    note: Optional[str] = None


class SalarySaveIn(BaseModel):
    rows: List[SalaryRowIn]


# ==================== 员工花名册 ====================
def _emp_out(e: models.Employee, users: dict) -> EmployeeOut:
    return EmployeeOut(
        id=e.id, emp_no=e.emp_no, name=e.name, department_id=e.department_id,
        department_name=(e.department.name if e.department else None),
        position=e.position, hire_date=e.hire_date, regular_date=e.regular_date,
        contract_end=e.contract_end, status=e.status, leave_date=e.leave_date,
        id_card=e.id_card, phone=e.phone, emergency_contact=e.emergency_contact,
        emergency_contact_phone=e.emergency_contact_phone,
        user_id=e.user_id, user_name=users.get(e.user_id), note=e.note)


@router.get("/employees", response_model=RosterOut)
async def list_employees(
    status: Optional[str] = Query(None),
    department_id: Optional[int] = Query(None),
    kw: Optional[str] = Query(None),
    _: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    # 🆕 反馈#214：花名册按工号(emp_no)升序;无工号的排最后,再按 id 兜底
    q = select(models.Employee).order_by(
        models.Employee.emp_no.is_(None), models.Employee.emp_no, models.Employee.id)
    if status:
        q = q.where(models.Employee.status == status)
    if department_id:
        q = q.where(models.Employee.department_id == department_id)
    if kw:
        k = f"%{kw.strip()}%"
        q = q.where(models.Employee.name.like(k) | models.Employee.phone.like(k)
                    | models.Employee.position.like(k))
    emps = list((await db.execute(q)).scalars().all())
    uids = [e.user_id for e in emps if e.user_id]
    users: dict = {}
    if uids:
        ur = await db.execute(select(models.User).where(models.User.id.in_(uids)))
        users = {u.id: (u.full_name or u.username) for u in ur.scalars().all()}
    # KPI 用全量（不受筛选影响）
    allr = list((await db.execute(select(models.Employee))).scalars().all())
    today = date.today()
    month = today.isoformat()[:7]
    d30 = today.toordinal() + 30
    def _exp30(e):
        if e.status == "离职" or not e.contract_end:
            return False
        try:
            return date.fromisoformat(e.contract_end).toordinal() <= d30
        except ValueError:
            return False
    stats = {
        "active": sum(1 for e in allr if e.status in ("在职", "试用")),
        "probation": sum(1 for e in allr if e.status == "试用"),
        "expiring30": sum(1 for e in allr if _exp30(e)),
        "joined_month": sum(1 for e in allr if (e.hire_date or "")[:7] == month),
        "left_month": sum(1 for e in allr if (e.leave_date or "")[:7] == month),
    }
    return RosterOut(rows=[_emp_out(e, users) for e in emps], stats=stats)


class BindableUserOut(BaseModel):
    id: int
    label: str                       # 显示名：姓名（用户名）
    username: str
    bound_to: Optional[str] = None   # 已被别的员工绑定 → 前端禁选


@router.get("/bindable-users", response_model=List[BindableUserOut])
async def bindable_users(
    _: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """可绑定的系统登录账号（供员工表单「登录账号」选择器）；标注已被别的员工绑定的账号。"""
    urs = list((await db.execute(select(models.User).where(
        models.User.is_active == True).order_by(models.User.id))).scalars().all())  # noqa: E712
    bemps = list((await db.execute(select(models.Employee).where(
        models.Employee.user_id.isnot(None)))).scalars().all())
    bound = {e.user_id: e.name for e in bemps}
    out: List[BindableUserOut] = []
    for u in urs:
        nm = u.full_name or u.username
        label = f"{nm}（{u.username}）" if (u.full_name and u.full_name != u.username) else u.username
        out.append(BindableUserOut(id=u.id, label=label, username=u.username, bound_to=bound.get(u.id)))
    return out


async def _apply_emp(db: AsyncSession, e: models.Employee, body: EmployeeIn,
                     current: models.User) -> None:
    """公共赋值 + 离职联动：状态改离职且挂了登录账号 → 提醒 admin 停用账号。"""
    if body.status not in _STATUSES:
        raise HTTPException(400, f"状态须为 {'/'.join(_STATUSES)}")
    was = e.status
    for k, v in body.model_dump().items():
        setattr(e, k, (v.strip() if isinstance(v, str) else v) or (None if isinstance(v, str) else v))
    e.name = body.name.strip()
    if e.status == "离职" and not e.leave_date:
        e.leave_date = date.today().isoformat()
    await db.commit()
    if e.status == "离职" and was != "离职" and e.user_id:
        await push_message(db, to_role="admin", kind="warn",
                           text=f"【离职提醒】{e.name} 已登记离职，其登录账号(uid={e.user_id})请及时停用。",
                           biz_type="employee", biz_id=e.id)


async def _check_user_unique(db: AsyncSession, user_id: Optional[int], exclude_eid: Optional[int]) -> None:
    """🆕 登录账号唯一：一个系统账号只能绑一个员工。在员工写入前调用（避免半成品被 autoflush）。"""
    if user_id is None:
        return
    dq = select(models.Employee).where(models.Employee.user_id == user_id)
    if exclude_eid is not None:
        dq = dq.where(models.Employee.id != exclude_eid)
    dup = (await db.execute(dq)).scalars().first()
    if dup:
        raise HTTPException(400, f"该登录账号已绑定到员工「{dup.name}」，一个账号只能绑一个员工")


@router.post("/employees", response_model=EmployeeOut)
async def create_employee(
    body: EmployeeIn,
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    await _check_user_unique(db, body.user_id, None)
    e = models.Employee(emp_no=await _next_emp_no(db))   # 🆕 #202 自动分配工号
    db.add(e)
    await _apply_emp(db, e, body, current)
    await write_audit(db, user=current, action="hr_emp_create", target_type="employee",
                      target_id=e.id, detail=f"{e.emp_no} {e.name}")
    r = await db.execute(select(models.Employee).where(models.Employee.id == e.id))
    return _emp_out(r.scalar_one(), {})


@router.put("/employees/{eid}", response_model=EmployeeOut)
async def update_employee(
    eid: int, body: EmployeeIn,
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    e = (await db.execute(select(models.Employee).where(
        models.Employee.id == eid))).scalar_one_or_none()
    if not e:
        raise HTTPException(404, "员工不存在")
    await _check_user_unique(db, body.user_id, eid)
    await _apply_emp(db, e, body, current)
    await write_audit(db, user=current, action="hr_emp_update", target_type="employee",
                      target_id=e.id, detail=e.name)
    r = await db.execute(select(models.Employee).where(models.Employee.id == eid))
    return _emp_out(r.scalar_one(), {})


@router.delete("/employees/{eid}", response_model=schemas.Msg)
async def delete_employee(
    eid: int,
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """物理删除只用于录错的行；正常人员变动请用「离职」状态留痕。"""
    e = (await db.execute(select(models.Employee).where(
        models.Employee.id == eid))).scalar_one_or_none()
    if not e:
        raise HTTPException(404, "员工不存在")
    name = e.name
    await db.delete(e)
    await db.commit()
    await write_audit(db, user=current, action="hr_emp_delete", target_type="employee",
                      target_id=eid, detail=name)
    return schemas.Msg(message=f"已删除 {name}")


# ==================== 花名册 Excel 导入 ====================
_EMP_IMPORT_COLS = ["姓名*", "部门", "岗位", "入职日期", "转正日期", "合同到期日",
                    "身份证", "电话", "紧急联系人", "紧急联系人电话", "状态", "备注"]


def _norm_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _norm_date(v) -> Optional[str]:
    from datetime import datetime as _dt, date as _d
    if v in (None, ""):
        return None
    if isinstance(v, (_dt, _d)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().replace(".", "-").replace("/", "-")
    return s[:10] or None


@router.get("/employees/import-template")
async def emp_import_template(
    _: models.User = Depends(require_roles(*_HR)),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "员工导入"
    ws.append(_EMP_IMPORT_COLS)
    hfill = PatternFill("solid", fgColor="FEF3E8")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.append(["张三", "生产部", "装配工", "2025-03-01", "2025-06-01", "2027-02-28",
               "", "13800000000", "李四 139...", "在职", "示例行，导入前请删除"])
    for i, w in enumerate([10, 12, 12, 12, 12, 12, 22, 14, 20, 8, 20], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("填写说明")
    for line in [["员工花名册导入说明"], [""],
                 ["1. 带 * 的列为必填：姓名。"],
                 ["2. 部门须与「OA审批-部门」里的名称一致，匹配不上会跳过该行并提示（不会自动新建部门）。"],
                 ["3. 日期格式 YYYY-MM-DD；状态：试用 / 在职 / 离职（留空默认在职）。"],
                 ["4. 姓名已存在（非离职）时只补空缺字段，不覆盖已维护的资料。"],
                 ["5. 第 2 行为示例，导入前请删除。"]]:
        ws2.append(line)
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 80
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('员工导入模板.xlsx')}"})


@router.post("/employees/import")
async def import_employees(
    file: UploadFile = File(...),
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "无法解析该文件，请使用模板导出的 .xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件为空或只有表头")
    header = [(_norm_str(c) or "").replace("*", "").strip() for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header) if name}
    if "姓名" not in col:
        raise HTTPException(400, "表头缺少必填列「姓名」，请用模板")

    def cell(row, name):
        i = col.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    depts = {d.name: d.id for d in (await db.execute(
        select(models.Department))).scalars().all()}
    existing = {e.name: e for e in (await db.execute(select(models.Employee).where(
        models.Employee.status != "离职"))).scalars().all()}
    created = updated = 0
    errors: list[str] = []
    _emp_seq = int(await _next_emp_no(db))   # 🆕 #202 导入自动发号,本批递增
    for rn, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        name = _norm_str(cell(row, "姓名"))
        if not name:
            errors.append(f"第 {rn} 行：姓名为空，已跳过")
            continue
        dept_name = _norm_str(cell(row, "部门"))
        dept_id = depts.get(dept_name) if dept_name else None
        if dept_name and dept_id is None:
            errors.append(f"第 {rn} 行：部门「{dept_name}」不存在（请先在 OA-部门 里维护），已跳过")
            continue
        status = _norm_str(cell(row, "状态")) or "在职"
        if status not in _STATUSES:
            errors.append(f"第 {rn} 行：状态「{status}」无效（试用/在职/离职），已跳过")
            continue
        vals = dict(
            department_id=dept_id, position=_norm_str(cell(row, "岗位")),
            hire_date=_norm_date(cell(row, "入职日期")), regular_date=_norm_date(cell(row, "转正日期")),
            contract_end=_norm_date(cell(row, "合同到期日")), id_card=_norm_str(cell(row, "身份证")),
            phone=_norm_str(cell(row, "电话")), emergency_contact=_norm_str(cell(row, "紧急联系人")),
            emergency_contact_phone=_norm_str(cell(row, "紧急联系人电话")),
            note=_norm_str(cell(row, "备注")))
        e = existing.get(name)
        if e is None:
            e = models.Employee(name=name, status=status, emp_no=f"{_emp_seq:05d}", **vals)
            _emp_seq += 1
            db.add(e)
            existing[name] = e
            created += 1
        else:   # 已存在只补空缺
            for k, v in vals.items():
                if v and not getattr(e, k, None):
                    setattr(e, k, v)
            updated += 1
    await db.commit()
    await write_audit(db, user=current, action="hr_emp_import", target_type="employee",
                      detail=f"新建{created} 补全{updated} 跳过{len(errors)}")
    msg = f"导入完成：新建 {created} 人、补全 {updated} 人"
    if errors:
        msg += f"；{len(errors)} 行跳过"
    return {"message": msg, "created": created, "updated": updated, "errors": errors[:20]}


# ==================== 部门月度工资总额 ====================
@router.get("/payroll")
async def get_payroll(
    month: str = Query(..., description="YYYY-MM"),
    _: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """指定月份的各部门工资总额（未填的部门也返回,金额 0 供录入）。"""
    depts = list((await db.execute(select(models.Department).where(
        models.Department.enabled == True)  # noqa: E712
        .order_by(models.Department.sort_order, models.Department.id))).scalars().all())
    cur = {r.department_id: r for r in (await db.execute(
        select(models.PayrollMonthly).where(models.PayrollMonthly.month == month))).scalars().all()}
    rows = [{"department_id": d.id, "department_name": d.name,
             "total_amount": (cur[d.id].total_amount if d.id in cur else 0),
             "note": (cur[d.id].note if d.id in cur else None)} for d in depts]
    return {"month": month, "rows": rows,
            "total": round(sum(r["total_amount"] or 0 for r in rows), 2)}


@router.put("/payroll/{month}", response_model=schemas.Msg)
async def save_payroll(
    month: str,
    body: PayrollSaveIn,
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    if len(month) != 7 or month[4] != "-":
        raise HTTPException(400, "month 格式应为 YYYY-MM")
    cur = {r.department_id: r for r in (await db.execute(
        select(models.PayrollMonthly).where(models.PayrollMonthly.month == month))).scalars().all()}
    n = 0
    for r in body.rows:
        row = cur.get(r.department_id)
        if row is None:
            db.add(models.PayrollMonthly(month=month, department_id=r.department_id,
                                         total_amount=r.total_amount or 0,
                                         note=(r.note or "").strip() or None))
        else:
            row.total_amount = r.total_amount or 0
            row.note = (r.note or "").strip() or None
        n += 1
    await db.commit()
    await write_audit(db, user=current, action="hr_payroll_save", target_type="payroll",
                      detail=f"{month} {n} 个部门")
    return schemas.Msg(message=f"已保存 {month} 工资总额（{n} 个部门）")


# ==================== 🆕 考勤(按月每人,人事录入) ====================
async def _active_emps(db: AsyncSession, period: Optional[str] = None) -> list[models.Employee]:
    """某月在册员工(在职/试用 + 该月仍在职的离职者),按工号(emp_no)升序,供考勤/工资逐人录入。
    🆕 反馈#225：与花名册(#214)一致,考勤/工资也按工号排,无工号排最后。
    🆕 反馈#229：离职当月仍要发工资/记考勤 —— 原先一标离职就整个从表里消失,
      最后一个月工资没法发。改为按离职日期判断：离职日期所在月(及之前)仍在列,之后不在。
      离职但没填离职日期的按老口径排除(数据不全,无法判断哪个月离的)。
      period=None(不传)时退化为老口径(仅在职/试用),供不按月的场景用。"""
    q = select(models.Employee)
    if period:
        q = q.where(or_(
            models.Employee.status != "离职",
            and_(models.Employee.leave_date.is_not(None),
                 func.substr(models.Employee.leave_date, 1, 7) >= period),
        ))
    else:
        q = q.where(models.Employee.status != "离职")
    return list((await db.execute(q.order_by(
        models.Employee.emp_no.is_(None), models.Employee.emp_no, models.Employee.id))).scalars().all())


def _is_full_attendance(a) -> bool:
    """🆕 #239 是否全勤——自动推导，不单独存字段(避免与明细打架)：
    迟到/早退/缺卡次数全为 0 且 请假 0 天 且 实出勤 ≥ 应出勤(且应出勤已填)。
    口径唯一：前端只展示，导入/导出/列表都走这里。"""
    if a is None:
        return False
    return bool((a.should_days or 0) > 0
                and (a.actual_days or 0) >= (a.should_days or 0)
                and (a.leave_days or 0) == 0
                and (a.late_count or 0) == 0
                and (a.early_leave_count or 0) == 0
                and (a.missing_card_count or 0) == 0)


@router.get("/attendance")
async def get_attendance(
    period: str = Query(..., description="YYYY-MM"),
    _: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """指定月份的员工考勤(该月在册员工都返回,未填的记 0 供录入)。
    🆕 #229 离职当月仍在列；🆕 #239 迟到/早退/缺卡 + 是否全勤(自动推导)。"""
    emps = await _active_emps(db, period)
    cur = {r.employee_id: r for r in (await db.execute(select(models.AttendanceMonthly).where(
        models.AttendanceMonthly.period == period))).scalars().all()}
    rows = []
    for e in emps:
        a = cur.get(e.id)
        rows.append({
            "employee_id": e.id, "emp_no": e.emp_no, "name": e.name,
            "department_name": (e.department.name if e.department else None),
            "should_days": (a.should_days if a else 0), "actual_days": (a.actual_days if a else 0),
            "leave_days": (a.leave_days if a else 0), "overtime_hours": (a.overtime_hours if a else 0),
            "late_count": (a.late_count if a else 0),
            "early_leave_count": (a.early_leave_count if a else 0),
            "missing_card_count": (a.missing_card_count if a else 0),
            "full_attendance": _is_full_attendance(a),
            "left_this_month": bool(e.status == "离职" and (e.leave_date or "")[:7] == period),
            "leave_date": e.leave_date,
            "note": (a.note if a else None)})
    return {"period": period, "rows": rows}


@router.put("/attendance/{period}", response_model=schemas.Msg)
async def save_attendance(
    period: str,
    body: AttendanceSaveIn,
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    if len(period) != 7 or period[4] != "-":
        raise HTTPException(400, "period 格式应为 YYYY-MM")
    cur = {r.employee_id: r for r in (await db.execute(select(models.AttendanceMonthly).where(
        models.AttendanceMonthly.period == period))).scalars().all()}
    n = 0
    for r in body.rows:
        row = cur.get(r.employee_id)
        if row is None:
            db.add(models.AttendanceMonthly(
                employee_id=r.employee_id, period=period,
                should_days=r.should_days or 0, actual_days=r.actual_days or 0,
                leave_days=r.leave_days or 0, overtime_hours=r.overtime_hours or 0,
                late_count=r.late_count or 0, early_leave_count=r.early_leave_count or 0,
                missing_card_count=r.missing_card_count or 0,
                note=(r.note or "").strip() or None))
        else:
            row.should_days = r.should_days or 0
            row.actual_days = r.actual_days or 0
            row.leave_days = r.leave_days or 0
            row.overtime_hours = r.overtime_hours or 0
            row.late_count = r.late_count or 0
            row.early_leave_count = r.early_leave_count or 0
            row.missing_card_count = r.missing_card_count or 0
            row.note = (r.note or "").strip() or None
        n += 1
    await db.commit()
    await write_audit(db, user=current, action="hr_attendance_save", target_type="attendance",
                      detail=f"{period} {n} 人")
    return schemas.Msg(message=f"已保存 {period} 考勤（{n} 人）")


# 🆕 #240 考勤批量导入：表头即模板列（顺序=界面列顺序，人事对着填）
_ATT_IMPORT_COLS = ["工号*", "姓名", "应出勤(天)", "实出勤(天)", "请假(天)",
                    "加班(工时)", "迟到(次)", "早退(次)", "缺卡(次)", "备注"]


@router.get("/attendance/import-template")
async def attendance_import_template(
    _: models.User = Depends(require_roles(*_HR)),
):
    """🆕 #240 下载考勤导入模板（.xlsx，含表头 + 示例行 + 填写说明）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤导入"
    ws.append(_ATT_IMPORT_COLS)
    hfill = PatternFill("solid", fgColor="E8F0FE")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.append(["00015", "张明宇（仅参考，导入按工号匹配）", 21.5, 21.5, 0, 8, 0, 0, 0, "示例行，导入前请删除"])
    for i, w in enumerate([10, 24, 12, 12, 10, 12, 10, 10, 10, 24], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("填写说明")
    for line in [
        ["考勤导入说明"],
        [""],
        ["1. 按「工号」匹配员工——工号必填且必须已在花名册存在，姓名列仅供人工核对，不参与匹配。"],
        ["2. 导入哪个月，由页面上选的月份决定（不在表里填月份），导入前请先确认页面月份选对。"],
        ["3. 同一工号在该月已有考勤 → 整行覆盖；没有 → 新建。表里没出现的人保持原样，不会被清空。"],
        ["4. 「是否全勤」不用填，系统自动算：迟到/早退/缺卡/请假都为 0 且 实出勤 ≥ 应出勤。"],
        ["5. 数字列留空按 0 处理。第 2 行为示例，导入前请删除。"],
        ["6. 离职员工：离职当月仍可导入（末月工资/考勤要发得出来），之后的月份不在名单里。"],
    ]:
        ws2.append(line)
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 84
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = "考勤导入模板.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.post("/attendance/{period}/import")
async def import_attendance(
    period: str,
    file: UploadFile = File(...),
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #240 批量导入某月考勤：按工号匹配，已有则整行覆盖，未出现的人不动。"""
    if len(period) != 7 or period[4] != "-":
        raise HTTPException(400, "period 格式应为 YYYY-MM")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "无法解析该文件，请使用模板导出的 .xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件为空或只有表头")
    header = [(str(c).strip().replace("*", "") if c is not None else "") for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header) if name}
    if "工号" not in col:
        raise HTTPException(400, "表头缺少必填列「工号」，请用模板")

    def cell(row, name):
        i = col.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    def num(v) -> float:
        if v is None or str(v).strip() == "":
            return 0.0
        try:
            return float(str(v).strip())
        except ValueError:
            return 0.0

    # 该月在册员工(含当月离职)：工号 -> 员工
    emp_by_no = {(e.emp_no or "").strip(): e for e in await _active_emps(db, period) if e.emp_no}
    cur = {r.employee_id: r for r in (await db.execute(select(models.AttendanceMonthly).where(
        models.AttendanceMonthly.period == period))).scalars().all()}

    created = updated = 0
    skipped: list[str] = []
    for row in rows[1:]:
        raw_no = cell(row, "工号")
        if raw_no is None or str(raw_no).strip() == "":
            continue
        # Excel 常把 00015 读成数字 15 → 补零回 5 位再匹配
        no = str(raw_no).strip()
        if no.isdigit() and no not in emp_by_no:
            no = no.zfill(5)
        e = emp_by_no.get(no)
        if not e:
            skipped.append(str(raw_no).strip())
            continue
        vals = dict(
            should_days=num(cell(row, "应出勤(天)")), actual_days=num(cell(row, "实出勤(天)")),
            leave_days=num(cell(row, "请假(天)")), overtime_hours=num(cell(row, "加班(工时)")),
            late_count=int(num(cell(row, "迟到(次)"))), early_leave_count=int(num(cell(row, "早退(次)"))),
            missing_card_count=int(num(cell(row, "缺卡(次)"))),
            note=(str(cell(row, "备注")).strip() if cell(row, "备注") not in (None, "") else None),
        )
        a = cur.get(e.id)
        if a is None:
            db.add(models.AttendanceMonthly(employee_id=e.id, period=period, **vals))
            created += 1
        else:
            for k, v in vals.items():
                setattr(a, k, v)
            updated += 1
    await db.commit()
    await write_audit(db, user=current, action="hr_attendance_import", target_type="attendance",
                      detail=f"{period} 新增{created} 覆盖{updated} 跳过{len(skipped)}")
    msg = f"导入完成：新增 {created} 人，覆盖 {updated} 人"
    if skipped:
        head = "、".join(skipped[:5]) + ("…" if len(skipped) > 5 else "")
        msg += f"；{len(skipped)} 个工号在该月名单里找不到，已跳过（{head}）"
    return {"message": msg, "created": created, "updated": updated, "skipped": skipped}


# ==================== 🆕 个人工资(按月每人,人事录入,敏感) ====================
def _salary_net(s) -> float:
    """实发 = 基本+绩效+加班费+补贴 − 社保公积金扣款 − 个税 − 其他扣款。"""
    return round((s.base or 0) + (s.merit or 0) + (s.overtime_pay or 0) + (s.allowance or 0)
                 - (s.social_deduct or 0) - (getattr(s, "personal_tax", 0) or 0) - (s.other_deduct or 0), 2)


@router.get("/salary")
async def get_salary(
    period: str = Query(..., description="YYYY-MM"),
    _: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """指定月份的个人工资(该月在册员工都返回,未填的记 0 供录入)。敏感,仅人事+管理层。
    🆕 #229：离职当月仍在列——最后一个月工资要发得出来。"""
    emps = await _active_emps(db, period)
    cur = {r.employee_id: r for r in (await db.execute(select(models.EmployeeSalaryMonthly).where(
        models.EmployeeSalaryMonthly.period == period))).scalars().all()}
    rows = []
    for e in emps:
        s = cur.get(e.id)
        rows.append({
            "employee_id": e.id, "emp_no": e.emp_no, "name": e.name,
            "department_name": (e.department.name if e.department else None),
            "base": (s.base if s else 0), "merit": (s.merit if s else 0),
            "overtime_pay": (s.overtime_pay if s else 0), "allowance": (s.allowance if s else 0),
            "social_deduct": (s.social_deduct if s else 0),
            "personal_tax": (s.personal_tax if s else 0), "other_deduct": (s.other_deduct if s else 0),
            "net": (_salary_net(s) if s else 0), "note": (s.note if s else None),
            # 🆕 #229 当月离职：人事需要一眼看出这是末月工资(否则容易照上月全额发)
            "left_this_month": bool(e.status == "离职" and (e.leave_date or "")[:7] == period),
            "leave_date": e.leave_date})
    return {"period": period, "rows": rows,
            "total_net": round(sum(r["net"] or 0 for r in rows), 2)}


@router.put("/salary/{period}", response_model=schemas.Msg)
async def save_salary(
    period: str,
    body: SalarySaveIn,
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    if len(period) != 7 or period[4] != "-":
        raise HTTPException(400, "period 格式应为 YYYY-MM")
    cur = {r.employee_id: r for r in (await db.execute(select(models.EmployeeSalaryMonthly).where(
        models.EmployeeSalaryMonthly.period == period))).scalars().all()}
    n = 0
    for r in body.rows:
        row = cur.get(r.employee_id)
        if row is None:
            db.add(models.EmployeeSalaryMonthly(
                employee_id=r.employee_id, period=period,
                base=r.base or 0, merit=r.merit or 0, overtime_pay=r.overtime_pay or 0,
                allowance=r.allowance or 0, social_deduct=r.social_deduct or 0,
                personal_tax=r.personal_tax or 0,
                other_deduct=r.other_deduct or 0, note=(r.note or "").strip() or None))
        else:
            row.base = r.base or 0
            row.merit = r.merit or 0
            row.overtime_pay = r.overtime_pay or 0
            row.allowance = r.allowance or 0
            row.social_deduct = r.social_deduct or 0
            row.personal_tax = r.personal_tax or 0
            row.other_deduct = r.other_deduct or 0
            row.note = (r.note or "").strip() or None
        n += 1
    await db.commit()
    await write_audit(db, user=current, action="hr_salary_save", target_type="salary",
                      detail=f"{period} {n} 人")
    return schemas.Msg(message=f"已保存 {period} 工资（{n} 人）")


# 🆕 #249 工资批量导入：表头即模板列（顺序=界面列顺序，人事对着填）
_SAL_IMPORT_COLS = ["工号*", "姓名", "基本工资", "绩效/奖金", "加班费", "补贴",
                    "社保公积金", "个税", "其他扣款", "备注"]


@router.get("/salary/import-template")
async def salary_import_template(
    _: models.User = Depends(require_roles(*_HR)),
):
    """🆕 #249 下载工资导入模板（.xlsx，含表头 + 示例行 + 填写说明）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "工资导入"
    ws.append(_SAL_IMPORT_COLS)
    hfill = PatternFill("solid", fgColor="FDF3E3")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.append(["00015", "张明宇（仅参考，导入按工号匹配）", 6000, 1500, 300, 200, 800, 120, 0, "示例行，导入前请删除"])
    for i, w in enumerate([10, 24, 12, 12, 10, 10, 12, 10, 12, 24], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("填写说明")
    for line in [
        ["工资导入说明"],
        [""],
        ["1. 按「工号」匹配员工——工号必填且必须已在花名册存在，姓名列仅供人工核对，不参与匹配。"],
        ["2. 导入哪个月，由页面上选的月份决定（不在表里填月份），导入前请先确认页面月份选对。"],
        ["3. 同一工号在该月已有工资 → 整行覆盖；没有 → 新建。表里没出现的人保持原样，不会被清空。"],
        ["4. 实发系统自动算 = 基本 + 绩效 + 加班费 + 补贴 − 社保公积金 − 个税 − 其他扣款，无需填。"],
        ["5. 金额列留空按 0 处理。第 2 行为示例，导入前请删除。"],
        ["6. 离职员工：离职当月仍可导入（末月工资要发得出来），之后的月份不在名单里。"],
    ]:
        ws2.append(line)
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 84
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = "工资导入模板.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.post("/salary/{period}/import")
async def import_salary(
    period: str,
    file: UploadFile = File(...),
    current: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #249 批量导入某月工资：按工号匹配，已有则整行覆盖，未出现的人不动。"""
    if len(period) != 7 or period[4] != "-":
        raise HTTPException(400, "period 格式应为 YYYY-MM")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "无法解析该文件，请使用模板导出的 .xlsx")
    rows = list(wb.active.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件为空或只有表头")
    header = [(str(c).strip().replace("*", "") if c is not None else "") for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header) if name}
    if "工号" not in col:
        raise HTTPException(400, "表头缺少必填列「工号」，请用模板")

    def cell(row, name):
        i = col.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    def num(v) -> float:
        if v is None or str(v).strip() == "":
            return 0.0
        try:
            return float(str(v).replace(",", "").replace("￥", "").replace("¥", "").strip())
        except ValueError:
            return 0.0

    emp_by_no = {(e.emp_no or "").strip(): e for e in await _active_emps(db, period) if e.emp_no}
    cur = {r.employee_id: r for r in (await db.execute(select(models.EmployeeSalaryMonthly).where(
        models.EmployeeSalaryMonthly.period == period))).scalars().all()}

    created = updated = 0
    skipped: list[str] = []
    for row in rows[1:]:
        raw_no = cell(row, "工号")
        if raw_no is None or str(raw_no).strip() == "":
            continue
        no = str(raw_no).strip()
        if no.isdigit() and no not in emp_by_no:
            no = no.zfill(5)
        e = emp_by_no.get(no)
        if not e:
            skipped.append(str(raw_no).strip())
            continue
        vals = dict(
            base=num(cell(row, "基本工资")), merit=num(cell(row, "绩效/奖金")),
            overtime_pay=num(cell(row, "加班费")), allowance=num(cell(row, "补贴")),
            social_deduct=num(cell(row, "社保公积金")), personal_tax=num(cell(row, "个税")),
            other_deduct=num(cell(row, "其他扣款")),
            note=(str(cell(row, "备注")).strip() if cell(row, "备注") not in (None, "") else None),
        )
        s = cur.get(e.id)
        if s is None:
            db.add(models.EmployeeSalaryMonthly(employee_id=e.id, period=period, **vals))
            created += 1
        else:
            for k, v in vals.items():
                setattr(s, k, v)
            updated += 1
    await db.commit()
    await write_audit(db, user=current, action="hr_salary_import", target_type="salary",
                      detail=f"{period} 新增{created} 覆盖{updated} 跳过{len(skipped)}")
    msg = f"导入完成：新增 {created} 人，覆盖 {updated} 人"
    if skipped:
        head = "、".join(skipped[:5]) + ("…" if len(skipped) > 5 else "")
        msg += f"；{len(skipped)} 个工号在该月名单里找不到，已跳过（{head}）"
    return {"message": msg, "created": created, "updated": updated, "skipped": skipped}


@router.get("/payroll-summary")
async def payroll_summary(
    year: int = Query(...),
    _: models.User = Depends(require_roles(*_HR)),
    db: AsyncSession = Depends(get_db),
):
    """年度逐月合计（人工成本走势;毛利榜人工分摊后续也读 payroll_monthly）。"""
    r = await db.execute(
        select(models.PayrollMonthly.month, func.sum(models.PayrollMonthly.total_amount))
        .where(models.PayrollMonthly.month.like(f"{year:04d}-%"))
        .group_by(models.PayrollMonthly.month))
    by = {m: (t or 0) for m, t in r.all()}
    rows = [{"month": f"{year:04d}-{m:02d}", "total": round(by.get(f"{year:04d}-{m:02d}", 0), 2)}
            for m in range(1, 13)]
    return {"year": year, "rows": rows, "total": round(sum(x["total"] for x in rows), 2)}
