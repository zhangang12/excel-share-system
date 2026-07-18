"""🆕 采购管理模块：供应商档案 / 采购明细 / 账目一览 / 请款流程 / 汇总报表 / 历史导入"""
from datetime import datetime, timezone, date as _date, datetime as _dt, timedelta as _td
from io import BytesIO
from typing import Optional, List, Dict
from collections import defaultdict
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, delete, or_
from sqlalchemy.ext.asyncio import AsyncSession

from ..database import get_db
from .. import models, schemas
from ..deps import require_roles, get_current_user
from ..utils import write_audit

router = APIRouter(prefix="/api/purchase-mgmt", tags=["采购管理"])

# 🆕 权限统一:tab 可见性由「二级菜单权限」决定,读接口按菜单内全部角色放行(admin/manager 自动)
_PURCHASE_ROLES = ("buyer", "buyer_lead", "finance", "finance_lead", "buyer_standard", "buyer_outsource")
# 🆕 采购员（标准件/外协）也可新增/编辑/删除采购明细与供应商
_WRITE_ROLES = ("buyer", "buyer_lead", "buyer_standard", "buyer_outsource")
# 🆕 采购收货入库：仓库填送货单号/到货日期/后填价格 → 自动入库。
# 收货归仓库（采购只下单、不收货）；admin/manager 由 require_roles 自动放行
_RECEIVE_ROLES = ("warehouse", "warehouse_lead")

# 🆕 R4/A6：采购员按清单分工（沿用采购部项目目录「按人分表」的可见性）。
# 仅限这三名采购员各管自己的清单；其他采购员 + 采购主管 + admin/manager 不受限（看全部）。
_BUYER_SHEET_MAP: dict[str, set[str]] = {
    "lixinxin": {"standard", "elec_po"},   # 李新新：标准件清单 + 电工采购单
    "wangqin": {"material", "laser"},       # 王芹：不锈钢原料下料单 + 激光件清单
    "fangbusen": {"outsource"},             # 方步森：外协加工
}


def _allowed_sheet_keys(user: Optional[models.User]) -> Optional[set[str]]:
    """该采购员可下单的清单集合；None = 不限制（看全部清单）。"""
    if not user:
        return None
    return _BUYER_SHEET_MAP.get((user.username or "").lower())


def _uname(u: Optional[models.User]) -> Optional[str]:
    return (u.full_name or u.username) if u else None


def _buyer_restricted(current: models.User) -> bool:
    """下单采购员(buyer 家族)只看自己下的单——判定依据只是"是不是采购下单角色"，
    与是否**兼任** finance / logistics / sales 无关（兼任财务不解除采购的行级隔离；
    这是反复出问题的坑：王芹=buyer/finance/logistics 曾因带 finance 而被判为看全部）。
    只有 采购主管 / admin / manager 看全部；纯 finance(不带 buyer)不受此限、可看全部对账。"""
    return current.has_role("buyer", "buyer_standard", "buyer_outsource") and not current.has_role(
        "buyer_lead", "admin", "manager"
    )


# ==================== 供应商 ====================

@router.get("/suppliers", response_model=List[schemas.SupplierOut])
async def list_suppliers(
    category: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    owned_only: bool = Query(False),
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(models.Supplier).order_by(models.Supplier.name)
    if category:
        stmt = stmt.where(models.Supplier.category == category)
    if status:
        stmt = stmt.where(models.Supplier.status == status)
    # 需求六(收紧)：采购员**只看自己新增的供应商**(created_by==本人)——别人建的、历史无归属(NULL)的
    #   都对采购员隐藏;管理层/采购主管看全部。下单选供应商 与 采购明细筛选下拉 口径一致(owned_only 不再区分)。
    #   注：老无归属供应商仅管理层可见;若采购员要用,需管理层把它归属给该采购员或该采购员重新新增。
    if _buyer_restricted(current):
        stmt = stmt.where(models.Supplier.created_by == current.id)
    if q:
        from sqlalchemy import or_
        stmt = stmt.where(
            or_(
                models.Supplier.name.ilike(f"%{q}%"),
                models.Supplier.code.ilike(f"%{q}%"),
                models.Supplier.contact.ilike(f"%{q}%"),
            )
        )
    r = await db.execute(stmt)
    sups = r.scalars().all()
    names = await _uid_name_map(db, [s.created_by for s in sups if s.created_by])
    return [_sup_out(s, names.get(s.created_by)) for s in sups]


async def _uid_name_map(db: AsyncSession, uids: list) -> dict:
    ids = {u for u in uids if u}
    if not ids:
        return {}
    r = await db.execute(select(models.User).where(models.User.id.in_(ids)))
    return {u.id: _uname(u) for u in r.scalars().all()}


def _sup_out(s: models.Supplier, creator_name: Optional[str] = None) -> schemas.SupplierOut:
    return schemas.SupplierOut(
        id=s.id, name=s.name, code=s.code, category=s.category,
        contact=s.contact, phone=s.phone, address=s.address,
        tax_no=s.tax_no, bank_name=s.bank_name, bank_account=s.bank_account,
        settlement_type=s.settlement_type, credit_days=s.credit_days,
        status=s.status, notes=s.notes,
        created_by=s.created_by, created_by_name=creator_name,
        created_at=s.created_at,
    )


@router.post("/suppliers", response_model=schemas.SupplierOut)
async def create_supplier(
    body: schemas.SupplierCreate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    s = models.Supplier(**body.model_dump(), created_by=current.id)  # 🆕 需求五：记录建档采购员
    db.add(s)
    await db.commit()
    await db.refresh(s)
    return _sup_out(s, _uname(current))


@router.put("/suppliers/{sid}", response_model=schemas.SupplierOut)
async def update_supplier(
    sid: int,
    body: schemas.SupplierUpdate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.Supplier).where(models.Supplier.id == sid))
    s = r.scalar_one_or_none()
    if not s:
        raise HTTPException(404, "供应商不存在")
    for k, v in body.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    await db.commit()
    await db.refresh(s)
    names = await _uid_name_map(db, [s.created_by])
    return _sup_out(s, names.get(s.created_by))


@router.put("/suppliers/{sid}/toggle")
async def toggle_supplier(
    sid: int,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.Supplier).where(models.Supplier.id == sid))
    s = r.scalar_one_or_none()
    if not s:
        raise HTTPException(404, "供应商不存在")
    s.status = "inactive" if s.status == "active" else "active"
    await db.commit()
    return {"status": s.status}


@router.delete("/suppliers/{sid}")
async def delete_supplier(
    sid: int,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """删除供应商：仅当该供应商没有任何采购明细/请款记录时可硬删除（保护账务历史）。
    有交易记录的供应商请改用「停用」。"""
    r = await db.execute(select(models.Supplier).where(models.Supplier.id == sid))
    s = r.scalar_one_or_none()
    if not s:
        raise HTTPException(404, "供应商不存在")
    cnt_items = await db.scalar(
        select(func.count()).select_from(models.PurchaseItem).where(
            models.PurchaseItem.supplier_id == sid))
    if cnt_items:
        raise HTTPException(400, f"该供应商已有 {cnt_items} 条采购明细，不能删除；如不再使用请「停用」。")
    cnt_pr = await db.scalar(
        select(func.count()).select_from(models.PaymentRequest).where(
            models.PaymentRequest.supplier_id == sid))
    if cnt_pr:
        raise HTTPException(400, f"该供应商已有 {cnt_pr} 条请款记录，不能删除；如不再使用请「停用」。")
    # 无交易记录：连带清理期初余额后硬删除
    await db.execute(delete(models.SupplierOpeningBalance).where(
        models.SupplierOpeningBalance.supplier_id == sid))
    await db.delete(s)
    await db.commit()
    return {"message": "供应商已删除"}


@router.post("/suppliers/{sid}/opening-balance", response_model=schemas.SupplierOpeningBalanceOut)
async def set_opening_balance(
    sid: int,
    body: schemas.SupplierOpeningBalanceIn,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(
        select(models.SupplierOpeningBalance).where(
            models.SupplierOpeningBalance.supplier_id == sid
        )
    )
    ob = r.scalar_one_or_none()
    if ob:
        ob.balance_date = body.balance_date
        ob.outstanding_amount = body.outstanding_amount
        ob.notes = body.notes
    else:
        ob = models.SupplierOpeningBalance(
            supplier_id=sid,
            balance_date=body.balance_date,
            outstanding_amount=body.outstanding_amount,
            notes=body.notes,
        )
        db.add(ob)
    await db.commit()
    await db.refresh(ob)
    return schemas.SupplierOpeningBalanceOut(
        id=ob.id, supplier_id=ob.supplier_id,
        balance_date=ob.balance_date,
        outstanding_amount=ob.outstanding_amount,
        notes=ob.notes,
    )


@router.get("/suppliers/{sid}/opening-balance", response_model=Optional[schemas.SupplierOpeningBalanceOut])
async def get_opening_balance(
    sid: int,
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    # 越权修复：受限采购员只能查看自己可见供应商(本人建 或 遗留共享)的期初欠款，
    # 不能凭 sid 直取他人供应商的期初金额（/suppliers 列表本按 created_by 过滤，这里补上同样边界）。
    if _buyer_restricted(current):
        sup_r = await db.execute(select(models.Supplier.created_by).where(models.Supplier.id == sid))
        creator = sup_r.scalar_one_or_none()
        if creator not in (None, current.id):
            raise HTTPException(403, "无权查看该供应商期初欠款")
    r = await db.execute(
        select(models.SupplierOpeningBalance).where(
            models.SupplierOpeningBalance.supplier_id == sid
        )
    )
    ob = r.scalar_one_or_none()
    if not ob:
        return None
    return schemas.SupplierOpeningBalanceOut(
        id=ob.id, supplier_id=ob.supplier_id,
        balance_date=ob.balance_date,
        outstanding_amount=ob.outstanding_amount,
        notes=ob.notes,
    )


# ==================== 采购明细 ====================

def _apply_sheet_type_filter(stmt, sheet_type: Optional[str]):
    """🆕 ④ 采购明细按清单类型过滤：source_sheet_id 的来源表名匹配该类型；'loose'=散单(无来源)。"""
    if not sheet_type:
        return stmt
    if sheet_type == "loose":
        return stmt.where(models.PurchaseItem.source_sheet_id.is_(None))
    conf = _PURCHASABLE_SHEETS.get(sheet_type)
    if not conf:
        return stmt
    sub = select(models.Datasheet.id).where(models.Datasheet.name == conf[0])
    return stmt.where(models.PurchaseItem.source_sheet_id.in_(sub))


# 注意：summary 路由须在 /{iid} 之前，避免 "summary" 被解析为 id 参数
@router.get("/items/summary", response_model=schemas.PurchaseItemSummary)
async def items_summary(
    supplier_id: Optional[int] = Query(None),
    project_code: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    invoice_status: Optional[str] = Query(None),
    sheet_type: Optional[str] = Query(None),
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(models.PurchaseItem)
    if _buyer_restricted(current):
        stmt = stmt.where(models.PurchaseItem.buyer_id == current.id)
    if supplier_id:
        stmt = stmt.where(models.PurchaseItem.supplier_id == supplier_id)
    if project_code:
        stmt = stmt.where(models.PurchaseItem.project_code.ilike(f"%{project_code}%"))
    if month:
        stmt = stmt.where(models.PurchaseItem.delivery_date.startswith(month))
    if invoice_status:
        # 用户要求「对账状态去掉已开票」：已开票并入已对账 → 筛选「已对账」也含存量「已开票」
        if invoice_status == "已对账":
            stmt = stmt.where(models.PurchaseItem.invoice_status.in_(["已对账", "已开票"]))
        else:
            stmt = stmt.where(models.PurchaseItem.invoice_status == invoice_status)
    stmt = _apply_sheet_type_filter(stmt, sheet_type)
    r = await db.execute(stmt)
    items = r.scalars().all()
    received = sum(i.received_amount or 0 for i in items)
    invoiced = sum(i.invoice_amount or 0 for i in items)
    paid = sum(i.paid_amount or 0 for i in items)
    return schemas.PurchaseItemSummary(
        received_total=received,
        uninvoiced=max(received - invoiced, 0),
        paid_total=paid,
        outstanding=max(received - paid, 0),
        count=len(items),
    )


@router.get("/items", response_model=List[schemas.PurchaseItemOut])
async def list_items(
    supplier_id: Optional[int] = Query(None),
    project_code: Optional[str] = Query(None),
    month: Optional[str] = Query(None),
    invoice_status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    sheet_type: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """采购明细（服务端分页：总条数用 /items/summary 的 count，随筛选联动）。"""
    stmt = select(models.PurchaseItem).order_by(
        models.PurchaseItem.delivery_date.desc(),
        models.PurchaseItem.id.desc(),
    )
    if _buyer_restricted(current):
        stmt = stmt.where(models.PurchaseItem.buyer_id == current.id)
    if supplier_id:
        stmt = stmt.where(models.PurchaseItem.supplier_id == supplier_id)
    if project_code:
        stmt = stmt.where(models.PurchaseItem.project_code.ilike(f"%{project_code}%"))
    if month:
        stmt = stmt.where(models.PurchaseItem.delivery_date.startswith(month))
    if invoice_status:
        # 用户要求「对账状态去掉已开票」：已开票并入已对账 → 筛选「已对账」也含存量「已开票」
        if invoice_status == "已对账":
            stmt = stmt.where(models.PurchaseItem.invoice_status.in_(["已对账", "已开票"]))
        else:
            stmt = stmt.where(models.PurchaseItem.invoice_status == invoice_status)
    stmt = _apply_sheet_type_filter(stmt, sheet_type)
    if category:
        # 🆕 按供应商分类筛选（明细本身无分类，取供应商分类）
        stmt = stmt.where(models.PurchaseItem.supplier_id.in_(
            select(models.Supplier.id).where(models.Supplier.category == category)))
    r = await db.execute(stmt.limit(page_size).offset((page - 1) * page_size))
    return await _attach_pay_status(db, [_item_out(i) for i in r.scalars().all()])


@router.get("/orders/{po_no}", response_model=List[schemas.PurchaseItemOut])
async def get_order(
    po_no: str,
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """按采购单号取整单明细（补打印采购单用）。"""
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.po_no == po_no)
                         .order_by(models.PurchaseItem.id))
    rows = list(r.scalars().all())
    # 🆕 #146 需求五：受限采购员只看自己建的采购单明细
    if _buyer_restricted(current):
        rows = [x for x in rows if x.buyer_id == current.id]
    if not rows:
        raise HTTPException(404, "采购单不存在")
    return await _attach_pay_status(db, [_item_out(x) for x in rows])


# ==================== 🆕 采购单 PDF（服务端直出，替代浏览器弹窗打印） ====================
_PO_COMPANY = "同辉智能装备（无锡）有限公司"
_PREPAY_METHODS = ("现金预付", "对公预付")


def _po_pay_label(method: Optional[str], ratio: Optional[float]) -> str:
    if not method:
        return ""
    if method in _PREPAY_METHODS and ratio is not None:
        return f"{method}（预付{ratio:g}%）"
    return method


def _render_po_pdf(po_no: str, supplier_name: str, rows: list) -> bytes:
    """用 reportlab 把整单渲染成 A4 PDF（内置 STSong-Light CID 字体，中文无需外部字体文件）。
    版式对齐原打印模板：抬头(需/供方/日期/付款方式/单号) + 明细表 + 合计 + 签字栏。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    FONT = "STSong-Light"
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("po_h1", parent=styles["Title"], fontName=FONT, fontSize=20,
                        alignment=TA_CENTER, spaceAfter=10)
    cell = ParagraphStyle("po_cell", parent=styles["Normal"], fontName=FONT, fontSize=8, leading=11)

    def P(t):
        return Paragraph(("" if t is None else str(t)).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), cell)

    first = rows[0]
    pay = _po_pay_label(first.payment_method, first.prepay_ratio)
    meta = [
        [P("需方"), P(_PO_COMPANY), P("下单日期"), P(first.delivery_date or "")],
        [P("供方"), P(supplier_name), P("付款方式"), P(pay)],
        [P("采购单号"), P(po_no), "", ""],
    ]
    meta_tbl = Table(meta, colWidths=[22 * mm, 60 * mm, 22 * mm, 60 * mm])
    meta_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("SPAN", (1, 2), (3, 2)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f2f2f2")),
        ("BACKGROUND", (2, 0), (2, 1), colors.HexColor("#f2f2f2")),
    ]))

    header = ["序号", "订单编号", "名称", "规格型号", "数量", "单价", "合计金额", "备注"]
    data = [[P(h) for h in header]]
    total = 0.0
    for i, x in enumerate(rows, 1):
        amt = x.received_amount if x.received_amount is not None else ((x.qty or 0) * (x.unit_price or 0))
        total += amt or 0
        # 🆕 成套：数量显示「N 套」，单价即套单价
        qty_txt = "" if x.qty is None else (f"{x.qty:g} 套" if getattr(x, "is_kit", False) else f"{x.qty:g}")
        data.append([
            P(i), P(x.project_code or ""), P(x.item_name or ""), P(x.spec or ""),
            P(qty_txt),
            P("" if x.unit_price is None else f"{x.unit_price:g}"),
            P("" if amt is None else f"{amt:,.2f}"), P(x.notes or ""),
        ])
    data.append([P("合计"), "", "", "", "", "", P(f"￥{total:,.2f}"), ""])
    col_w = [10 * mm, 24 * mm, 34 * mm, 30 * mm, 13 * mm, 18 * mm, 24 * mm, 25 * mm]
    items_tbl = Table(data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f5ee")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("SPAN", (0, -1), (5, -1)),
        ("ALIGN", (4, 1), (6, -1), "RIGHT"),
    ]))

    # 🆕 成套：把各成套行的「套内零件清单」附在明细表下方，供供应商核对
    kit_flow = []
    kb = ParagraphStyle("po_kit", parent=cell, fontName=FONT, fontSize=9, spaceBefore=4, spaceAfter=2)
    for x in rows:
        if not getattr(x, "is_kit", False):
            continue
        parts = x.kit_parts or []
        if not parts:
            continue
        kit_flow.append(Spacer(1, 5 * mm))
        kit_flow.append(Paragraph(f"套内零件清单（{x.item_name or ''}，每套）", kb))
        pdata = [[P("序号"), P("名称"), P("规格型号"), P("每套数量")]]
        for j, p in enumerate(parts, 1):
            pq = p.get("qty")
            pdata.append([P(j), P(p.get("name") or ""), P(p.get("spec") or ""),
                          P("" if pq in (None, "") else f"{pq:g}")])
        ptbl = Table(pdata, colWidths=[12 * mm, 60 * mm, 60 * mm, 24 * mm], repeatRows=1)
        ptbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#888888")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef3f8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (3, 1), (3, -1), "RIGHT"),
        ]))
        kit_flow.append(ptbl)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm, title=f"采购单 {po_no}")
    doc.build([
        Paragraph("采购单", h1), meta_tbl, Spacer(1, 6 * mm), items_tbl, *kit_flow,
        Spacer(1, 16 * mm),
        Paragraph("采购（签字）：____________　　　　　供方（盖章）：____________", cell),
    ])
    return buf.getvalue()


def _render_preq_pdf(pr: models.PurchaseRequest, requester_name: str, buyer_name: str) -> bytes:
    """🆕 反馈#232：把采购申请渲染成正式「采购申请单」A4 PDF(可查看/打印)。
    抬头(申请编号/申请人/申请日期/指定采购员/备注) + 物料明细表。"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    FONT = "STSong-Light"
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("pr_h1", parent=styles["Title"], fontName=FONT, fontSize=20,
                        alignment=TA_CENTER, spaceAfter=10)
    cell = ParagraphStyle("pr_cell", parent=styles["Normal"], fontName=FONT, fontSize=8, leading=11)

    def P(t):
        return Paragraph(("" if t is None else str(t)).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), cell)

    no = f"SQ{pr.id:05d}"
    created = pr.created_at.strftime("%Y-%m-%d") if pr.created_at else ""
    meta = [
        [P("申请编号"), P(no), P("申请日期"), P(created)],
        [P("申请人"), P(requester_name or ""), P("指定采购员"), P(buyer_name or "（未指定）")],
        [P("备注"), P(pr.notes or ""), "", ""],
    ]
    meta_tbl = Table(meta, colWidths=[22 * mm, 60 * mm, 24 * mm, 58 * mm])
    meta_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("SPAN", (1, 2), (3, 2)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f2f2f2")),
        ("BACKGROUND", (2, 0), (2, 1), colors.HexColor("#f2f2f2")),
    ]))

    header = ["序号", "名称", "规格型号", "数量", "项目编号", "备注"]
    data = [[P(h) for h in header]]
    for i, l in enumerate(pr.lines, 1):
        data.append([P(i), P(l.item_name or ""), P(l.spec or ""),
                     P("" if l.qty is None else f"{l.qty:g}"), P(l.project_code or ""), P(l.notes or "")])
    col_w = [12 * mm, 44 * mm, 40 * mm, 16 * mm, 26 * mm, 30 * mm]
    items_tbl = Table(data, colWidths=col_w, repeatRows=1)
    items_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8f5ee")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=16 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm, title=f"采购申请单 {no}")
    doc.build([
        Paragraph("采购申请单", h1), meta_tbl, Spacer(1, 6 * mm), items_tbl,
        Spacer(1, 16 * mm),
        Paragraph("申请（签字）：____________　　　　　采购（签字）：____________", cell),
    ])
    return buf.getvalue()


@router.get("/orders/{po_no}/pdf")
async def download_order_pdf(
    po_no: str,
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES, "warehouse", "warehouse_lead")),
    db: AsyncSession = Depends(get_db),
):
    """🆕 采购单直接下载 PDF：服务端渲染返回真正的 PDF 文件，替代原来靠浏览器
    弹窗 + window.print 的方式（移动端/微信浏览器经常无响应）。
    🆕 反馈#234/#235：仓库(采购收货)也可查看/打印采购单,故放开 warehouse 角色。"""
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.po_no == po_no)
                         .order_by(models.PurchaseItem.id))
    rows = list(r.scalars().all())
    # 🆕 #146 需求五：受限采购员只能下载自己建的采购单
    if _buyer_restricted(current):
        rows = [x for x in rows if x.buyer_id == current.id]
    if not rows:
        raise HTTPException(404, "采购单不存在")
    # 🆕 #146/#148 兜底：一单不该混供应商(存量迁移已拆),若仍有混合只渲染首供应商那部分,避免抬头串台
    sid0 = rows[0].supplier_id
    rows = [x for x in rows if x.supplier_id == sid0]
    sup = (await db.execute(select(models.Supplier).where(
        models.Supplier.id == sid0))).scalar_one_or_none()
    pdf_bytes = _render_po_pdf(po_no, sup.name if sup else "", rows)
    fname = quote(f"采购单_{po_no}.pdf")
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"})


async def _attach_pay_status(db: AsyncSession, outs: list) -> list:
    """按 已付金额 + 关联请款单状态 计算每条采购明细的付款状态（B1=a：记录付款才算已付）。
    优先级：已付款 > 部分付款 > 已批待付 > 已请款 > 未付款。"""
    if not outs:
        return outs
    ids = [o.id for o in outs]
    # 每条明细关联到的请款单状态集合
    r = await db.execute(
        select(models.PaymentRequestItem.item_id, models.PaymentRequest.status)
        .join(models.PaymentRequest, models.PaymentRequestItem.request_id == models.PaymentRequest.id)
        .where(models.PaymentRequestItem.item_id.in_(ids))
    )
    pr_by_item: dict[int, set] = {}
    for item_id, st in r.all():
        pr_by_item.setdefault(item_id, set()).add(st)
    for o in outs:
        prs = pr_by_item.get(o.id, set())
        recv = o.received_amount or 0
        paid = o.paid_amount or 0
        if recv > 0 and paid >= recv - 0.005:
            o.pay_status = "已付款"
        elif paid > 0:
            o.pay_status = "部分付款"
        elif "approved" in prs:
            o.pay_status = "已批待付"
        elif "pending" in prs:
            o.pay_status = "已请款"
        else:
            o.pay_status = "未付款"
    return outs


def _item_out(i: models.PurchaseItem) -> schemas.PurchaseItemOut:
    return schemas.PurchaseItemOut(
        id=i.id, po_no=i.po_no, supplier_id=i.supplier_id,
        supplier_name=i.supplier.name if i.supplier else "",
        delivery_date=i.delivery_date, contract_no=i.contract_no,
        project_code=i.project_code, delivery_note_no=i.delivery_note_no,
        arrival_date=i.arrival_date,
        item_name=i.item_name, spec=i.spec, brand=i.brand, qty=i.qty, unit_price=i.unit_price,
        received_amount=i.received_amount or 0,
        invoice_date=i.invoice_date, invoice_no=i.invoice_no, tax_rate=i.tax_rate,
        invoice_amount=i.invoice_amount or 0,
        paid_amount=i.paid_amount or 0, paid_date=i.paid_date,
        payment_method=i.payment_method, prepay_ratio=i.prepay_ratio,
        invoice_status=i.invoice_status,
        custom_values=i.custom_values or {},
        buyer_id=i.buyer_id,
        buyer_name=_uname(i.buyer),
        is_kit=bool(i.is_kit), kit_parts=i.kit_parts, is_stock=bool(i.is_stock),
        stock_location=i.stock_location,
        notes=i.notes, created_at=i.created_at,
    )


def _maybe_auto_reconcile(item: models.PurchaseItem) -> None:
    """🆕 对账状态自动化：现金全款/对公全款 + 已收货 + 已付清 → 自动置「已对账」，
    不需要人工手动切换；账期/现金预付/对公预付这几种（有账期天数/预付比例要核对）仍走人工流程。
    只从「待对账」推进一次，已经是「已对账/已开票」的不回退/不覆盖，避免打乱已经人工确认过的状态。"""
    if item.invoice_status != "待对账":
        return
    if item.payment_method not in ("现金全款", "对公全款"):
        return
    if not item.arrival_date:
        return
    recv = item.received_amount or 0
    if recv > 0 and (item.paid_amount or 0) >= recv - 0.005:
        item.invoice_status = "已对账"


@router.post("/items", response_model=schemas.PurchaseItemOut)
async def create_item(
    body: schemas.PurchaseItemCreate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    data = body.model_dump()
    if not data.get("received_amount") and data.get("qty") and data.get("unit_price"):
        data["received_amount"] = round((data["qty"] or 0) * (data["unit_price"] or 0), 4)
    data["custom_values"] = await _clean_custom(db, data.get("custom_values"))
    data["buyer_id"] = current.id
    item = models.PurchaseItem(**data)
    db.add(item)
    await db.commit()
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == item.id))
    return _item_out(r.scalar_one())


async def _next_po_no(db: AsyncSession) -> str:
    """采购单号：TH{yyyymmdd}-{当日序号3位}（TH=同辉）。同一天多张单顺序递增。

    🆕 #147/#148 修复：序号取「当日现有最大后缀 + 1」，而非 COUNT(DISTINCT)——
    否则删单留下空档后 COUNT 会算出一个已存在的号,新单并进老单里串供应商。MAX+1 永不重用已删号。"""
    from datetime import date as _date
    prefix = f"TH{_date.today().strftime('%Y%m%d')}-"
    r = await db.execute(
        select(models.PurchaseItem.po_no)
        .where(models.PurchaseItem.po_no.like(f"{prefix}%")).distinct()
    )
    mx = 0
    for (po,) in r.all():
        try:
            mx = max(mx, int(str(po).rsplit("-", 1)[-1]))
        except (ValueError, IndexError):
            pass
    return f"{prefix}{mx + 1:03d}"


@router.post("/orders", response_model=List[schemas.PurchaseItemOut])
async def create_purchase_order(
    body: schemas.PurchaseOrderCreate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 采购单：同一供应商一次录入多个零件行 → 生成一个采购单号(po_no)，批量建采购明细。

    - 表头共享：供应商 / 下单日期 / 合同编号 / 默认项目编号（行可覆盖项目编号）
    - 单价选填：后填价格流程留空即可（货到仓库再由仓库补送货单号/到货，采购/财务补价）
    """
    lines = [ln for ln in body.lines if (ln.item_name or "").strip()]
    if not lines:
        raise HTTPException(400, "请至少填写一行有效明细（名称必填）")
    po_no = await _next_po_no(db)
    for ln in lines:
        recv = ln.received_amount
        if not recv and ln.qty and ln.unit_price:
            recv = round((ln.qty or 0) * (ln.unit_price or 0), 4)
        cv = await _clean_custom(db, ln.custom_values)
        db.add(models.PurchaseItem(
            po_no=po_no,
            supplier_id=body.supplier_id,
            delivery_date=body.delivery_date,
            contract_no=body.contract_no,
            project_code=(ln.project_code or body.project_code),
            item_name=ln.item_name.strip(),
            spec=ln.spec, brand=ln.brand, qty=ln.qty, unit_price=ln.unit_price,
            received_amount=recv or 0,
            tax_rate=ln.tax_rate, notes=ln.notes,
            payment_method=body.payment_method, prepay_ratio=body.prepay_ratio,
            custom_values=cv,
            buyer_id=current.id,
            # is_stock 已废弃(收货一律只入库);兼容旧前端仍传值,不再落库为 False
            stock_location=(body.stock_location or "").strip() or None,   # 🆕 库位(整单一个)
        ))
    await db.commit()
    r = await db.execute(
        select(models.PurchaseItem).where(models.PurchaseItem.po_no == po_no)
        .order_by(models.PurchaseItem.id)
    )
    return [_item_out(x) for x in r.scalars().all()]


# ==================== 清单 → 采购下单 / 回写清单 ====================
_STD_SHEET = "标准件清单"

# 🆕 R1：可从清单下单的 5 张来源表 + 各表列名映射（列名各表不同）
# key: (表名, 名称列, 规格列, 数量列|None, 品牌列|None, 下单日期回写列, 到货日期回写列, 仓库签字回写列)
_PURCHASABLE_SHEETS = {
    "standard":  ("标准件清单",       "项目",     "规格型号", "数量", "品牌", "订购日期", "到货日期", "仓库签字"),
    "elec_po":   ("电工采购单",       "项目",     "规格型号", "数量", "品牌", "订购日期", "到货日期", "仓库签字"),
    "material":  ("不锈钢原料下料单", "材料类别", "规格型号", "数量", None,   "下单日期", "到料日期", "仓库"),
    "outsource": ("外协加工",         "名称",     "图纸名称", None,   None,   "发出日期", "到货日期", "仓库"),
    "laser":     ("激光件清单",       "名称",     "图纸名称", None,   None,   "发出日期", "到料日期", "仓库"),
}
# 回写清单时把所有可能的「下单日期/到货日期/仓库签字」列名都写一遍，只有该表存在的列会生效
_ALL_ORDER_DATE_COLS = ("订购日期", "下单日期", "发出日期")
_ALL_ARRIVAL_COLS = ("到货日期", "到料日期")
_ALL_WH_SIGN_COLS = ("仓库签字", "仓库")
_ALL_WH_LOC_COLS = ("库位",)   # 🆕 #250 收货时把库位也回写清单（列不存在的表自动跳过）


async def _sheet_fieldmap(db: AsyncSession, sheet_id: int) -> dict:
    fr = await db.execute(select(models.Field).where(models.Field.datasheet_id == sheet_id))
    return {f.name: str(f.id) for f in fr.scalars().all()}


async def _writeback_sheet_row(db: AsyncSession, sheet_id: int, record_id: int, updates: dict,
                               *, only_if_empty: Optional[set] = None) -> None:
    """把 {列名: 值} 写进某数据表某行 values（按列名匹配 field_id）；列不存在则跳过。
    only_if_empty 里的列名仅在当前单元格为空时才写（不覆盖已手填值，#255 采购负责人用）。"""
    if not sheet_id or not record_id:
        return
    only_if_empty = only_if_empty or set()
    name2id = await _sheet_fieldmap(db, sheet_id)
    rr = await db.execute(select(models.Record).where(models.Record.id == record_id))
    rec = rr.scalar_one_or_none()
    if not rec:
        return
    vals = dict(rec.values or {})
    hit = False
    for col, val in updates.items():
        fid = name2id.get(col)
        if not fid or val is None:
            continue
        if col in only_if_empty:
            cur = vals.get(fid)
            cur_empty = cur in (None, "") or (isinstance(cur, list) and not cur)
            if not cur_empty:
                continue   # 已手填，保留不覆盖
        vals[fid] = val
        hit = True
    if hit:
        rec.values = vals


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# ==================== R6：采购自定义字段 ====================
_FIELD_ADMIN_ROLES = ("buyer_lead",)   # 配置字段：采购主管（admin/manager 由 require_roles 自动放行）


async def _custom_fields(db: AsyncSession, enabled_only: bool = False):
    q = select(models.PurchaseCustomField).order_by(
        models.PurchaseCustomField.sort_order, models.PurchaseCustomField.id)
    if enabled_only:
        q = q.where(models.PurchaseCustomField.enabled == True)  # noqa: E712
    r = await db.execute(q)
    return list(r.scalars().all())


async def _clean_custom(db: AsyncSession, custom_values: Optional[dict]) -> dict:
    """校验必填、净化自定义字段值（只保留启用字段、去空）。"""
    fields = await _custom_fields(db, enabled_only=True)
    cv = custom_values or {}
    clean: dict = {}
    missing: list[str] = []
    for f in fields:
        key = str(f.id)
        val = cv.get(key)
        sval = "" if val is None else str(val).strip()
        if f.required and not sval:
            missing.append(f.label)
        elif sval:
            clean[key] = val
    if missing:
        raise HTTPException(400, f"必填自定义字段未填写：{'、'.join(missing)}")
    return clean


@router.get("/custom-fields", response_model=List[schemas.PurchaseCustomFieldOut])
async def list_custom_fields(
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """采购自定义字段列表（所有采购/财务角色可读，用于渲染列与输入框）。"""
    return [schemas.PurchaseCustomFieldOut.model_validate(f) for f in await _custom_fields(db)]


@router.post("/custom-fields", response_model=schemas.PurchaseCustomFieldOut)
async def create_custom_field(
    body: schemas.PurchaseCustomFieldIn,
    current: models.User = Depends(require_roles(*_FIELD_ADMIN_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    f = models.PurchaseCustomField(**body.model_dump())
    db.add(f)
    await db.commit()
    await db.refresh(f)
    await write_audit(db, user=current, action="create", target_type="purchase_custom_field", target_id=f.id)
    return schemas.PurchaseCustomFieldOut.model_validate(f)


@router.put("/custom-fields/{fid}", response_model=schemas.PurchaseCustomFieldOut)
async def update_custom_field(
    fid: int,
    body: schemas.PurchaseCustomFieldIn,
    current: models.User = Depends(require_roles(*_FIELD_ADMIN_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.PurchaseCustomField).where(models.PurchaseCustomField.id == fid))
    f = r.scalar_one_or_none()
    if not f:
        raise HTTPException(404, "字段不存在")
    for k, v in body.model_dump().items():
        setattr(f, k, v)
    await db.commit()
    await db.refresh(f)
    return schemas.PurchaseCustomFieldOut.model_validate(f)


@router.delete("/custom-fields/{fid}", response_model=schemas.Msg)
async def delete_custom_field(
    fid: int,
    current: models.User = Depends(require_roles(*_FIELD_ADMIN_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """删除字段定义（已录入明细里的历史值保留在 custom_values 中，只是不再展示/校验）。"""
    r = await db.execute(select(models.PurchaseCustomField).where(models.PurchaseCustomField.id == fid))
    f = r.scalar_one_or_none()
    if not f:
        raise HTTPException(404, "字段不存在")
    await db.delete(f)
    await db.commit()
    await write_audit(db, user=current, action="delete", target_type="purchase_custom_field", target_id=fid)
    return schemas.Msg(message="已删除该自定义字段")


@router.get("/purchasable/{project_id}", response_model=List[schemas.PurchasableRow])
async def purchasable(
    project_id: int,
    sheet: str = Query("standard", description="清单类型: standard/elec_po/material/outsource/laser"),
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """取项目某来源清单的可采购行 + 采购状态 + 现有库存/建议采购量。
    5 张来源表列名各异（外协/激光无数量列），按 _PURCHASABLE_SHEETS 映射取数。"""
    conf = _PURCHASABLE_SHEETS.get(sheet)
    if not conf:
        raise HTTPException(400, "未知清单类型")
    # 🆕 R4/A6：采购员只能取自己负责清单的可采购行（管理员/主管/其他采购员不限）
    allowed = _allowed_sheet_keys(current)
    if allowed is not None and sheet not in allowed:
        raise HTTPException(403, "你没有该清单的采购权限")
    r = await db.execute(select(models.Datasheet).where(
        models.Datasheet.project_id == project_id, models.Datasheet.name == conf[0]))
    ds = r.scalar_one_or_none()
    if not ds:
        return []
    stock_by_key = await _build_stock_by_key(db)
    return await _purchasable_rows(db, ds, conf, sheet, stock_by_key)


def _stock_key(nm, sp):
    return ((nm or "").strip(), (sp or "").strip())


async def _build_stock_by_key(db: AsyncSession) -> dict:
    """按 名称+规格 汇总实时库存（供 purchasable 类端点复用；跨项目聚合时只算一次）。"""
    from .warehouse_router import _stock_map
    mr = await db.execute(select(models.WhMaterial))
    mats = list(mr.scalars().all())
    stock_by_id = await _stock_map(db) if mats else {}
    stock_by_key: dict = {}
    for m in mats:
        k = _stock_key(m.name, m.spec)
        stock_by_key[k] = stock_by_key.get(k, 0) + stock_by_id.get(m.id, m.init_stock or 0)
    return stock_by_key


async def _purchasable_rows(db: AsyncSession, ds: models.Datasheet, conf: tuple, sheet_key: str,
                            stock_by_key: dict, *, only_pending: bool = False,
                            project_id: Optional[int] = None, project_code: Optional[str] = None,
                            project_name: Optional[str] = None,
                            name2id: Optional[dict] = None, by_rec: Optional[dict] = None,
                            records: Optional[list] = None) -> list:
    """某数据表的可采购行（归一化 5 类清单异构列 → 统一 PurchasableRow）。供单项目/跨项目端点复用。
    name2id/by_rec/records 由跨项目聚合端点批量预取传入，避免逐张清单 N+1 查询。"""
    item_col, spec_col, qty_col, brand_col = conf[1:5]
    if name2id is None:
        name2id = await _sheet_fieldmap(db, ds.id)
    if by_rec is None:
        lr = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.source_sheet_id == ds.id))
        by_rec = {}
        for pi in lr.scalars().all():
            by_rec.setdefault(pi.source_record_id, []).append(pi)
    if records is None:
        rr = await db.execute(select(models.Record).where(
            models.Record.datasheet_id == ds.id).order_by(models.Record.sort_order, models.Record.id))
        records = list(rr.scalars().all())
    out = []
    for rec in records:
        v = rec.values or {}

        def gv(col):
            if not col:
                return None
            fid = name2id.get(col)
            x = v.get(fid) if fid else None
            if isinstance(x, list):
                x = "、".join(str(i) for i in x)
            return str(x).strip() if x not in (None, "") else None

        name = gv(item_col)
        if not name:
            continue
        pis = by_rec.get(rec.id, [])
        if not pis:
            status = "未下单"
        elif all(p.arrival_date for p in pis):
            status = "已到货"
        else:
            status = "已下单"
        if only_pending and status != "未下单":
            continue
        spec = gv(spec_col)
        qty = _num(gv(qty_col)) if qty_col else None
        stock = round(stock_by_key.get(_stock_key(name, spec), 0), 4)
        suggest = round(max(0.0, (qty or 0) - stock), 4) if qty is not None else 0
        out.append(schemas.PurchasableRow(
            sheet_id=ds.id, record_id=rec.id, item_name=name, spec=spec, brand=gv(brand_col),
            material=gv("材质"), drawing=gv("图纸名称"), qty=qty, stock=stock, suggest_purchase=suggest,
            notes=gv("备注"), status=status,
            sheet_key=sheet_key, project_id=project_id,
            project_code=project_code, project_name=project_name))
    return out



@router.post("/orders/from-list", response_model=List[schemas.PurchaseItemOut])
async def create_order_from_list(
    body: schemas.OrderFromListCreate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """从清单选定行 + 选供应商 → 生成一张采购单；回写清单的订购日期/采购负责人。"""
    lines = [l for l in body.lines if (l.item_name or "").strip()]
    if not lines:
        raise HTTPException(400, "请至少选择一行")
    # 🆕 R4/A6：采购员只能对自己负责清单下单（管理员/主管/其他采购员不限）
    allowed = _allowed_sheet_keys(current)
    if allowed is not None:
        sheet_ids = {l.source_sheet_id for l in lines if l.source_sheet_id}
        if sheet_ids:
            name2key = {conf[0]: k for k, conf in _PURCHASABLE_SHEETS.items()}
            dsr = await db.execute(select(models.Datasheet).where(models.Datasheet.id.in_(sheet_ids)))
            for ds in dsr.scalars().all():
                key = name2key.get(ds.name)
                if key and key not in allowed:
                    raise HTTPException(403, f"你没有「{ds.name}」的采购权限")
    po_no = await _next_po_no(db)
    today = _date.today().isoformat()
    uname = current.full_name or current.username
    # 🆕 #253：未显式带项目号时，从来源清单回溯项目编号落库（否则仓库收货看不到订单编号）
    sheet_code_map: dict = {}
    if not (body.project_code or "").strip():
        sids = {l.source_sheet_id for l in lines if l.source_sheet_id}
        if sids:
            dsr = await db.execute(select(models.Datasheet.id, models.Datasheet.project_id)
                                   .where(models.Datasheet.id.in_(sids)))
            pmap = {sid: pid for sid, pid in dsr.all() if pid}
            if pmap:
                pr = await db.execute(select(models.Project.id, models.Project.code)
                                      .where(models.Project.id.in_(set(pmap.values()))))
                cbp = {pid: code for pid, code in pr.all()}
                sheet_code_map = {sid: cbp.get(pid) for sid, pid in pmap.items()}
    for l in lines:
        recv = None
        if l.qty and l.unit_price:
            recv = round((l.qty or 0) * (l.unit_price or 0), 4)
        db.add(models.PurchaseItem(
            po_no=po_no, source_sheet_id=l.source_sheet_id, source_record_id=l.source_record_id,
            supplier_id=body.supplier_id, delivery_date=body.delivery_date,
            project_code=(body.project_code or sheet_code_map.get(l.source_sheet_id)),
            item_name=l.item_name.strip(), spec=l.spec, brand=l.brand, qty=l.qty, unit_price=l.unit_price,
            received_amount=recv or 0, payment_method=(l.payment_method or body.payment_method),
            prepay_ratio=(l.prepay_ratio if l.prepay_ratio is not None else body.prepay_ratio),
            notes=l.notes, buyer_id=current.id,   # is_stock 已废弃:收货一律只入库(默认True)
            stock_location=(body.stock_location or "").strip() or None))   # 🆕 库位(整单一个)
        if l.source_sheet_id and l.source_record_id:
            # 各来源表的「下单日期」列名不同（订购/下单/发出日期），全写一遍只有存在的列生效
            wb = {"采购负责人": uname}
            for c in _ALL_ORDER_DATE_COLS:
                wb[c] = (body.delivery_date or today)
            # 🆕 #255：采购负责人若已在表格手填，则保留、不被下单人覆盖；下单日期照常回写
            await _writeback_sheet_row(db, l.source_sheet_id, l.source_record_id, wb,
                                       only_if_empty={"采购负责人"})
    await db.commit()
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.po_no == po_no)
                         .order_by(models.PurchaseItem.id))
    return [_item_out(x) for x in r.scalars().all()]


@router.post("/orders/kit-from-list", response_model=schemas.PurchaseItemOut)
async def create_kit_order_from_list(
    body: schemas.KitFromListCreate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 按套下单（从清单下单的扩展版）：从清单勾选一组零件(同一供应商)打包成「一套」，
    建**一条**成套采购明细(is_kit)，并回写这些清单行的订购日期/采购负责人(与从清单下单完全一致)。
    - item_name=套名, qty=套数, unit_price=套单价(=套总价/套数), received_amount=套总价
    - kit_parts=套内零件清单(勾选行的 名称/规格/每套数量, 描述性)；整套后续按一个总走收货/入库/开票/请款/付款。
    """
    parts_in = [p for p in body.parts if (p.name or "").strip()]
    if not parts_in:
        raise HTTPException(400, "请至少勾选一行清单零件")
    # 🆕 R4/A6：采购员只能对自己负责清单下单（镜像从清单下单）
    allowed = _allowed_sheet_keys(current)
    if allowed is not None and body.source_sheet_id:
        name2key = {conf[0]: k for k, conf in _PURCHASABLE_SHEETS.items()}
        ds = (await db.execute(select(models.Datasheet).where(
            models.Datasheet.id == body.source_sheet_id))).scalar_one_or_none()
        key = name2key.get(ds.name) if ds else None
        if key and key not in allowed:
            raise HTTPException(403, f"你没有「{ds.name}」的采购权限")
    qty = body.kit_qty
    unit_price = round(body.kit_total / qty, 4) if qty else 0
    kit_parts = [{"name": p.name.strip(), "spec": p.spec, "qty": p.qty} for p in parts_in]
    po_no = await _next_po_no(db)
    item = models.PurchaseItem(
        po_no=po_no, source_sheet_id=body.source_sheet_id,
        supplier_id=body.supplier_id, delivery_date=body.delivery_date,
        project_code=body.project_code,
        item_name=body.kit_name.strip(), spec=f"成套（{len(kit_parts)}项零件）",
        qty=qty, unit_price=unit_price, received_amount=round(body.kit_total, 2),
        payment_method=body.payment_method, prepay_ratio=body.prepay_ratio,
        is_kit=True, kit_parts=kit_parts, notes=body.notes, buyer_id=current.id,
        # is_stock 已废弃:收货一律只入库(默认True)
        stock_location=(body.stock_location or "").strip() or None,
    )
    db.add(item)
    # 回写各来源清单行（与从清单下单一致：采购负责人 + 下单日期）
    today = _date.today().isoformat()
    uname = current.full_name or current.username
    if body.source_sheet_id:
        for p in parts_in:
            if p.source_record_id:
                wb = {"采购负责人": uname}
                for c in _ALL_ORDER_DATE_COLS:
                    wb[c] = (body.delivery_date or today)
                # 🆕 #255：采购负责人若已手填则保留、不被下单人覆盖
                await _writeback_sheet_row(db, body.source_sheet_id, p.source_record_id, wb,
                                           only_if_empty={"采购负责人"})
    await db.commit()
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == item.id))
    return _item_out(r.scalar_one())


async def _auto_stock_in(db: AsyncSession, item: models.PurchaseItem, current: models.User) -> None:
    """采购收货 → 自动入库（带采购单价/金额）；非备货(is_stock=False)再自动一笔「采购领用」出库
    (直发对应项目，净库存过账为0)；备货(is_stock=True)只入库、留库存。幂等：同一采购明细只过账一次。"""
    if not item.qty or item.qty <= 0:
        return
    ex = await db.execute(select(models.WhTxn).where(
        models.WhTxn.purchase_item_id == item.id, models.WhTxn.is_reversal == False))  # noqa: E712
    if ex.scalars().first():
        return
    spec = (item.spec or "").strip() or None
    mq = select(models.WhMaterial).where(models.WhMaterial.name == item.item_name)
    mq = mq.where(models.WhMaterial.spec == spec) if spec else mq.where(models.WhMaterial.spec.is_(None))
    m = (await db.execute(mq)).scalar_one_or_none()
    loc = (item.stock_location or "").strip() or None
    if not m:
        m = models.WhMaterial(name=item.item_name, spec=spec, unit="个", category="采购入库",
                              location=loc)   # 🆕 采购下单选的库位
        db.add(m)
        await db.flush()
    elif loc:
        m.location = loc   # 🆕 库位管理：最近一次收货放到哪,物料当前库位就是哪(单库位模型)
    pid = None
    if item.project_code:
        pr = await db.execute(select(models.Project.id).where(models.Project.code == item.project_code))
        pid = pr.scalar_one_or_none()
    bd = item.arrival_date or _date.today().isoformat()
    from .warehouse_router import _next_ref
    ref = await _next_ref(db, "in", bd)
    up = item.unit_price
    amt = round((item.qty or 0) * up, 4) if up is not None else None
    db.add(models.WhTxn(
        material_id=m.id, biz_date=bd, direction="in", qty=item.qty,
        unit_price=up, amount=amt, purchase_item_id=item.id,
        source="采购入库", party=(item.supplier.name if item.supplier else None),
        project_id=pid, location=(loc or m.location), ref_no=ref, operator_id=current.id))
    # 🆕 库位管理批次：**取消收货自动出库**（原 is_stock=False 自动生成「采购领用」出库已删）——
    #   收货一律只入库到所选库位;出库统一走仓库领料(出入库登记/物料需求一键领用),挂项目计成本。


# ==================== 采购历史数据 一键导入 ====================
# 模板列（顺序即模板列顺序；* 为必填）。解析时按表头名匹配，允许调整列序。
_IMPORT_COLS = [
    "供应商名称*", "采购单号", "下单日期", "订单编号", "名称*", "规格型号",
    "数量", "单价", "合计金额", "送货单号", "到货日期", "付款方式",
    "开票日期", "开票金额", "税率", "对账状态", "备注",
]


def _norm_date(v) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, (_dt, _date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().replace(".", "-").replace("/", "-")
    return s[:10] if s else None


def _norm_num(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("￥", "").replace("¥", "").strip())
    except (ValueError, TypeError):
        return None


def _norm_str(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


@router.get("/items/import-template")
async def import_template(
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
):
    """下载采购明细导入模板（.xlsx，含表头 + 示例行 + 填写说明）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "采购明细导入"
    ws.append(_IMPORT_COLS)
    hfill = PatternFill("solid", fgColor="E8F5EE")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.append(["无锡示例金属制品有限公司", "CG20260101-001", "2026-01-05", "2026-046",
               "轴承座", "SKF-6205", 4, 35, 140, "SF1234", "2026-01-08", "转账",
               "", "", "13%", "待对账", "示例行，导入前请删除"])
    widths = [26, 16, 12, 12, 18, 16, 8, 10, 12, 14, 12, 10, 12, 12, 8, 10, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("填写说明")
    for line in [
        ["采购历史数据导入说明"],
        [""],
        ["1. 带 * 的列为必填：供应商名称、名称。"],
        ["2. 供应商名称若系统中不存在，会自动新建（仅带名称，其余资料后续在供应商档案补全）。"],
        ["3. 日期格式 YYYY-MM-DD（如 2026-01-05），也兼容 2026.1.5 / 2026/1/5。"],
        ["4. 单价、合计金额只填数字；合计金额留空且有数量与单价时，系统自动=数量×单价。"],
        ["5. 付款方式：现金 / 转账 / 月结 / 承兑 / 预付（也可自定义）。"],
        ["6. 对账状态：待对账 / 已对账 / 已开票（留空默认待对账）。"],
        ["7. 第 2 行为示例，导入前请删除。可一次导入多个供应商、多条明细。"],
    ]:
        ws2.append(line)
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 80

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = "采购明细导入模板.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.post("/items/import", response_model=schemas.PurchaseImportResult)
async def import_items(
    file: UploadFile = File(...),
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """一键导入采购历史明细：按表头名匹配列；供应商按名称匹配，不存在则自动新建。"""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "无法解析该文件，请使用模板导出的 .xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件为空或只有表头")
    header = [(_norm_str(c) or "").replace("*", "").strip() for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header) if name}

    def cell(row, name):
        i = col.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    if "供应商名称" not in col or "名称" not in col:
        raise HTTPException(400, "表头缺少必填列「供应商名称」或「名称」，请用模板")

    # 供应商名称 -> id 缓存（一次性载入现有）
    sres = await db.execute(select(models.Supplier.id, models.Supplier.name))
    sup_map = {n: i for i, n in sres.all()}

    created = 0
    suppliers_created = 0
    errors: list[str] = []
    for rn, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        sup_name = _norm_str(cell(row, "供应商名称"))
        item_name = _norm_str(cell(row, "名称"))
        if not sup_name or not item_name:
            errors.append(f"第 {rn} 行：供应商名称/名称为空，已跳过")
            continue
        sid = sup_map.get(sup_name)
        if sid is None:
            s = models.Supplier(name=sup_name, status="active")
            db.add(s)
            await db.flush()
            sid = s.id
            sup_map[sup_name] = sid
            suppliers_created += 1
        qty = _norm_num(cell(row, "数量"))
        unit_price = _norm_num(cell(row, "单价"))
        recv = _norm_num(cell(row, "合计金额"))
        if recv is None and qty is not None and unit_price is not None:
            recv = round(qty * unit_price, 4)
        db.add(models.PurchaseItem(
            po_no=_norm_str(cell(row, "采购单号")),
            supplier_id=sid,
            delivery_date=_norm_date(cell(row, "下单日期")),
            project_code=_norm_str(cell(row, "订单编号")),
            item_name=item_name,
            spec=_norm_str(cell(row, "规格型号")),
            qty=qty, unit_price=unit_price, received_amount=recv or 0,
            delivery_note_no=_norm_str(cell(row, "送货单号")),
            arrival_date=_norm_date(cell(row, "到货日期")),
            payment_method=_norm_str(cell(row, "付款方式")),
            invoice_date=_norm_date(cell(row, "开票日期")),
            invoice_amount=_norm_num(cell(row, "开票金额")) or 0,
            tax_rate=_norm_str(cell(row, "税率")),
            invoice_status=_norm_str(cell(row, "对账状态")) or "待对账",
            notes=_norm_str(cell(row, "备注")),
            buyer_id=current.id,
        ))
        created += 1
    await db.commit()
    return schemas.PurchaseImportResult(
        created=created, suppliers_created=suppliers_created,
        failed=len(errors), errors=errors[:50],
    )


# ==================== 🆕 #191 供应商档案 一键导入 ====================
_SUP_IMPORT_COLS = [
    "供应商名称*", "编码", "分类", "结算方式", "账期天数", "联系人",
    "电话", "地址", "税号", "开户行", "银行账号", "备注",
]


@router.get("/suppliers/import-template")
async def supplier_import_template(
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
):
    """下载供应商导入模板（.xlsx，含表头 + 示例行 + 填写说明）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "供应商导入"
    ws.append(_SUP_IMPORT_COLS)
    hfill = PatternFill("solid", fgColor="E8F0FE")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = hfill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.append(["无锡示例金属制品有限公司", "GYS-001", "标准件", "月结", 30, "张三",
               "13800000000", "无锡市XX路1号", "91320200XXXXXXXXXX", "工商银行无锡支行",
               "1234567890123456789", "示例行，导入前请删除"])
    widths = [28, 12, 12, 10, 10, 10, 14, 26, 22, 22, 24, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    ws2 = wb.create_sheet("填写说明")
    for line in [
        ["供应商档案导入说明"],
        [""],
        ["1. 带 * 的列为必填：供应商名称。其余列选填，后续可在供应商档案里补。"],
        ["2. 名称已存在的供应商：只补空缺字段（不会覆盖已维护的资料）；不存在则新建。"],
        ["3. 分类建议与「字典设置-供应商分类」里的取值一致（不一致也能导入，下拉里可能选不到）。"],
        ["4. 结算方式：现金 / 月结 / 无账期。账期天数只填数字（月结供应商务必填，资金面板按它算应付到期）。"],
        ["5. 第 2 行为示例，导入前请删除。可一次导入多家。"],
    ]:
        ws2.append(line)
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.column_dimensions["A"].width = 80
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = "供应商导入模板.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


@router.post("/suppliers/import")
async def import_suppliers(
    file: UploadFile = File(...),
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #191 一键导入供应商：按名称匹配——已存在只补空缺字段，不存在新建(建档人=导入人)。"""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(BytesIO(await file.read()), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(400, "无法解析该文件，请使用模板导出的 .xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "文件为空或只有表头")
    header = [(_norm_str(c) or "").replace("*", "").strip() for c in rows[0]]
    col = {name: idx for idx, name in enumerate(header) if name}
    if "供应商名称" not in col:
        raise HTTPException(400, "表头缺少必填列「供应商名称」，请用模板")

    def cell(row, name):
        i = col.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    sres = await db.execute(select(models.Supplier))
    sup_by_name = {x.name: x for x in sres.scalars().all()}
    field_map = [("编码", "code"), ("分类", "category"), ("结算方式", "settlement_type"),
                 ("联系人", "contact"), ("电话", "phone"), ("地址", "address"),
                 ("税号", "tax_no"), ("开户行", "bank_name"), ("银行账号", "bank_account"),
                 ("备注", "notes")]
    created = updated = 0
    errors: list[str] = []
    for rn, row in enumerate(rows[1:], start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        name = _norm_str(cell(row, "供应商名称"))
        if not name:
            errors.append(f"第 {rn} 行：供应商名称为空，已跳过")
            continue
        credit = _norm_num(cell(row, "账期天数"))
        sup = sup_by_name.get(name)
        if sup is None:
            sup = models.Supplier(name=name, status="active", created_by=current.id)
            db.add(sup)
            sup_by_name[name] = sup
            created += 1
            touched = True
        else:
            touched = False
        # 已存在只补空缺，不覆盖已维护资料
        for cn, attr in field_map:
            v = _norm_str(cell(row, cn))
            if v and not getattr(sup, attr, None):
                setattr(sup, attr, v)
                touched = True
        if credit is not None and sup.credit_days is None:
            sup.credit_days = int(credit)
            touched = True
        if touched and sup.id:
            updated += 1
    await db.commit()
    msg = f"导入完成：新建 {created} 家、补全 {updated} 家"
    if errors:
        msg += f"；{len(errors)} 行跳过"
    return {"message": msg, "created": created, "updated": updated, "errors": errors[:20]}


@router.put("/items/{iid}", response_model=schemas.PurchaseItemOut)
async def update_item(
    iid: int,
    body: schemas.PurchaseItemUpdate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == iid))
    item = r.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "明细不存在")
    if _buyer_restricted(current) and item.buyer_id != current.id:
        raise HTTPException(403, "无权编辑他人明细")
    data = body.model_dump(exclude_unset=True)
    if "custom_values" in data:
        data["custom_values"] = await _clean_custom(db, data.get("custom_values"))
    for k, v in data.items():
        setattr(item, k, v)
    if ("qty" in data or "unit_price" in data) and "received_amount" not in data:
        if item.qty and item.unit_price:
            item.received_amount = round(item.qty * item.unit_price, 4)
    # 🆕 盈利改善1b·修真bug：先收货后补价 → 回写已生成的出入库流水。此前补价只更新
    #   PurchaseItem，收货时生成的 amount=NULL 的 wh_txn 永久无价，库存金额与项目材料成本双双偏低。
    if ("qty" in data or "unit_price" in data or "received_amount" in data) and item.unit_price is not None:
        tr = await db.execute(select(models.WhTxn).where(
            models.WhTxn.purchase_item_id == item.id,
            models.WhTxn.is_reversal == False))  # noqa: E712
        for t in tr.scalars().all():
            t.unit_price = item.unit_price
            t.amount = round((t.qty or 0) * item.unit_price, 4)
    _maybe_auto_reconcile(item)
    await db.commit()
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == iid))
    return _item_out(r.scalar_one())


@router.delete("/items/{iid}")
async def delete_item(
    iid: int,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """删除只删「采购明细」这一条记录本身，不会级联删除/冲正仓库入库或财务请款——
    如果链路已经走到那两边，直接物理删除会让仓库库存虚高、财务的钱对不上明细，所以直接拦截，
    要求先去对应模块走正常的撤销流程（仓库冲红 / 请款单处理），保证两边流水都有据可查。"""
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == iid))
    item = r.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "明细不存在")
    if _buyer_restricted(current) and item.buyer_id != current.id:
        raise HTTPException(403, "无权删除他人明细")
    if item.invoice_status in ("已对账", "已开票"):
        raise HTTPException(400, "已对账/已开票的明细不可删除")
    wh = await db.execute(select(models.WhTxn.id).where(
        models.WhTxn.purchase_item_id == iid, models.WhTxn.is_reversal == False,  # noqa: E712
        models.WhTxn.reversed == False))  # noqa: E712
    if wh.scalar_one_or_none():
        raise HTTPException(400, "该明细已收货入库，仓库库存已增加；请先到【仓库-出入库流水】把对应入库单冲红，再回来删除")
    pri = await db.execute(select(func.count(models.PaymentRequestItem.id)).where(
        models.PaymentRequestItem.item_id == iid))
    if pri.scalar():
        raise HTTPException(400, "该明细已被请款（含已批/已付），不可直接删除；请先到【请款记录】里处理对应请款单")
    await db.delete(item)
    await db.commit()
    return {"ok": True}


# ==================== 采购收货（仓库）====================
@router.get("/receiving", response_model=List[schemas.PurchaseItemOut])
async def list_receiving(
    supplier_id: Optional[int] = Query(None),
    po_no: Optional[str] = Query(None),
    received: bool = Query(False, description="False=待收货(默认) / True=已收货"),
    current: models.User = Depends(require_roles(*_RECEIVE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """采购收货清单：默认列出「待收货」明细（到货日期为空）供仓库确认收货、补送货单号/后填价格。"""
    stmt = select(models.PurchaseItem).order_by(
        models.PurchaseItem.po_no.desc().nullslast(),
        models.PurchaseItem.id.desc(),
    )
    if received:
        stmt = stmt.where(models.PurchaseItem.arrival_date.isnot(None))
    else:
        stmt = stmt.where(models.PurchaseItem.arrival_date.is_(None))
    if supplier_id:
        stmt = stmt.where(models.PurchaseItem.supplier_id == supplier_id)
    if po_no:
        stmt = stmt.where(models.PurchaseItem.po_no.ilike(f"%{po_no}%"))
    r = await db.execute(stmt.limit(300))
    items = list(r.scalars().all())
    outs = [_item_out(i) for i in items]
    # 🆕 #253：订单编号(project_code)为空、但来自项目清单的明细，回溯来源表所属项目编号补显示
    #   （外协/加工件下单时未带项目号，仓库收货看不到属于哪个项目/加工订单）
    need_sheets = {i.source_sheet_id for i in items
                   if i.source_sheet_id and not (i.project_code or "").strip()}
    if need_sheets:
        dsr = await db.execute(select(models.Datasheet.id, models.Datasheet.project_id)
                               .where(models.Datasheet.id.in_(need_sheets)))
        proj_by_sheet = {sid: pid for sid, pid in dsr.all() if pid}
        pids = set(proj_by_sheet.values())
        code_by_pid = {}
        if pids:
            pr = await db.execute(select(models.Project.id, models.Project.code)
                                  .where(models.Project.id.in_(pids)))
            code_by_pid = {pid: code for pid, code in pr.all()}
        code_by_item = {}
        for i in items:
            if i.source_sheet_id and not (i.project_code or "").strip():
                code = code_by_pid.get(proj_by_sheet.get(i.source_sheet_id))
                if code:
                    code_by_item[i.id] = code
        for o in outs:
            if not (o.project_code or "").strip() and o.id in code_by_item:
                o.project_code = code_by_item[o.id]
    # 🆕 需求十四：附上各明细的收货单数量
    if outs:
        rc = await db.execute(
            select(models.Attachment.biz_id, func.count(models.Attachment.id))
            .where(models.Attachment.biz_type == "receipt_doc",
                   models.Attachment.biz_id.in_([o.id for o in outs]))
            .group_by(models.Attachment.biz_id))
        cnt = {bid: c for bid, c in rc.all()}
        for o in outs:
            o.receipt_count = cnt.get(o.id, 0)
    return outs


async def _finish_receive(db: AsyncSession, item: models.PurchaseItem,
                          arrival_date: str, current: models.User) -> None:
    """收货共同尾部：回写清单(到货日期/进度/仓库签字) → 自动入库 → 自动对账。"""
    if item.source_sheet_id and item.source_record_id:
        wb = {"进度": "已到货"}
        for c in _ALL_ARRIVAL_COLS:
            wb[c] = arrival_date
        for c in _ALL_WH_SIGN_COLS:
            wb[c] = (current.full_name or current.username)
        # 🆕 #250 库位回写：收货填的库位写回清单「库位」列（此前只回写日期/签字，库位一直空）
        if item.stock_location:
            for c in _ALL_WH_LOC_COLS:
                wb[c] = item.stock_location
        await _writeback_sheet_row(db, item.source_sheet_id, item.source_record_id, wb)
    await _auto_stock_in(db, item, current)
    _maybe_auto_reconcile(item)


@router.put("/items/{iid}/receive", response_model=schemas.PurchaseItemOut)
async def receive_item(
    iid: int,
    body: schemas.PurchaseReceiveIn,
    current: models.User = Depends(require_roles(*_RECEIVE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """仓库收货：填送货单号 + 到货日期（后填价格流程一并补单价/收货金额）。
    到货日期一填即视为已收货。"""
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == iid))
    item = r.scalar_one_or_none()
    if not item:
        raise HTTPException(404, "明细不存在")
    if not body.arrival_date:
        raise HTTPException(400, "请填写到货日期")
    item.delivery_note_no = body.delivery_note_no
    item.arrival_date = body.arrival_date
    # 🆕 #204 库位改由仓库收货时填（取代采购下单填）；填了才写，不覆盖已有为空
    _loc = (body.stock_location or "").strip() or None
    if _loc:
        item.stock_location = _loc
    # 🆕 #253 订单编号：手工采购单没带项目号的，仓库收货时补/改（放在入库前，入库流水也能挂上项目）
    if body.project_code is not None:
        item.project_code = (body.project_code or "").strip() or None
    if body.unit_price is not None:
        item.unit_price = body.unit_price
    if body.received_amount is not None:
        item.received_amount = body.received_amount
    elif item.qty and item.unit_price:
        item.received_amount = round(item.qty * item.unit_price, 4)
    await _finish_receive(db, item, body.arrival_date, current)
    await db.commit()
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id == iid))
    return _item_out(r.scalar_one())


@router.post("/items/receive-batch", response_model=List[schemas.PurchaseItemOut])
async def receive_batch(
    body: schemas.BatchReceiveIn,
    current: models.User = Depends(require_roles(*_RECEIVE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 需求四：合并零件收货——一次收多条明细。两种填价方式：
    - 合并总价(total_amount)：按各行数量权重把总价分摊到 received_amount，单价=金额÷数量；
    - 逐行(lines)：各行分别填 单价/收货金额。
    公共：填送货单号 + 到货日期 → 自动入库 + 回写清单。"""
    if not body.arrival_date:
        raise HTTPException(400, "请填写到货日期")
    if not body.item_ids:
        raise HTTPException(400, "请选择要收货的明细")
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id.in_(body.item_ids)))
    got = {i.id: i for i in r.scalars().all()}
    ordered = [got[i] for i in body.item_ids if i in got]
    if not ordered:
        raise HTTPException(404, "明细不存在")
    if body.total_amount is not None:
        # 合并总价按数量权重分摊，末行兜余数保证合计精确
        weights = [(it.qty or 0) for it in ordered]
        wsum = sum(weights)
        n = len(ordered)
        allocated = 0.0
        for idx, it in enumerate(ordered):
            if idx == n - 1:
                share = round(body.total_amount - allocated, 2)
            else:
                share = (round(body.total_amount * (weights[idx] / wsum), 2)
                         if wsum > 0 else round(body.total_amount / n, 2))
                allocated += share
            it.received_amount = share
            it.unit_price = round(share / it.qty, 4) if it.qty else None
    else:
        line_map = {ln.item_id: ln for ln in body.lines}
        for it in ordered:
            ln = line_map.get(it.id)
            if not ln:
                continue
            if ln.unit_price is not None:
                it.unit_price = ln.unit_price
            if ln.received_amount is not None:
                it.received_amount = ln.received_amount
            elif it.qty and it.unit_price:
                it.received_amount = round(it.qty * it.unit_price, 4)
    _loc = (body.stock_location or "").strip() or None   # 🆕 #204 整批一个库位,仓库收货时填
    _pcode = (body.project_code or "").strip() or None    # 🆕 #253 整批一个订单编号
    for it in ordered:
        it.delivery_note_no = body.delivery_note_no
        it.arrival_date = body.arrival_date
        if _loc:
            it.stock_location = _loc
        if body.project_code is not None:
            it.project_code = _pcode
        await _finish_receive(db, it, body.arrival_date, current)
    await db.commit()
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id.in_(body.item_ids)))
    return [_item_out(i) for i in r.scalars().all()]


# ==================== 🆕 需求十四：采购收货单（图片）上传 ====================
@router.post("/items/{iid}/receipt", response_model=schemas.AttachmentOut)
async def upload_receipt(
    iid: int,
    file: UploadFile = File(...),
    current: models.User = Depends(require_roles(*_RECEIVE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """仓库收货时上传收货单（图片/PDF），按采购明细归档。"""
    item = (await db.execute(select(models.PurchaseItem).where(
        models.PurchaseItem.id == iid))).scalar_one_or_none()
    if not item:
        raise HTTPException(404, "明细不存在")
    from .attachments_router import save_upload
    att = await save_upload(db, file, biz_type="receipt_doc", biz_id=iid, user=current)
    await db.commit()
    await db.refresh(att)
    return schemas.AttachmentOut.model_validate(att)


@router.get("/items/{iid}/receipts", response_model=List[schemas.AttachmentOut])
async def list_receipts(
    iid: int,
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES, *_RECEIVE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    # 越权修复：受限采购员只能看自己明细的收货单附件（仓库角色不受限，需正常收货）
    if _buyer_restricted(current):
        it = (await db.execute(select(models.PurchaseItem.buyer_id).where(
            models.PurchaseItem.id == iid))).scalar_one_or_none()
        if it != current.id:
            raise HTTPException(403, "无权查看该明细的收货单")
    r = await db.execute(select(models.Attachment).where(
        models.Attachment.biz_type == "receipt_doc",
        models.Attachment.biz_id == iid).order_by(models.Attachment.id.desc()))
    return [schemas.AttachmentOut.model_validate(a) for a in r.scalars().all()]


@router.post("/items/batch-invoice")
async def batch_invoice(
    body: schemas.BatchInvoiceIn,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(
        select(models.PurchaseItem).where(models.PurchaseItem.id.in_(body.item_ids))
    )
    items = r.scalars().all()
    # 越权修复：受限采购员只能对本人明细批量开票（与 /items/set-invoice-no 一致，防篡改他人开票状态/金额）
    if _buyer_restricted(current):
        items = [i for i in items if i.buyer_id == current.id]
    # 🆕 #100 对账后合计开票：合计开票金额按各明细「收货金额」比例分摊到 invoice_amount，
    #     收货额全为 0 时均摊；末条兜余数，保证分摊合计精确等于填入的合计开票金额。
    amt = body.invoice_amount
    total_received = sum(float(i.received_amount or 0) for i in items)
    n = len(items)
    allocated = 0.0
    for idx, item in enumerate(items):
        item.invoice_status = "已开票"
        if body.invoice_date:
            item.invoice_date = body.invoice_date
        if amt is not None:
            if idx == n - 1:
                item.invoice_amount = round(amt - allocated, 2)
            else:
                share = round(amt * float(item.received_amount or 0) / total_received, 2) if total_received > 0 else round(amt / n, 2)
                item.invoice_amount = share
                allocated += share
    await db.commit()
    return {"updated": len(items)}


@router.post("/items/set-group-summary")
async def set_group_summary(
    body: schemas.GroupSummaryIn,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #4 合并父行整单维护(不分摊)：开票金额/已付款作为整单总额记在**首行**(其余对应字段置0，
    保证合并父行/账目的汇总合计=所填总额)；对账状态套用到**所有**子行。空字段不改。仅本人明细(受限时)。"""
    if not body.item_ids:
        raise HTTPException(400, "请选择合并单的明细")
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id.in_(body.item_ids)))
    got = {i.id: i for i in r.scalars().all()}
    items = [got[i] for i in body.item_ids if i in got]
    if _buyer_restricted(current):
        items = [i for i in items if i.buyer_id == current.id]
    if not items:
        raise HTTPException(404, "明细不存在或无权限")
    if len({i.supplier_id for i in items}) > 1:
        raise HTTPException(400, "跨供应商不能一起维护")
    for idx, it in enumerate(items):
        if body.invoice_amount is not None:
            it.invoice_amount = round(body.invoice_amount, 2) if idx == 0 else 0
        if body.paid_amount is not None:
            it.paid_amount = round(body.paid_amount, 2) if idx == 0 else 0
            it.paid_date = body.paid_date if idx == 0 else it.paid_date
        if body.invoice_status:
            it.invoice_status = body.invoice_status
    await db.commit()
    await write_audit(db, user=current, action="set_group_summary", target_type="purchase_item",
                      target_id=None, detail=f"整单维护 {len(items)} 条")
    return {"updated": len(items)}


@router.post("/items/set-invoice-no")
async def set_invoice_no(
    body: schemas.SetInvoiceNoIn,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 需求十三：对多个零件统一维护同一开票号；每行开票金额=该行收货金额，置「已开票」。"""
    ino = body.invoice_no.strip()
    if not ino:
        raise HTTPException(400, "请填写开票号")
    if not body.item_ids:
        raise HTTPException(400, "请至少选择一条采购明细")
    r = await db.execute(select(models.PurchaseItem).where(models.PurchaseItem.id.in_(body.item_ids)))
    items = r.scalars().all()
    if _buyer_restricted(current):
        items = [i for i in items if i.buyer_id == current.id]
    if not items:
        raise HTTPException(404, "明细不存在或无权限")
    # #170：一个开票号=一张发票=一个供应商，禁止跨供应商批量盖同号（前端也拦，这里兜底）
    if len({i.supplier_id for i in items}) > 1:
        raise HTTPException(400, "跨供应商不能一起维护开票号（一个开票号=一张发票=一个供应商）")
    # #2：未收货(收货金额<=0)的零件不能参与合并开票——必须先全部收货
    unrecv = [i for i in items if not (i.received_amount and i.received_amount > 0)]
    if unrecv:
        raise HTTPException(400, f"有 {len(unrecv)} 项尚未收货（收货金额为0），请先全部收货后再合并开票")
    # #2：合并开票金额(发票总额)必须与Σ勾选零件收货金额一致（≤1分误差）才放行
    recv_total = round(sum(i.received_amount or 0 for i in items), 2)
    if body.invoice_amount is not None and abs(round(body.invoice_amount, 2) - recv_total) > 0.01:
        raise HTTPException(400, f"合并开票金额 {round(body.invoice_amount, 2)} 与勾选零件收货金额合计 {recv_total} 不一致，无法开票")
    # 开票日期：合并开票必然有开票日期，未传则默认今天——必须同步到每个零件(修:原来选填留空则不同步)
    inv_date = (body.invoice_date or "").strip() or _date.today().isoformat()
    for item in items:
        item.invoice_no = ino
        item.invoice_amount = item.received_amount or 0   # 每个零件开票金额=各自收货金额
        item.invoice_status = "已开票"
        item.invoice_date = inv_date                      # 总是同步开票日期到每个零件
    await db.commit()
    await write_audit(db, user=current, action="set_invoice_no", target_type="purchase_item",
                      target_id=None, detail=f"开票号 {ino} → {len(items)} 条明细")
    return {"updated": len(items), "invoice_no": ino}


# ==================== 供应商账目一览 ====================

@router.get("/statements", response_model=schemas.SupplierStatementList)
async def supplier_statements(
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.Supplier).order_by(models.Supplier.name))
    all_suppliers = r.scalars().all()

    ob_r = await db.execute(select(models.SupplierOpeningBalance))
    ob_map = {ob.supplier_id: ob.outstanding_amount for ob in ob_r.scalars().all()}

    item_r = await db.execute(select(models.PurchaseItem))
    items = item_r.scalars().all()

    # 越权修复：受限采购员的每行金额与合计只能按本人明细汇总——否则同一(共享)供应商上
    # 其他采购员的收货/开票/已付金额会经该行汇总泄漏(原代码只用 my_sids 决定"哪几行出现"，
    # 加总却用了全量 items，导致列表合计 > 下钻明细合计，自相矛盾)。
    restricted = _buyer_restricted(current)
    # 需求六：受限采购员账目**只列本人新增的供应商**(created_by==本人),别人建/无归属的都不显示;
    #   每行金额仍只按本人明细(buyer_id==本人)汇总,不泄漏他人金额。
    my_sids: Optional[set] = None
    if restricted:
        my_sids = {s.id for s in all_suppliers if s.created_by == current.id}
    grp: dict = defaultdict(lambda: {"received": 0.0, "invoice": 0.0, "paid": 0.0, "count": 0})
    for i in items:
        if restricted and i.buyer_id != current.id:
            continue
        g = grp[i.supplier_id]
        g["received"] += i.received_amount or 0
        g["invoice"] += i.invoice_amount or 0
        g["paid"] += i.paid_amount or 0
        g["count"] += 1

    rows = []
    total_opening = total_received = total_paid = total_outstanding = 0.0
    for s in all_suppliers:
        if my_sids is not None and s.id not in my_sids:
            continue
        g = grp.get(s.id, {"received": 0.0, "invoice": 0.0, "paid": 0.0, "count": 0})
        ob = ob_map.get(s.id, 0.0)
        outstanding = ob + g["received"] - g["paid"]
        uninvoiced = g["received"] - g["invoice"]
        rows.append(schemas.SupplierStatementRow(
            supplier_id=s.id, supplier_name=s.name, category=s.category,
            opening_balance=ob, received_total=g["received"],
            invoice_total=g["invoice"], paid_total=g["paid"],
            outstanding=outstanding, uninvoiced=uninvoiced,
            item_count=g["count"],
        ))
        total_opening += ob
        total_received += g["received"]
        total_paid += g["paid"]
        total_outstanding += outstanding

    return schemas.SupplierStatementList(
        rows=rows, total_opening=total_opening,
        total_received=total_received, total_paid=total_paid,
        total_outstanding=total_outstanding,
    )


@router.get("/statements/{sid}/detail", response_model=List[schemas.PurchaseItemOut])
async def supplier_statement_detail(
    sid: int,
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(models.PurchaseItem).where(
        models.PurchaseItem.supplier_id == sid
    ).order_by(models.PurchaseItem.delivery_date.desc(), models.PurchaseItem.id.desc())
    if _buyer_restricted(current):
        stmt = stmt.where(models.PurchaseItem.buyer_id == current.id)
    r = await db.execute(stmt.limit(500))
    return await _attach_pay_status(db, [_item_out(i) for i in r.scalars().all()])


@router.get("/statements/{sid}/export")
async def export_supplier_statement(
    sid: int,
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 C6：导出某供应商采购对账单（.xlsx）：抬头 + 汇总 + 明细 + 合计。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sr = await db.execute(select(models.Supplier).where(models.Supplier.id == sid))
    sup = sr.scalar_one_or_none()
    if not sup:
        raise HTTPException(404, "供应商不存在")
    # 越权修复：受限采购员导出对账单只能包含本人明细，且不得导出自己从未经手的供应商
    # （原代码零过滤，任意 sid 都能拉到含其他采购员单据的整表 Excel）。与 /statements/{sid}/detail 一致。
    istmt = select(models.PurchaseItem).where(
        models.PurchaseItem.supplier_id == sid).order_by(
        models.PurchaseItem.delivery_date.asc().nullsfirst(), models.PurchaseItem.id.asc())
    if _buyer_restricted(current):
        istmt = istmt.where(models.PurchaseItem.buyer_id == current.id)
    ir = await db.execute(istmt)
    rows0 = ir.scalars().all()
    if _buyer_restricted(current) and not rows0:
        raise HTTPException(403, "无权导出该供应商对账单（您没有该供应商的采购明细）")
    items = await _attach_pay_status(db, [_item_out(i) for i in rows0])
    obr = await db.execute(select(models.SupplierOpeningBalance).where(
        models.SupplierOpeningBalance.supplier_id == sid))
    ob = obr.scalar_one_or_none()
    opening = (ob.outstanding_amount if ob else 0) or 0

    recv = sum(i.received_amount or 0 for i in items)
    inv = sum(i.invoice_amount or 0 for i in items)
    paid = sum(i.paid_amount or 0 for i in items)
    uninv = sum((i.received_amount or 0) for i in items if i.invoice_status != "已开票")
    outstanding = opening + recv - paid

    wb = Workbook()
    ws = wb.active
    ws.title = "采购对账单"
    thin = Side(style="thin", color="C0C4CC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = ["下单日期", "采购单号", "项目编号", "名称", "规格型号", "品牌", "数量", "单价",
            "收货金额", "开票日期", "开票金额", "已付款", "付款状态", "对账状态", "备注"]
    n = len(cols)

    def _cell(row, col, val, bold=False, fill=None, align="left", money=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        if money and isinstance(val, (int, float)):
            c.number_format = "#,##0.00"
        return c

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    t = _cell(1, 1, f"{sup.name}  采购对账单", bold=True, align="center")
    t.font = Font(bold=True, size=15)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    _cell(2, 1, f"分类：{sup.category or '—'}    联系人：{sup.contact or '—'}    电话：{sup.phone or '—'}",
          align="center")
    # 汇总条
    summ = (f"期初欠款：{opening:,.2f}    收货合计：{recv:,.2f}    开票合计：{inv:,.2f}    "
            f"待开票：{uninv:,.2f}    已付款：{paid:,.2f}    欠款余额：{outstanding:,.2f}")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n)
    _cell(3, 1, summ, bold=True, fill="FFF7E6", align="center")

    hrow = 5
    for j, name in enumerate(cols, start=1):
        c = _cell(hrow, j, name, bold=True, fill="E8F5EE", align="center")
        c.border = border
    r0 = hrow + 1
    for k, it in enumerate(items):
        vals = [it.delivery_date or "", it.po_no or "", it.project_code or "", it.item_name,
                it.spec or "", it.brand or "", it.qty, it.unit_price, it.received_amount or 0,
                it.invoice_date or "", it.invoice_amount or 0, it.paid_amount or 0,
                it.pay_status or "未付款", it.invoice_status, it.notes or ""]
        for j, v in enumerate(vals, start=1):
            money = j in (8, 9, 11, 12)
            align = "right" if j in (7, 8, 9, 11, 12) else "center" if j in (1, 2, 3, 10, 13, 14) else "left"
            c = _cell(r0 + k, j, v, align=align, money=money)
            c.border = border
    # 合计行
    tot_row = r0 + len(items)
    _cell(tot_row, 1, "合计", bold=True, fill="F5F7FA", align="center").border = border
    for j in range(2, n + 1):
        c = _cell(tot_row, j, "", fill="F5F7FA")
        c.border = border
    _cell(tot_row, 9, recv, bold=True, fill="F5F7FA", align="right", money=True).border = border
    _cell(tot_row, 11, inv, bold=True, fill="F5F7FA", align="right", money=True).border = border
    _cell(tot_row, 12, paid, bold=True, fill="F5F7FA", align="right", money=True).border = border

    widths = [12, 15, 12, 20, 18, 10, 8, 10, 13, 12, 13, 12, 10, 10, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=hrow, column=i).column_letter].width = w
    ws.freeze_panes = f"A{r0}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"{sup.name}_采购对账单.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ==================== 请款流程 ====================

async def _pr_out(db: AsyncSession, pr_id: int) -> schemas.PaymentRequestOut:
    r = await db.execute(select(models.PaymentRequest).where(models.PaymentRequest.id == pr_id))
    pr = r.scalar_one()
    ri = await db.execute(
        select(models.PaymentRequestItem, models.PurchaseItem)
        .join(models.PurchaseItem, models.PaymentRequestItem.item_id == models.PurchaseItem.id)
        .where(models.PaymentRequestItem.request_id == pr_id)
    )
    pairs = ri.all()
    item_rows = [
        {
            "item_id": pri.item_id,
            "item_name": pi.item_name,
            "po_no": pi.po_no,                       # 🆕 需求十六：付款时可见采购单
            "spec": pi.spec,
            "project_code": pi.project_code,
            "received_amount": pi.received_amount or 0,
            "allocated_amount": pri.allocated_amount,
        }
        for pri, pi in pairs
    ]
    # 🆕 盈利改善2·应付账期：最早到期日 = min(明细到货日) + 供应商账期天数——
    #   财务按到期日排程付款：有账期不必提前付(白放弃免息资金)，逾期未付有断供风险。
    earliest_due = None
    due_in_days = None
    arrivals = [pi.arrival_date for _pri, pi in pairs if pi.arrival_date]
    if arrivals and pr.supplier and pr.supplier.credit_days is not None:
        try:
            _d0 = _date.fromisoformat(min(arrivals)) + _td(days=pr.supplier.credit_days)
            earliest_due = _d0.isoformat()
            due_in_days = (_d0 - _date.today()).days
        except ValueError:
            pass
    po_nos = sorted({r["po_no"] for r in item_rows if r["po_no"]})
    voucher_name = None
    if pr.pay_voucher_file_id:
        ar = await db.execute(select(models.Attachment.name).where(
            models.Attachment.id == pr.pay_voucher_file_id))
        voucher_name = ar.scalar_one_or_none()
    sup = pr.supplier
    return schemas.PaymentRequestOut(
        id=pr.id, supplier_id=pr.supplier_id,
        supplier_name=sup.name if sup else "",
        requested_amount=pr.requested_amount,
        requester_id=pr.requester_id,
        requester_name=_uname(pr.requester),
        status=pr.status, notes=pr.notes,
        finance_approver_id=pr.finance_approver_id,
        approver_name=_uname(pr.finance_approver),
        approved_at=pr.approved_at,
        paid_amount=pr.paid_amount, paid_date=pr.paid_date,
        payment_method=pr.payment_method,
        pay_voucher_file_id=pr.pay_voucher_file_id,
        pay_voucher_name=voucher_name,
        reject_reason=pr.reject_reason,
        # 🆕 需求十六：付款时可见收款账户信息 + 关联采购单
        supplier_bank_name=(sup.bank_name if sup else None),
        supplier_bank_account=(sup.bank_account if sup else None),
        supplier_tax_no=(sup.tax_no if sup else None),
        po_nos=po_nos,
        earliest_due=earliest_due, due_in_days=due_in_days,
        created_at=pr.created_at,
        items=item_rows,
    )


@router.post("/payment-requests", response_model=schemas.PaymentRequestOut)
async def create_payment_request(
    body: schemas.PaymentRequestCreate,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    # 越权修复：受限采购员只能对本人下单的采购明细发起请款，防止引用他人明细
    if _buyer_restricted(current) and body.items:
        ids = [it.item_id for it in body.items]
        owned = set((await db.execute(
            select(models.PurchaseItem.id).where(
                models.PurchaseItem.id.in_(ids),
                models.PurchaseItem.buyer_id == current.id,
            ))).scalars().all())
        if set(ids) - owned:
            raise HTTPException(403, "请款单只能包含本人下单的采购明细")
    # #177：防重复请款——同一采购明细若已有未完成(待审批/已批待付)的请款单，禁止再次请款
    if body.items:
        ids2 = [it.item_id for it in body.items]
        dup_r = await db.execute(
            select(models.PurchaseItem.item_name)
            .join(models.PaymentRequestItem, models.PaymentRequestItem.item_id == models.PurchaseItem.id)
            .join(models.PaymentRequest, models.PaymentRequestItem.request_id == models.PaymentRequest.id)
            .where(models.PaymentRequestItem.item_id.in_(ids2),
                   models.PaymentRequest.status.in_(("pending", "approved")))
        )
        dups = list(dict.fromkeys(dup_r.scalars().all()))
        if dups:
            raise HTTPException(400, f"以下明细已有未完成的请款单，请勿重复请款：{'、'.join(dups)[:120]}")
    pr = models.PaymentRequest(
        supplier_id=body.supplier_id,
        requested_amount=body.requested_amount,
        requester_id=current.id,
        notes=body.notes,
        status="pending",
    )
    db.add(pr)
    await db.flush()
    for item_in in body.items:
        db.add(models.PaymentRequestItem(
            request_id=pr.id,
            item_id=item_in.item_id,
            allocated_amount=item_in.allocated_amount,
        ))
    await db.commit()
    # notify finance users
    try:
        fin_r = await db.execute(
            select(models.User).join(models.Role, models.User.role_id == models.Role.id)
            .where(models.Role.code == "finance", models.User.is_active == True)  # noqa: E712
        )
        for u in fin_r.scalars().all():
            db.add(models.Message(
                to_user_id=u.id, kind="info",
                text=f"采购请款：{_uname(current)} 发起请款 ¥{body.requested_amount:.2f}，请审批",
                biz_type="payment_request", biz_id=pr.id,
            ))
        await db.commit()
    except Exception:
        pass
    return await _pr_out(db, pr.id)


@router.get("/payment-requests", response_model=List[schemas.PaymentRequestOut])
async def list_payment_requests(
    status: Optional[str] = Query(None),
    supplier_id: Optional[int] = Query(None),
    current: models.User = Depends(require_roles(*_PURCHASE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(models.PaymentRequest).order_by(models.PaymentRequest.created_at.desc())
    if status:
        stmt = stmt.where(models.PaymentRequest.status == status)
    if supplier_id:
        stmt = stmt.where(models.PaymentRequest.supplier_id == supplier_id)
    if _buyer_restricted(current):
        # 需求四：请款记录按「下单采购员」隔离——请款单本身无 buyer_id，
        # 经 PaymentRequestItem → PurchaseItem.buyer_id 反查，关联明细里有本人下的单即可见。
        mine = (
            select(models.PaymentRequestItem.request_id)
            .join(
                models.PurchaseItem,
                models.PaymentRequestItem.item_id == models.PurchaseItem.id,
            )
            .where(models.PurchaseItem.buyer_id == current.id)
        )
        stmt = stmt.where(models.PaymentRequest.id.in_(mine))
    r = await db.execute(stmt)
    return [await _pr_out(db, pr.id) for pr in r.scalars().all()]


@router.put("/payment-requests/{prid}/approve")
async def approve_payment_request(
    prid: int,
    current: models.User = Depends(require_roles("finance")),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.PaymentRequest).where(models.PaymentRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "请款单不存在")
    if pr.status != "pending":
        raise HTTPException(400, "只有待审状态的请款单可审批")
    # 🆕 反馈#237 内控：不能审批自己提交的请款单（兼任采购+财务的账号最容易踩，如采购员兼开票）。
    #   与下方付款端点的「审批人不能给自己审过的单付款」同属职责分离，一并对管理层生效（不留后门）。
    if pr.requester_id and pr.requester_id == current.id:
        raise HTTPException(400, "职责分离：不能审批自己提交的请款单，请由另一位财务审批")
    pr.status = "approved"
    pr.finance_approver_id = current.id
    pr.approved_at = datetime.now(timezone.utc)
    await db.commit()
    return {"ok": True}


@router.put("/payment-requests/{prid}/reject")
async def reject_payment_request(
    prid: int,
    body: schemas.PaymentRejectIn,
    current: models.User = Depends(require_roles("finance")),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.PaymentRequest).where(models.PaymentRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "请款单不存在")
    if pr.status != "pending":
        raise HTTPException(400, "只有待审状态可驳回")
    pr.status = "rejected"
    pr.finance_approver_id = current.id
    pr.reject_reason = body.reason
    await db.commit()
    return {"ok": True}


@router.put("/payment-requests/{prid}/pay")
async def pay_payment_request(
    prid: int,
    paid_amount: float = Form(...),
    paid_date: str = Form(...),
    payment_method: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
    current: models.User = Depends(require_roles("finance")),
    db: AsyncSession = Depends(get_db),
):
    """财务付款：记录金额/日期/方式，并可上传付款凭证（付款单据）。
    🆕 需求十六：审批与付款需不同人操作（内控职责分离）——审批人不能给自己审过的单付款。"""
    r = await db.execute(select(models.PaymentRequest).where(models.PaymentRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "请款单不存在")
    if pr.status != "approved":
        raise HTTPException(400, "只有已审批的请款单可付款")
    # 🆕 需求十六：请款审批与付款分成两项、不是一个人操作
    if pr.finance_approver_id and pr.finance_approver_id == current.id:
        raise HTTPException(400, "职责分离：请款审批与付款需由不同人操作，请由另一位财务/出纳付款")
    pr.status = "paid"
    pr.paid_amount = paid_amount
    pr.paid_date = paid_date
    pr.payment_method = payment_method
    if file is not None and file.filename:
        from .attachments_router import save_upload
        att = await save_upload(db, file, biz_type="payment_voucher", biz_id=pr.id, user=current)
        await db.flush()
        pr.pay_voucher_file_id = att.id

    # 回写明细 paid_amount
    ri = await db.execute(
        select(models.PaymentRequestItem).where(models.PaymentRequestItem.request_id == prid)
    )
    pri_rows = ri.scalars().all()
    if pri_rows:
        total_alloc = sum(p.allocated_amount for p in pri_rows)
        for pri in pri_rows:
            ir = await db.execute(
                select(models.PurchaseItem).where(models.PurchaseItem.id == pri.item_id)
            )
            item = ir.scalar_one_or_none()
            if item:
                if total_alloc > 0:
                    ratio = pri.allocated_amount / total_alloc
                    add = round(paid_amount * ratio, 4)
                else:
                    add = round(paid_amount / len(pri_rows), 4)
                item.paid_amount = (item.paid_amount or 0) + add
                item.paid_date = paid_date
                _maybe_auto_reconcile(item)

    await db.commit()
    return {"ok": True}


@router.delete("/payment-requests/{prid}")
async def delete_payment_request(
    prid: int,
    current: models.User = Depends(require_roles("finance")),
    db: AsyncSession = Depends(get_db),
):
    """🆕 请款单全流程删除：待审/已批/已驳/已付 任意状态均可删。
    已付款的先按付款时同一口径把付款从采购明细 paid_amount 冲销掉（必要时回退自动对账状态），
    再删除请款单、分配明细与付款凭证，避免删掉后供应商账上留下「幽灵付款」。"""
    pr = (await db.execute(select(models.PaymentRequest).where(
        models.PaymentRequest.id == prid))).scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "请款单不存在")
    st = pr.status
    pri_rows = list((await db.execute(select(models.PaymentRequestItem).where(
        models.PaymentRequestItem.request_id == prid))).scalars().all())

    # 已付款：反算冲销采购明细 paid_amount（与 /pay 写回口径一致）
    if st == "paid" and pr.paid_amount and pri_rows:
        total_alloc = sum(p.allocated_amount for p in pri_rows)
        for pri in pri_rows:
            item = (await db.execute(select(models.PurchaseItem).where(
                models.PurchaseItem.id == pri.item_id))).scalar_one_or_none()
            if not item:
                continue
            if total_alloc > 0:
                sub = round(pr.paid_amount * (pri.allocated_amount / total_alloc), 4)
            else:
                sub = round(pr.paid_amount / len(pri_rows), 4)
            item.paid_amount = max(0.0, round((item.paid_amount or 0) - sub, 4))
            # 冲销后不再付清、且此前是「自动」置的已对账 → 回退待对账（不动已开票/人工确认态）
            if item.invoice_status == "已对账":
                recv = item.received_amount or 0
                if not (recv > 0 and (item.paid_amount or 0) >= recv - 0.005):
                    item.invoice_status = "待对账"

    # 删付款凭证附件（先断 FK 再删行/物理文件）
    if pr.pay_voucher_file_id:
        att = (await db.execute(select(models.Attachment).where(
            models.Attachment.id == pr.pay_voucher_file_id))).scalar_one_or_none()
        pr.pay_voucher_file_id = None
        await db.flush()
        if att:
            from .attachments_router import delete_attachment_file
            await delete_attachment_file(db, att)

    # 删分配明细 + 请款单（显式删明细，不依赖 DB 级联——SQLite 默认不强制外键）
    await db.execute(delete(models.PaymentRequestItem).where(
        models.PaymentRequestItem.request_id == prid))
    await db.delete(pr)
    await db.commit()
    await write_audit(db, user=current, action="delete", target_type="payment_request",
                      target_id=prid, detail=f"删除请款单(原状态={st})")
    return {"ok": True, "message": "请款单已删除"}


# ==================== 汇总报表 ====================

async def _report_items(db: AsyncSession, current: models.User):
    """🆕 需求五：汇总报表数据也按采购员隔离——受限采购员只统计自己建的采购明细。"""
    stmt = select(models.PurchaseItem)
    if _buyer_restricted(current):
        stmt = stmt.where(models.PurchaseItem.buyer_id == current.id)
    return list((await db.execute(stmt)).scalars().all())


@router.get("/reports/overview", response_model=schemas.PurchaseKPI)
async def report_overview(
    current: models.User = Depends(require_roles("buyer", "buyer_lead", "finance")),
    db: AsyncSession = Depends(get_db),
):
    from datetime import date
    today = date.today()
    this_month = today.strftime("%Y-%m")
    q_start_month = ((today.month - 1) // 3) * 3 + 1
    this_quarter_prefix = f"{today.year:04d}-{q_start_month:02d}"
    this_year = str(today.year)

    items = await _report_items(db, current)

    month_amount = sum(
        i.received_amount or 0 for i in items
        if (i.delivery_date or "").startswith(this_month)
    )
    quarter_amount = sum(
        i.received_amount or 0 for i in items
        if (i.delivery_date or "") >= f"{today.year:04d}-{q_start_month:02d}-01"
    )
    year_amount = sum(
        i.received_amount or 0 for i in items
        if (i.delivery_date or "").startswith(this_year)
    )

    # 🆕 需求五：受限采购员的期初/待审请款也按本人隔离（管理层/主管看全部）
    restricted = _buyer_restricted(current)
    my_sids = None
    if restricted:
        my_sids = {s.id for s in (await db.execute(select(models.Supplier).where(
            models.Supplier.created_by == current.id))).scalars().all()}
    ob_r = await db.execute(select(models.SupplierOpeningBalance))
    ob_total = sum(ob.outstanding_amount or 0 for ob in ob_r.scalars().all()
                   if my_sids is None or ob.supplier_id in my_sids)
    total_received = sum(i.received_amount or 0 for i in items)
    total_paid = sum(i.paid_amount or 0 for i in items)
    total_outstanding = ob_total + total_received - total_paid

    pr_stmt = select(func.count(models.PaymentRequest.id)).where(
        models.PaymentRequest.status == "pending")
    if restricted:
        # 需求四：与请款列表口径一致——按关联采购明细的下单采购员隔离
        mine = (
            select(models.PaymentRequestItem.request_id)
            .join(
                models.PurchaseItem,
                models.PaymentRequestItem.item_id == models.PurchaseItem.id,
            )
            .where(models.PurchaseItem.buyer_id == current.id)
        )
        pr_stmt = pr_stmt.where(models.PaymentRequest.id.in_(mine))
    pr_r = await db.execute(pr_stmt)
    pending_count = pr_r.scalar() or 0

    return schemas.PurchaseKPI(
        month_amount=month_amount,
        quarter_amount=quarter_amount,
        year_amount=year_amount,
        total_outstanding=max(total_outstanding, 0),
        pending_requests=pending_count,
    )


@router.get("/reports/monthly-trend", response_model=List[schemas.PurchaseMonthlyPoint])
async def report_monthly_trend(
    months: int = Query(12),
    current: models.User = Depends(require_roles("buyer", "buyer_lead", "finance")),
    db: AsyncSession = Depends(get_db),
):
    from datetime import date
    items = await _report_items(db, current)

    by_month: dict = defaultdict(lambda: {"amount": 0.0, "paid": 0.0})
    for i in items:
        m = (i.delivery_date or "")[:7]
        if m:
            by_month[m]["amount"] += i.received_amount or 0
            by_month[m]["paid"] += i.paid_amount or 0

    today = date.today()
    result = []
    for delta in range(months - 1, -1, -1):
        total_months = today.year * 12 + today.month - 1 - delta
        y = total_months // 12
        m = total_months % 12 + 1
        key = f"{y:04d}-{m:02d}"
        g = by_month.get(key, {"amount": 0.0, "paid": 0.0})
        result.append(schemas.PurchaseMonthlyPoint(month=key, amount=g["amount"], paid=g["paid"]))
    return result


@router.get("/reports/by-buyer", response_model=List[schemas.PurchaseBuyerRow])
async def report_by_buyer(
    current: models.User = Depends(require_roles("buyer", "buyer_lead", "finance")),
    db: AsyncSession = Depends(get_db),
):
    items = await _report_items(db, current)
    by_buyer: dict = defaultdict(lambda: {"name": "未知", "amount": 0.0, "count": 0})
    for i in items:
        bid = i.buyer_id or 0
        by_buyer[bid]["amount"] += i.received_amount or 0
        by_buyer[bid]["count"] += 1
        if i.buyer:
            by_buyer[bid]["name"] = _uname(i.buyer) or "未知"
    return [
        schemas.PurchaseBuyerRow(
            buyer_id=bid if bid else None,
            buyer_name=v["name"],
            amount=v["amount"],
            count=v["count"],
        )
        for bid, v in sorted(by_buyer.items(), key=lambda x: -x[1]["amount"])
    ]


@router.get("/reports/by-project", response_model=List[schemas.PurchaseProjectRow])
async def report_by_project(
    q: Optional[str] = Query(None),
    current: models.User = Depends(require_roles("buyer", "buyer_lead", "finance")),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(models.PurchaseItem)
    if _buyer_restricted(current):   # 🆕 需求五：采购员只看自己
        stmt = stmt.where(models.PurchaseItem.buyer_id == current.id)
    if q:
        stmt = stmt.where(models.PurchaseItem.project_code.ilike(f"%{q}%"))
    r = await db.execute(stmt)
    items = r.scalars().all()
    by_proj: dict = defaultdict(lambda: {"amount": 0.0, "count": 0})
    for i in items:
        code = i.project_code or "（无项目编号）"
        by_proj[code]["amount"] += i.received_amount or 0
        by_proj[code]["count"] += 1
    return [
        schemas.PurchaseProjectRow(project_code=code, amount=v["amount"], count=v["count"])
        for code, v in sorted(by_proj.items(), key=lambda x: -x[1]["amount"])
    ]


@router.get("/reports/top-suppliers", response_model=List[schemas.PurchaseTopSupplier])
async def report_top_suppliers(
    limit: int = Query(10),
    current: models.User = Depends(require_roles("buyer", "buyer_lead", "finance")),
    db: AsyncSession = Depends(get_db),
):
    items = await _report_items(db, current)
    by_sup: dict = defaultdict(lambda: {"name": "", "amount": 0.0, "count": 0})
    for i in items:
        by_sup[i.supplier_id]["amount"] += i.received_amount or 0
        by_sup[i.supplier_id]["count"] += 1
        if i.supplier:
            by_sup[i.supplier_id]["name"] = i.supplier.name
    top = sorted(by_sup.items(), key=lambda x: -x[1]["amount"])[:limit]
    return [
        schemas.PurchaseTopSupplier(
            supplier_id=sid, supplier_name=v["name"],
            amount=v["amount"], count=v["count"],
        )
        for sid, v in top
    ]


@router.get("/reports/supplier-trend")
async def report_supplier_trend(
    months: int = Query(12, ge=1, le=36),
    top: int = Query(5, ge=1, le=10),
    current: models.User = Depends(require_roles("buyer", "buyer_lead", "finance")),
    db: AsyncSession = Depends(get_db),
):
    """🆕 需求十二：Top-N 供应商近 N 月采购额趋势（多折线图数据源）。
    返回 { months:[YYYY-MM...], series:[{supplier_id, supplier_name, points:[各月采购额], total}] }。
    ——体现：谁是主力供应商、各供应商采购额随时间的走势、是否有集中/波动。"""
    from datetime import date
    items = await _report_items(db, current)
    today = date.today()
    month_keys = []
    for delta in range(months - 1, -1, -1):
        tm = today.year * 12 + today.month - 1 - delta
        month_keys.append(f"{tm // 12:04d}-{tm % 12 + 1:02d}")
    idx = {k: i for i, k in enumerate(month_keys)}
    by_sup_month: dict = defaultdict(lambda: [0.0] * months)
    total_by_sup: dict = defaultdict(float)
    name_by_sup: dict = {}
    for it in items:
        m = (it.delivery_date or "")[:7]
        amt = it.received_amount or 0
        total_by_sup[it.supplier_id] += amt
        if it.supplier:
            name_by_sup[it.supplier_id] = it.supplier.name
        if m in idx:
            by_sup_month[it.supplier_id][idx[m]] += amt
    top_sids = [sid for sid, _ in sorted(total_by_sup.items(), key=lambda x: -x[1])[:top]]
    series = [{
        "supplier_id": sid,
        "supplier_name": name_by_sup.get(sid, f"#{sid}"),
        "points": [round(x, 2) for x in by_sup_month[sid]],
        "total": round(total_by_sup[sid], 2),
    } for sid in top_sids if total_by_sup[sid] > 0]
    return {"months": month_keys, "series": series}


# ==================== 🆕 #167 仓库采购申请（仓库发起 → 采购部处理）====================
# 🆕 请购申请人角色：仓库 + 设计师(设计师请购单,与仓库同一流程/推送)
# 🆕 #198 电工部、🆕 生产部 也可提请购单（与仓库/设计同一流程：指定采购员→推送→行级隔离）
_PREQ_WAREHOUSE = ("warehouse", "warehouse_lead", "designer", "design_lead", "electrician", "electric_lead",
                   "pm_lead", "production_clerk", "assembler", "sheetmetal", "sealing")  # 🆕 #223 封板组漏加→请购单403
_PREQ_VIEW = _PREQ_WAREHOUSE + ("buyer", "buyer_lead", "buyer_standard", "buyer_outsource",
                                "finance", "finance_lead")


def _preq_out(pr: models.PurchaseRequest, atts: Optional[list[dict]] = None) -> schemas.PurchaseRequestOut:
    return schemas.PurchaseRequestOut(
        id=pr.id, requester_id=pr.requester_id, requester_name=_uname(pr.requester),
        buyer_id=pr.buyer_id, buyer_name=_uname(pr.buyer),
        status=pr.status, notes=pr.notes, handler_name=_uname(pr.handler),
        handled_at=pr.handled_at, reject_reason=pr.reject_reason, created_at=pr.created_at,
        lines=[schemas.PurchaseRequestLineOut(
            id=l.id, item_name=l.item_name, spec=l.spec, qty=l.qty,
            project_code=l.project_code, notes=l.notes) for l in pr.lines],
        attachments=atts or [],
    )


async def _preq_atts(db: AsyncSession, pr_ids: list[int]) -> dict[int, list[dict]]:
    """🆕 #245/#246 批量取请购单直传附件 {pr_id: [{id,name}]}。"""
    if not pr_ids:
        return {}
    rows = (await db.execute(select(models.Attachment).where(
        models.Attachment.biz_type == "purchase_request",
        models.Attachment.biz_id.in_(pr_ids)).order_by(models.Attachment.id))).scalars().all()
    out: dict[int, list[dict]] = {}
    for a in rows:
        out.setdefault(a.biz_id, []).append({"id": a.id, "name": a.name})
    return out


@router.get("/buyers")
async def list_buyers(
    current: models.User = Depends(require_roles(*_PREQ_VIEW)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #2 采购员下拉（仓库提采购申请时指定推给谁）。取有采购下单角色的在职用户。"""
    r = await db.execute(
        select(models.User).where(models.User.is_active == True)  # noqa: E712
        .order_by(models.User.full_name))
    codes = {"buyer", "buyer_lead", "buyer_standard", "buyer_outsource"}
    out = []
    for u in r.scalars().all():
        if u.role_codes & codes:
            out.append({"id": u.id, "name": _uname(u)})
    return out


def _preq_warehouse_only(u: models.User) -> bool:
    return u.has_role(*_PREQ_WAREHOUSE) and not u.has_role(
        "buyer", "buyer_lead", "buyer_standard", "buyer_outsource", "admin", "manager")


@router.get("/purchase-requests/{prid}/pdf")
async def download_purchase_request_pdf(
    prid: int,
    current: models.User = Depends(require_roles(*_PREQ_VIEW)),
    db: AsyncSession = Depends(get_db),
):
    """🆕 反馈#232：采购申请查看/打印成正式采购申请单 PDF。仓库/设计师(非采购)只能看自己提的。"""
    r = await db.execute(select(models.PurchaseRequest).where(models.PurchaseRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "采购申请不存在")
    if _preq_warehouse_only(current) and pr.requester_id != current.id:
        raise HTTPException(403, "只能查看自己提交的采购申请")
    pdf_bytes = _render_preq_pdf(pr, _uname(pr.requester), _uname(pr.buyer))
    fname = quote(f"采购申请单_SQ{pr.id:05d}.pdf")
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"})


@router.post("/purchase-requests", response_model=schemas.PurchaseRequestOut)
async def create_purchase_request(
    body: schemas.PurchaseRequestCreate,
    current: models.User = Depends(require_roles(*_PREQ_WAREHOUSE)),
    db: AsyncSession = Depends(get_db),
):
    """仓库/设计师发起采购申请（列出要买什么）→ 推送指定采购员(未指定则全体)，采购部处理。"""
    lines = [l for l in body.lines if (l.item_name or "").strip()]
    # 🆕 #245/#246 二选一：逐行填明细 或 直接上传文件——两者至少有其一
    if not lines and not body.attachment_ids:
        raise HTTPException(400, "请至少填写一行要采购的物料，或上传采购文件")
    pr = models.PurchaseRequest(requester_id=current.id, buyer_id=body.buyer_id,
                                status="pending", notes=body.notes)
    db.add(pr)
    await db.flush()
    for l in lines:
        db.add(models.PurchaseRequestLine(
            request_id=pr.id, item_name=l.item_name.strip(), spec=l.spec,
            qty=l.qty, project_code=l.project_code, notes=l.notes))
    # 🆕 #245/#246 把预上传的附件（biz_id 尚为空）绑定到本单——只认自己刚传的 purchase_request 附件
    if body.attachment_ids:
        atts = (await db.execute(select(models.Attachment).where(
            models.Attachment.id.in_(body.attachment_ids),
            models.Attachment.biz_type == "purchase_request",
            models.Attachment.biz_id.is_(None),
            models.Attachment.uploaded_by == current.id))).scalars().all()
        for a in atts:
            a.biz_id = pr.id
    await db.commit()
    # 🆕 反馈#242：指定了采购员 → 只推给他本人（他人的任务单不该骚扰无关采购员，
    #   尤其王芹这类兼任 buyer 副角色的账号）。未指定 → 才推给全体采购员/主管。
    #   —— 这半是回退 #205「一律推全体」的推送口径；列表可见性不动（仍全员可见）。
    #   （#205 原意是"列表全员可见，所以也通知全员"，但用户反馈指派后仍全员通知太吵。）
    try:
        bname = None
        if body.buyer_id:
            bu = (await db.execute(select(models.User).where(
                models.User.id == body.buyer_id))).scalar_one_or_none()
            bname = _uname(bu) if bu else None
        uids: set = set()
        if body.buyer_id:
            # 指定了采购员：只推给他一人
            text = f"采购申请：{_uname(current)} 提交 {len(lines)} 项，指派给你，请处理"
            uids = {body.buyer_id}
        else:
            # 未指定：推给全体采购员/主管（主/副角色并集去重——含以副角色持有 buyer 的用户）
            text = f"采购申请：{_uname(current)} 提交 {len(lines)} 项待采购物料，请处理"
            rids = [r for (r,) in (await db.execute(select(models.Role.id).where(
                models.Role.code.in_(("buyer", "buyer_lead", "buyer_standard", "buyer_outsource"))))).all()]
            if rids:
                sub = select(models.UserRole.user_id).where(models.UserRole.role_id.in_(rids))
                urs = await db.execute(select(models.User.id).where(
                    models.User.is_active == True,  # noqa: E712
                    or_(models.User.role_id.in_(rids), models.User.id.in_(sub))))
                uids = {u for (u,) in urs.all()}
        for uid in uids:
            db.add(models.Message(to_user_id=uid, kind="info", text=text,
                                  biz_type="purchase_request", biz_id=pr.id))
        await db.commit()
    except Exception:
        pass
    r = await db.execute(select(models.PurchaseRequest).where(models.PurchaseRequest.id == pr.id))
    atts = await _preq_atts(db, [pr.id])
    return _preq_out(r.scalar_one(), atts.get(pr.id))


@router.get("/purchase-requests", response_model=List[schemas.PurchaseRequestOut])
async def list_purchase_requests(
    status: Optional[str] = Query(None),
    current: models.User = Depends(require_roles(*_PREQ_VIEW)),
    db: AsyncSession = Depends(get_db),
):
    """采购申请列表：仓库/设计师(非采购)只看自己提的；采购员/主管看全部。"""
    stmt = select(models.PurchaseRequest).order_by(models.PurchaseRequest.created_at.desc())
    if status:
        stmt = stmt.where(models.PurchaseRequest.status == status)
    if _preq_warehouse_only(current):
        stmt = stmt.where(models.PurchaseRequest.requester_id == current.id)
    r = await db.execute(stmt)
    prs = list(r.scalars().all())
    atts = await _preq_atts(db, [pr.id for pr in prs])
    return [_preq_out(pr, atts.get(pr.id)) for pr in prs]


@router.put("/purchase-requests/{prid}/handle")
async def handle_purchase_request(
    prid: int,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    """采购员/主管把申请标记为「已处理」(已按此下单)。"""
    r = await db.execute(select(models.PurchaseRequest).where(models.PurchaseRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "申请不存在")
    pr.status = "done"
    pr.handled_by = current.id
    pr.handled_at = datetime.now(timezone.utc)
    if pr.requester_id:
        db.add(models.Message(to_user_id=pr.requester_id, kind="info",
                              text=f"你的采购申请 #{pr.id} 已被采购部处理", biz_type="purchase_request", biz_id=pr.id))
    await db.commit()
    return {"ok": True}


@router.put("/purchase-requests/{prid}/reject")
async def reject_purchase_request(
    prid: int,
    body: schemas.PaymentRejectIn,
    current: models.User = Depends(require_roles(*_WRITE_ROLES)),
    db: AsyncSession = Depends(get_db),
):
    r = await db.execute(select(models.PurchaseRequest).where(models.PurchaseRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "申请不存在")
    pr.status = "rejected"
    pr.handled_by = current.id
    pr.handled_at = datetime.now(timezone.utc)
    pr.reject_reason = (body.reason or "").strip() or None
    if pr.requester_id:
        db.add(models.Message(to_user_id=pr.requester_id, kind="warn",
                              text=f"你的采购申请 #{pr.id} 被采购部驳回：{pr.reject_reason or '无原因'}",
                              biz_type="purchase_request", biz_id=pr.id))
    await db.commit()
    return {"ok": True}


@router.delete("/purchase-requests/{prid}")
async def delete_purchase_request(
    prid: int,
    current: models.User = Depends(require_roles(*_PREQ_VIEW)),
    db: AsyncSession = Depends(get_db),
):
    """删除采购申请：仅本人(申请人)或管理层可删。"""
    r = await db.execute(select(models.PurchaseRequest).where(models.PurchaseRequest.id == prid))
    pr = r.scalar_one_or_none()
    if not pr:
        raise HTTPException(404, "申请不存在")
    if pr.requester_id != current.id and not current.has_role("admin", "manager"):
        raise HTTPException(403, "只能删除自己提交的采购申请")
    await db.delete(pr)
    await db.commit()
    return {"ok": True}
