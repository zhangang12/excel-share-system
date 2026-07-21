"""🆕 v3 部门任务单：派单 → 分派 → 接单 → 完成 全状态机（M04/M17）。

状态机：pending_assign 待分派 → assigned 待接单 → in_progress 进行中
        → done 已完成（可 reopen 回 in_progress）；任意未完成态可 voided 作废。

口径：
- B2 仅「进行中」项目可下单
- B5 开始/预计时间一旦填写本人不可改（仅管理层）
- C1-C3 效率口径见 dept_config.compute_efficiency（报表共用）
- D1 设计完成前置校验=五表有 Excel 导入记录（datasheets.imported_at）；#4 起已放开为可选
- P-13 作废留痕（不删单），管理层收通知后可重新下单
- 接单/换人回传一览「设计师/电工」列，设计开始/完成回传「制图开始/制图结束」
"""
import logging
from datetime import date, datetime, timezone, timedelta
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update, or_, delete

from ..database import get_db
from .. import models, schemas
from ..deps import get_current_user, require_roles
from ..dept_config import DEPTS, OUTSOURCE_WORKERS, compute_efficiency
from ..notify import push_message
from ..utils import write_audit
from ..sheet_templates import SHEET_TEMPLATES, OVERVIEW_HEADER_ALIAS
from .attachments_router import save_upload, delete_attachment_file
from .projects_router import OVERVIEW_KEY_PREFIX, HEADER_KEY_PREFIX
from ..sheet_templates import ELEC_PO_SHEET_NAME
from ..config import settings

router = APIRouter(prefix="/api/orders", tags=["部门任务单"])
log = logging.getLogger("app.orders")


def _read_excel_rows(fpath) -> list[list]:
    """读取 Excel 首个工作表为行列表，兼容 .xls(xlrd) 与 .xlsx/.xlsm(openpyxl)。
    日期单元格统一转 YYYY-MM-DD 字符串。"""
    suffix = fpath.suffix.lower()
    rows: list[list] = []
    if suffix == ".xls":
        import xlrd
        from xlrd.xldate import xldate_as_datetime
        try:
            wb = xlrd.open_workbook(str(fpath), encoding_override="cp936")
        except Exception:
            wb = xlrd.open_workbook(str(fpath))
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            row = []
            for ci in range(ws.ncols):
                v = ws.cell_value(r, ci)
                if ws.cell_type(r, ci) == 3:  # xlrd 日期类型
                    try:
                        v = xldate_as_datetime(v, wb.datemode)
                    except Exception:
                        pass
                row.append(v)
            rows.append(row)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return rows


def _cell_to_str(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


async def _populate_elec_po_from_excel(db: AsyncSession, project_id: int, project_code: str, att: models.Attachment) -> int:
    """电工首次上传采购清单 → 解析 Excel 自动填充项目「电工采购单」第5表（§十六）。

    仅当第5表当前无数据行时填充（避免覆盖采购/仓库后续手填的列）。
    按表头名与第5表列名交集映射；'项目'列自动填项目编号。解析失败记日志并跳过（仅留附件引用）。
    返回写入行数。调用方负责 commit。兼容 .xls(xlrd) 与 .xlsx/.xlsm(openpyxl)。
    """
    from pathlib import Path
    from sqlalchemy import func as _f

    res = await db.execute(select(models.Datasheet).where(
        models.Datasheet.project_id == project_id,
        models.Datasheet.name == ELEC_PO_SHEET_NAME,
    ))
    ds = res.scalar_one_or_none()
    if not ds:
        return 0
    # 已有数据行则不覆盖
    cnt = await db.execute(select(_f.count(models.Record.id)).where(models.Record.datasheet_id == ds.id))
    if (cnt.scalar() or 0) > 0:
        return 0

    res = await db.execute(select(models.Field).where(models.Field.datasheet_id == ds.id)
                           .order_by(models.Field.sort_order))
    fields = list(res.scalars().all())
    name_to_fid = {f.name: f.id for f in fields}

    try:
        rows = _read_excel_rows(Path(settings.files_dir) / att.path)   # 兼容 .xls / .xlsx
    except Exception as e:  # noqa: BLE001
        log.warning("[elec_po] 解析采购清单失败 project=%s file=%s err=%s", project_code, att.name, e)
        return 0
    if not rows:
        return 0
    # 找表头行：前 5 行中非空单元格最多的一行
    head_idx, best = 0, -1
    for i, r in enumerate(rows[:5]):
        n = sum(1 for c in r if c not in (None, ""))
        if n > best:
            best, head_idx = n, i
    headers = [str(c).strip() if c is not None else "" for c in rows[head_idx]]
    # 表头列名 → Excel 列下标（命中第5表列名的）
    col_map = {h: i for i, h in enumerate(headers) if h in name_to_fid and h != "项目"}
    # 🆕 #196：'项目'列 = 物料名称（与标准件清单口径一致），优先取 Excel 里的
    #   项目/名称/品名/物料名称 列；Excel 没有该列或该格为空才回退填项目编号。
    #   （旧逻辑硬编码写项目编号，把电工清单里的名称整列丢了。）
    name_col = next((i for i, h in enumerate(headers)
                     if h in ("项目", "名称", "品名", "物料名称")), None)

    written = 0
    for r in rows[head_idx + 1:]:
        if not any(c not in (None, "") for c in r):
            continue
        nm = ""
        if name_col is not None and name_col < len(r) and r[name_col] not in (None, ""):
            nm = _cell_to_str(r[name_col])
        values: dict[str, str] = {str(name_to_fid["项目"]): nm or project_code}
        has_data = False
        for h, ci in col_map.items():
            v = r[ci] if ci < len(r) else None
            if v not in (None, ""):
                values[str(name_to_fid[h])] = _cell_to_str(v)
                has_data = True
        if not has_data:
            continue
        db.add(models.Record(datasheet_id=ds.id, sort_order=written, values=values))
        written += 1
    if written == 0:
        log.warning("[elec_po] 采购清单解析到 0 行 project=%s file=%s 表头=%s 命中列=%s",
                    project_code, att.name, headers, list(col_map.keys()))
    return written


# ==================== 工具 ====================
def _is_mgr(u: models.User) -> bool:
    return u.has_role("admin", "manager")


def _is_lead(u: models.User, dept: str) -> bool:
    return u.has_role(DEPTS[dept]["lead_role"])


def _is_worker_role(u: models.User, dept: str) -> bool:
    return u.has_role(DEPTS[dept]["worker_role"])


def _dept_or_400(dept: str) -> dict:
    cfg = DEPTS.get(dept)
    if not cfg:
        raise HTTPException(400, f"未知部门: {dept}")
    return cfg


async def _order_or_404(db: AsyncSession, oid: int) -> models.DeptOrder:
    res = await db.execute(select(models.DeptOrder).where(models.DeptOrder.id == oid))
    o = res.scalar_one_or_none()
    if not o:
        raise HTTPException(404, "任务单不存在")
    return o


def _uname(u: Optional[models.User]) -> Optional[str]:
    if not u:
        return None
    return u.full_name or u.username


def _writeback_overview(p: models.Project, label: str, value) -> None:
    """回写一览列（__o__label），alias 字段同步双写 __h__（与 header-cell 接口同规则）。
    value=None 表示清除。调用方负责 commit。"""
    extra = dict(p.extra or {})
    o_key = f"{OVERVIEW_KEY_PREFIX}{label}"
    h_alias = OVERVIEW_HEADER_ALIAS.get(label)
    if value is None or value == "":
        extra.pop(o_key, None)
        if h_alias:
            extra.pop(f"{HEADER_KEY_PREFIX}{h_alias}", None)
    else:
        extra[o_key] = value
        if h_alias:
            extra[f"{HEADER_KEY_PREFIX}{h_alias}"] = value
    p.extra = extra


def _valid_date(s: str) -> bool:
    try:
        y, m, d = s.split("-")
        date(int(y), int(m), int(d))
        return True
    except Exception:
        return False


async def _files_of(db: AsyncSession, order_ids: list[int]) -> dict[int, dict[str, list]]:
    """批量取任务单附件并按 (order, 类别) 分组，避免 N+1。"""
    out: dict[int, dict[str, list]] = {
        oid: {"input": [], "start": [], "output": []} for oid in order_ids
    }
    if not order_ids:
        return out
    res = await db.execute(
        select(models.Attachment).where(
            models.Attachment.biz_type.in_(
                ("order_input", "order_start_output", "order_output")),
            models.Attachment.biz_id.in_(order_ids),
        ).order_by(models.Attachment.id)
    )
    group = {"order_input": "input", "order_start_output": "start", "order_output": "output"}
    for a in res.scalars().all():
        out[a.biz_id][group[a.biz_type]].append(schemas.AttachmentOut.model_validate(a))
    return out


_PG_NAME = {"sheetmetal": "钣金", "assembly": "装配", "sealing": "封板"}  # 任务跟踪父视图短标签(🆕 反馈#209)


def _order_to_out(o: models.DeptOrder, files: dict[str, list],
                  notify_name: Optional[str] = None,
                  produce_groups: Optional[list] = None,
                  standard_datasheet_id: Optional[int] = None,
                  packlist_status: Optional[str] = None,
                  material_locations: Optional[list] = None) -> schemas.OrderOut:
    eff, on_time, _od = compute_efficiency(o.start_date, o.due_date, o.done_date)
    today_s = date.today().isoformat()
    overdue = bool(
        (o.status == "in_progress" and o.due_date and today_s > o.due_date)
        or (o.status == "done" and o.done_date and o.due_date and o.done_date > o.due_date)
    )
    return schemas.OrderOut(
        id=o.id, project_id=o.project_id,
        project_code=o.project.code if o.project else "",
        project_name=o.project.name if o.project else "",
        dept=o.dept, status=o.status,
        worker_id=o.worker_id, worker_name=_uname(o.worker),
        req_text=o.req_text,
        start_date=o.start_date, due_date=o.due_date, done_date=o.done_date,
        notify_user_id=o.notify_user_id, notify_user_name=notify_name,
        eff_pct=eff if o.status == "done" else None,
        on_time=on_time if o.status == "done" else None,
        overdue=overdue, created_at=o.created_at,
        design_done_flag=bool(getattr(o, "design_done_flag", False)),
        electric_done_flag=bool(getattr(o, "electric_done_flag", False)),
        ship_prep_done=bool(getattr(o, "ship_prep_done", False)),
        input_files=files.get("input", []),
        start_files=files.get("start", []),
        output_files=files.get("output", []),
        produce_groups=produce_groups,
        standard_datasheet_id=standard_datasheet_id,
        packlist_status=packlist_status,
        material_locations=material_locations or [],
    )


# ==================== 查询 ====================
@router.get("", response_model=List[schemas.OrderOut])
async def list_orders(
    dept: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    year: Optional[str] = Query(None),
    proj_status: Optional[str] = Query(None),
    month: Optional[str] = Query(None, description="YYYY-MM；按接单/制图开始(start_date)月份过滤"),
    worker_id: Optional[int] = Query(None, description="按负责人筛选；仅负责人/管理层可用"),
    limit: int = Query(200, ge=1, le=500),
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """工作台列表：工人=本人任务；部门负责人=本部门全量；管理层=全量（可按 dept 过滤）。"""
    q = (select(models.DeptOrder)
         .join(models.Project, models.DeptOrder.project_id == models.Project.id)
         .where(models.Project.is_deleted == False))  # noqa: E712
    # 调货订单不流转生产，工作台不展示其任务单（防御：存量被误建的 DeptOrder 也隐藏）
    delivery_pids = select(models.SalesLedger.project_id).where(
        models.SalesLedger.order_type == "调货订单")
    q = q.where(models.DeptOrder.project_id.not_in(delivery_pids))
    if year:
        q = q.where(models.Project.code.like(f"{year}-%"))
    if proj_status:
        q = q.where(models.Project.status == proj_status)
    if month:
        # 按接单/制图开始月份过滤（start_date 为 'YYYY-MM-DD' 字符串，前缀匹配）
        q = q.where(models.DeptOrder.start_date.like(f"{month}%"))
    if dept:
        _dept_or_400(dept)
        q = q.where(models.DeptOrder.dept == dept)
    if status:
        q = q.where(models.DeptOrder.status == status)

    # can_filter_worker：能看全部门的负责人/管理层才允许按指定负责人筛选；
    # 工人被强制只看自己，忽略其传入的 worker_id（防越权看他人）
    can_filter_worker = False
    if _is_mgr(current):
        can_filter_worker = True
    elif dept and _is_lead(current, dept):
        can_filter_worker = True
    elif dept and _is_worker_role(current, dept):
        q = q.where(models.DeptOrder.worker_id == current.id)
    else:
        # 未指定 dept 的非管理层：按角色推断本部门
        my_dept = next(
            (k for k, c in DEPTS.items()
             if current.has_role(c["worker_role"], c["lead_role"])),
            None,
        )
        if not my_dept:
            raise HTTPException(403, "无任务单查看权限")
        q = q.where(models.DeptOrder.dept == my_dept)
        if _is_worker_role(current, my_dept):
            q = q.where(models.DeptOrder.worker_id == current.id)
        else:
            can_filter_worker = True

    if worker_id and can_filter_worker:
        q = q.where(models.DeptOrder.worker_id == worker_id)

    res = await db.execute(q.order_by(models.Project.code.desc()).limit(limit))
    orders = list(res.scalars().all())
    files = await _files_of(db, [o.id for o in orders])

    # 通知人姓名批量取
    nids = {o.notify_user_id for o in orders if o.notify_user_id}
    names: dict[int, str] = {}
    if nids:
        r2 = await db.execute(select(models.User).where(models.User.id.in_(nids)))
        names = {u.id: _uname(u) for u in r2.scalars().all()}

    # 🆕 生产单两组(钣金/装配)预计完成/完成，批量取（任务跟踪父视图展示「钣金 X / 装配 Y」）
    pg_map: dict[int, list] = {}
    produce_oids = [o.id for o in orders if o.dept == "produce"]
    if produce_oids:
        gr = await db.execute(select(models.ProduceGroupTask).where(
            models.ProduceGroupTask.order_id.in_(produce_oids)))
        for gt in gr.scalars().all():
            pg_map.setdefault(gt.order_id, []).append(schemas.ProduceGroupBrief(
                group=gt.group, name=_PG_NAME.get(gt.group, gt.group),
                due_date=gt.due_date,
                done_date=(gt.done_at + timedelta(hours=8)).strftime("%Y-%m-%d") if gt.done_at else None,
            ))
    # 🆕 #6 各订单所属项目的「标准件清单」数据表 id（电工部只读引用用）
    std_map: dict[int, int] = {}
    proj_ids = list({o.project_id for o in orders})
    if proj_ids:
        sres = await db.execute(select(models.Datasheet.project_id, models.Datasheet.id).where(
            models.Datasheet.project_id.in_(proj_ids),
            models.Datasheet.name == "标准件清单"))
        for pid_, did_ in sres.all():
            std_map.setdefault(pid_, did_)

    # 🆕 发货清单：批量取各项目当前 packlist_status（设计部对话框用）
    packlist_map: dict[int, str] = {}
    if proj_ids:
        plr = await db.execute(select(models.Shipment.project_id, models.Shipment.packlist_status)
                               .where(models.Shipment.project_id.in_(proj_ids)))
        for pid_, st_ in plr.all():
            packlist_map[pid_] = st_

    # 🆕 #204 各项目材料库位（入库流水去过的库位,去重保序）→ 同步到设计/电工/生产任务跟踪表
    loc_map: dict[int, list[str]] = {}
    if proj_ids:
        lr = await db.execute(select(models.WhTxn.project_id, models.WhTxn.location).where(
            models.WhTxn.project_id.in_(proj_ids), models.WhTxn.direction == "in",
            models.WhTxn.is_reversal == False,  # noqa: E712
            models.WhTxn.location.isnot(None)).order_by(models.WhTxn.id))
        for pid_, loc_ in lr.all():
            loc_ = (loc_ or "").strip()
            if loc_ and loc_ not in loc_map.setdefault(pid_, []):
                loc_map[pid_].append(loc_)

    return [
        _order_to_out(o, files[o.id], names.get(o.notify_user_id), pg_map.get(o.id),
                      std_map.get(o.project_id), packlist_map.get(o.project_id),
                      loc_map.get(o.project_id))
        for o in orders
    ]


@router.get("/options", response_model=schemas.OrderOptionsOut)
async def order_options(
    dept: str = Query(...),
    _: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """工作台下拉数据：可分派工人 / 完成通知人池 / 部门配置。"""
    cfg = _dept_or_400(dept)

    async def _role_users(code: str) -> list[schemas.OrderOptionUser]:
        # 多角色感知：锚点角色 role_id 命中，或 user_roles 关联命中任一即算该角色
        # （修复：兼任负责人/管理层的设计师等，因锚点非该角色而漏出可分派名单）
        rres = await db.execute(select(models.Role.id).where(models.Role.code == code))
        role_id = rres.scalar_one_or_none()
        if role_id is None:
            return []
        sub = select(models.UserRole.user_id).where(models.UserRole.role_id == role_id)
        res = await db.execute(
            select(models.User).where(
                models.User.is_active == True,  # noqa: E712
                or_(models.User.role_id == role_id, models.User.id.in_(sub)),
            ).order_by(models.User.id)
        )
        return [schemas.OrderOptionUser(id=u.id, name=_uname(u)) for u in res.scalars().all()]

    async def _outsource_users() -> list[schemas.OrderOptionUser]:
        """🆕 外协人员（dept_config 按 username 配置）→ 前端据此筛出「外协订单」tab 的数据。"""
        names = OUTSOURCE_WORKERS.get(dept, [])
        if not names:
            return []
        res = await db.execute(
            select(models.User).where(models.User.username.in_(names)).order_by(models.User.id)
        )
        return [schemas.OrderOptionUser(id=u.id, name=_uname(u)) for u in res.scalars().all()]

    return schemas.OrderOptionsOut(
        workers=await _role_users(cfg["worker_role"]),
        outsource_workers=await _outsource_users(),
        notify_pool=await _role_users(cfg["notify_pool"]),
        notify_label=cfg["notify_label"],
        dept_name=cfg["name"],
        sheet_check=cfg["sheet_check"],
        start_outputs=cfg["start_outputs"],
        outputs=cfg["outputs"],
        start_label=cfg["start_label"], end_label=cfg["end_label"], done_label=cfg["done_label"],
    )


# ==================== 下单 ====================
@router.post("", response_model=schemas.OrderOut)
async def create_order(
    data: schemas.OrderCreate,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """下单（双入口）：管理层目录下单可带 worker_id 直接指派；
    销售下单（M02 内部调用 create_order_internal）建待分派单。"""
    if not _is_mgr(current):
        raise HTTPException(403, "仅管理层可在项目目录下单")
    cfg = _dept_or_400(data.dept)

    res = await db.execute(select(models.Project).where(
        models.Project.id == data.project_id, models.Project.is_deleted == False))  # noqa: E712
    p = res.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "项目不存在")
    if p.status != "进行中":
        raise HTTPException(400, "仅「进行中」项目可下单")  # B2

    o = models.DeptOrder(
        project_id=p.id, dept=data.dept, req_text=(data.req_text or "").strip() or None,
        created_by=current.id,
    )
    if data.worker_id:
        res = await db.execute(select(models.User).where(models.User.id == data.worker_id))
        w = res.scalar_one_or_none()
        if not w or not w.is_active or not _is_worker_role(w, data.dept):
            raise HTTPException(400, f"指派对象必须是{cfg['name']}{DEPTS[data.dept]['worker_role']}角色")
        o.worker_id = w.id
        o.status = "assigned"
    db.add(o)
    await db.commit()
    await db.refresh(o)

    if o.status == "assigned":
        await push_message(db, to_user_id=o.worker_id, kind="info",
                           text=f"【分派】{p.code} {p.name} 已分派给你，请到{cfg['name']}工作台接单填写时间。",
                           biz_type="order", biz_id=o.id)
    else:
        await push_message(db, to_role=cfg["lead_role"], kind="info",
                           text=f"【待分派】{p.code} {p.name} 新{cfg['name']}任务待分派。",
                           biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="create", target_type="dept_order",
                      target_id=o.id, detail=f"{p.code} {data.dept}")
    return _order_to_out(o, {"input": [], "start": [], "output": []})


async def create_order_internal(
    db: AsyncSession, *, project: models.Project, dept: str,
    req_text: Optional[str], created_by: int,
) -> models.DeptOrder:
    """供销售下单（M02）在同一事务内建待分派任务。调用方负责 commit + 推送。"""
    _dept_or_400(dept)
    o = models.DeptOrder(
        project_id=project.id, dept=dept,
        req_text=(req_text or "").strip() or None, created_by=created_by,
    )
    db.add(o)
    await db.flush()
    return o


@router.post("/spare", response_model=schemas.SalesOrderOut)
async def create_spare_order(
    data: schemas.SpareOrderCreate,
    current: models.User = Depends(require_roles("admin", "manager", "design_lead")),
    db: AsyncSession = Depends(get_db),
):
    """🆕 备机下单（设计部负责人/管理层）：建项目(进项目目录/详单)+派各部门任务，
    复用下单派单/推送逻辑，但不建销售台账(不在销售部管理)。直接生效、无需审批。"""
    from .projects_router import create_default_template_sheets, _add_all_active_users_as_members
    name = data.name.strip()
    if not name:
        raise HTTPException(400, "请填写设备名称")
    code = (data.code or "").strip()
    if not code:
        raise HTTPException(400, "请填写项目编号")
    depts = [d for d in data.depts if d in DEPTS] or ["produce", "electric"]
    res = await db.execute(select(models.Project).where(models.Project.code == code))
    if res.scalar_one_or_none():
        raise HTTPException(409, f"项目编号 {code} 已存在")

    p = models.Project(code=code, name=name, status="进行中", manager_id=None)
    db.add(p)
    await db.flush()
    await create_default_template_sheets(db, p.id)
    await _add_all_active_users_as_members(db, p.id)
    u = data.unit if data.unit in ("台", "套") else "台"
    n = data.qty if isinstance(data.qty, int) and data.qty >= 1 else 1
    _writeback_overview(p, "数量", f"{n}{u}")
    _writeback_overview(p, "销售", f"备机·{_uname(current)}")  # 标记备机便于在项目目录区分

    req = data.req_text.strip() or f"（备机下单）{name}"
    order_ids = []
    for d in depts:
        o = await create_order_internal(db, project=p, dept=d, req_text=req, created_by=current.id)
        order_ids.append(o.id)
    await db.commit()

    for d in depts:
        await push_message(db, to_role=DEPTS[d]["lead_role"], kind="info",
                           text=f"【备机下单】{code} {name} 新{DEPTS[d]['name']}任务待分派（下单：{_uname(current)}）。",
                           biz_type="project", biz_id=p.id)
    await write_audit(db, user=current, action="create_spare", target_type="project",
                      target_id=p.id, detail=f"{code} 备机 派往{','.join(depts)}")
    return schemas.SalesOrderOut(project_id=p.id, code=code, order_ids=order_ids)


# ==================== 附件 ====================
@router.post("/{oid}/input-files", response_model=List[schemas.AttachmentOut])
async def upload_input_files(
    oid: int,
    files: List[UploadFile] = File(...),
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """下发资料（下单方/管理层上传）。"""
    o = await _order_or_404(db, oid)
    if not (_is_mgr(current) or o.created_by == current.id):
        raise HTTPException(403, "仅下单方或管理层可传下发资料")
    outs = []
    for f in files:
        a = await save_upload(db, f, biz_type="order_input", biz_id=o.id,
                              project_id=o.project_id, user=current)
        outs.append(a)
    await db.commit()
    return [schemas.AttachmentOut.model_validate(a) for a in outs]


@router.post("/{oid}/start-upload", response_model=List[schemas.AttachmentOut])
async def start_upload(
    oid: int,
    kind: str = Query(...),
    files: List[UploadFile] = File(...),
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """接单后上传（设计图纸包→钣金 / 电工采购清单→采购），多文件累加。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    so = next((s for s in cfg["start_outputs"] if s["k"] == kind), None)
    if not so:
        raise HTTPException(400, f"{cfg['name']}没有 {kind} 类型的接单上传")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可上传")
    # 🆕 #5 例外：设计部「设计资料」(CAD激光图纸/外购附图)在设计完成(done)后仍可更换
    is_design_resource = o.dept == "design" and kind in ("sheetpkg", "outsource_img", "sealing_pkg", "coldwork_pkg")
    if o.status not in ("in_progress", "assigned") and not (is_design_resource and o.status == "done"):
        raise HTTPException(400, "任务未在进行中")

    outs = []
    for f in files:
        a = await save_upload(db, f, biz_type="order_start_output", biz_id=o.id,
                              kind=kind, project_id=o.project_id, user=current)
        outs.append(a)
    p = o.project
    # 🆕 M12：电工采购清单 → 首次自动解析写入项目「电工采购单」第5表（兼容 .xls/.xlsx）
    elec_po_written: Optional[int] = None
    if o.dept == "electric" and kind == "plist" and outs:
        try:
            elec_po_written = await _populate_elec_po_from_excel(db, o.project_id, p.code, outs[0])
        except Exception as e:  # noqa: BLE001  解析失败不阻塞上传
            log.warning("[elec_po] 自动入表异常 project=%s err=%s", p.code, e)
            elec_po_written = 0
    await db.commit()

    # 解析未写入任何行 → 给上传者一条提示（不阻塞上传，便于排查表头/格式）
    if elec_po_written == 0:
        await push_message(db, to_user_id=current.id, kind="warn",
                           text=f"【电工采购单】{p.code} 采购清单已上传，但未能自动解析入表（请检查表头列名是否与「电工采购单」一致），可重新上传或在项目详单手动录入。",
                           biz_type="project", biz_id=o.project_id)

    # 电工采购清单：直接推送给 lixinxin（李新新），不广播给全体采购角色
    if o.dept == "electric" and kind == "plist":
        res2 = await db.execute(select(models.User).where(
            models.User.username == "lixinxin",
            models.User.is_active == True,  # noqa: E712
        ))
        lxx = res2.scalar_one_or_none()
        if lxx:
            await push_message(db, to_user_id=lxx.id, kind="info",
                               text=f"【采购清单】{p.code} {cfg['name']}已上传采购清单 {len(outs)} 个文件，请查收。",
                               biz_type="order", biz_id=o.id)
    else:
        await push_message(db, to_role=so["to_role"], kind="info",
                           text=f"【{so['label']}】{cfg['name']}已上传 {p.code} {so['label']} {len(outs)} 个文件，请查收。",
                           biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="upload", target_type="dept_order",
                      target_id=o.id, detail=f"start:{kind} x{len(outs)}")
    return [schemas.AttachmentOut.model_validate(a) for a in outs]


@router.post("/{oid}/output-upload", response_model=List[schemas.AttachmentOut])
async def output_upload(
    oid: int,
    kind: str = Query(...),
    files: List[UploadFile] = File(...),
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """完成产物上传（完成弹窗内逐项上传，complete 时校验必传项）。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not any(x["k"] == kind for x in cfg["outputs"]):
        raise HTTPException(400, f"{cfg['name']}没有 {kind} 类型的产物")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可上传")
    # 🆕 #50 状态守卫：已完成/已作废单不得再挂产物（避免改下游资料/破坏留痕），与 start_upload 对齐
    # 🆕 #5 例外：设计部「发货准备」(说明书/铭牌)在设计完成(done)后仍可补传/替换
    # 🆕 #197：电工「电路图」同设计说明书/铭牌口径——完成(done)后仍可补传（发货准备）
    is_ship_prep = (o.dept == "design" and kind in ("manual", "nameplate")) \
        or (o.dept == "electric" and kind == "circuit")
    if o.status not in ("in_progress", "assigned") and not (is_ship_prep and o.status == "done"):
        raise HTTPException(400, "任务未在进行中，不能上传产物")
    outs = []
    for f in files:
        a = await save_upload(db, f, biz_type="order_output", biz_id=o.id,
                              kind=kind, project_id=o.project_id, user=current)
        outs.append(a)
    await db.commit()

    # 产物上传后按 dept_config.outputs[].to_role 推送给下游部门
    ot_cfg = next((x for x in cfg["outputs"] if x["k"] == kind), None)
    if ot_cfg and ot_cfg.get("to_role"):
        p = o.project
        await push_message(
            db, to_role=ot_cfg["to_role"], kind="info",
            text=f"【{ot_cfg['label']}】{cfg['name']}已上传 {p.code} {ot_cfg['label']} {len(outs)} 个文件，请查收。",
            biz_type="order", biz_id=o.id,
        )

    await write_audit(db, user=current, action="upload", target_type="dept_order",
                      target_id=o.id, detail=f"output:{kind} x{len(outs)}")
    return [schemas.AttachmentOut.model_validate(a) for a in outs]


@router.delete("/{oid}/attachments/{aid}", response_model=schemas.Msg)
async def remove_order_attachment(
    oid: int, aid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    o = await _order_or_404(db, oid)
    if not (_is_mgr(current) or o.worker_id == current.id or o.created_by == current.id):
        raise HTTPException(403, "无权移除")
    res = await db.execute(select(models.Attachment).where(
        models.Attachment.id == aid, models.Attachment.biz_id == oid,
        models.Attachment.biz_type.in_(
            ("order_input", "order_start_output", "order_output")),
    ))
    a = res.scalar_one_or_none()
    if not a:
        raise HTTPException(404, "附件不存在")
    # 🆕 #53 已完成/已作废单的附件不可移除（防止完成后删空必传产物形成非法终态）
    # 🆕 #5 例外：设计部已完成单允许替换「设计资料/产物」(CAD激光图纸/外购附图=order_start_output、
    #         说明书/铭牌=order_output)，方便发货前更正
    if o.status == "voided" or (o.status == "done" and not (
            (o.dept == "design" and a.biz_type in ("order_output", "order_start_output"))
            or (o.dept == "electric" and a.biz_type == "order_output"))):  # 🆕 #197 电路图可替换
        raise HTTPException(400, "任务已完成/作废，附件不可移除（如需修改请先改回进行中）")
    name = a.name
    await delete_attachment_file(db, a)
    await db.commit()
    await write_audit(db, user=current, action="delete", target_type="dept_order",
                      target_id=oid, detail=f"附件:{name}")
    return schemas.Msg(message="已移除")


@router.post("/{oid}/ship-prep-done", response_model=schemas.Msg)
async def ship_prep_done(
    oid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #5 设计部「发货准备完成」：设计完成(done)后，确认说明书/铭牌等发货资料已备齐，
    标记 ship_prep_done 并通知物流。仅设计部已完成单可点（说明书/铭牌仍为选填，按需）。"""
    o = await _order_or_404(db, oid)
    if o.dept not in ("design", "electric"):   # 🆕 #197 电工部同口径（电路图=发货资料）
        raise HTTPException(400, "仅设计部/电工部任务支持发货准备完成")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可操作")
    if o.status != "done":
        raise HTTPException(400, "请先完成任务（设计完成/接线完成），再做发货准备完成")
    o.ship_prep_done = True
    await db.commit()
    p = o.project
    is_spare = str((p.extra or {}).get("__o__销售") or "").startswith("备机")
    if not is_spare:
        what = "设计部说明书/铭牌" if o.dept == "design" else "电工部电路图"
        await push_message(db, to_role="logistics", kind="info",
                           text=f"【发货准备完成】{p.code} {p.name} {what}等发货资料已备齐。",
                           biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="ship_prep_done", target_type="dept_order",
                      target_id=o.id)
    return schemas.Msg(message="已标记发货准备完成" + ("，并通知物流" if not is_spare else ""))


@router.post("/{oid}/ship-list-upload", response_model=schemas.AttachmentOut)
async def ship_list_upload(
    oid: int,
    file: UploadFile = File(...),
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🆕 发货清单：设计部上传发货清单文件 → 存为项目级 ship_list（仓库「发货清单」可下载/打印），
    并把该项目发货单据标记为「待备货」、通知仓库；仓库备货完成后通知物流。可重复上传替换。"""
    o = await _order_or_404(db, oid)
    if o.dept != "design":
        raise HTTPException(400, "仅设计部任务支持上传发货清单")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可操作")
    if o.status != "done":
        raise HTTPException(400, "请先点「设计完成」，再上传发货清单")
    a = await save_upload(db, file, biz_type="ship_list", biz_id=o.project_id,
                          project_id=o.project_id, user=current)
    # 发货单据标记为待备货（若存在且尚未备货完成）
    r = await db.execute(select(models.Shipment).where(models.Shipment.project_id == o.project_id))
    sh = r.scalar_one_or_none()
    if sh and sh.packlist_status != "ready":
        sh.packlist_status = "requested"
        sh.packlist_requested_at = datetime.now(timezone.utc)
        sh.packlist_requested_by = current.id
    await db.commit()
    await db.refresh(a)
    p = o.project
    for role in ("warehouse", "warehouse_lead"):
        await push_message(db, to_role=role, kind="info",
                           text=f"【发货清单待备货】{p.code} {p.name} 设计部已下发发货清单，请仓库备货。",
                           biz_type="project", biz_id=p.id)
    # 🆕 设计同时直推发货部/物流：让发货部提前拿到发货清单，仓库备齐后再安排发货
    await push_message(db, to_role="logistics", kind="info",
                       text=f"【发货清单已下发】{p.code} {p.name} 设计部已下发发货清单，待仓库备齐后安排发货。",
                       biz_type="project", biz_id=p.id)
    await write_audit(db, user=current, action="ship_list_upload", target_type="attachment",
                      target_id=a.id)
    return schemas.AttachmentOut.model_validate(a)


@router.post("/{oid}/revision-request", response_model=schemas.Msg)
async def revision_request(
    oid: int, data: schemas.RevisionRequestIn,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🆕 #1 设计/电工对销售下发的「合同技术资料」提修订意见 → 推送对应销售员 + 抄送销售主管。
    销售用「更换技术资料」上传修正版后该意见自动标记 resolved 并回通知提出人。"""
    o = await _order_or_404(db, oid)
    if o.dept not in ("design", "electric"):
        raise HTTPException(400, "仅设计/电工部可对技术资料提修订意见")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可提修订意见")
    reason = (data.reason or "").strip()
    if not reason:
        raise HTTPException(400, "请填写修订意见")
    db.add(models.RevisionRequest(
        project_id=o.project_id, order_id=o.id, dept=o.dept,
        reason=reason, raised_by=current.id))
    await db.commit()
    p = o.project
    res = await db.execute(select(models.SalesLedger).where(
        models.SalesLedger.project_id == o.project_id))
    led = res.scalar_one_or_none()
    short = reason if len(reason) <= 40 else reason[:40] + "…"
    dept_name = DEPTS[o.dept]["name"]
    if led and led.sales_uid:
        await push_message(db, to_user_id=led.sales_uid, kind="warn",
                           text=f"【技术资料待修订】{p.code} {dept_name}提出修订意见：{short}，请到销售台账「更换技术资料」上传修正版。",
                           biz_type="project", biz_id=o.project_id)
    await push_message(db, to_role="sales_lead", kind="info",
                       text=f"【技术资料修订】{p.code} {dept_name}提出修订意见：{short}",
                       biz_type="project", biz_id=o.project_id)
    await write_audit(db, user=current, action="revision_request", target_type="dept_order",
                      target_id=o.id, detail=reason[:200])
    return schemas.Msg(message="修订意见已提交，已通知销售更换技术资料")


# ==================== 状态流转 ====================
@router.post("/{oid}/assign", response_model=schemas.Msg)
async def assign_order(
    oid: int, data: schemas.OrderAssignIn,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """部门负责人分派给具体工人。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not (_is_mgr(current) or _is_lead(current, o.dept)):
        raise HTTPException(403, f"仅{cfg['name']}负责人可分派")
    if o.status != "pending_assign":
        raise HTTPException(400, "仅待分派任务可分派")
    res = await db.execute(select(models.User).where(models.User.id == data.worker_id))
    w = res.scalar_one_or_none()
    if not w or not w.is_active or not _is_worker_role(w, o.dept):
        raise HTTPException(400, f"分派对象必须是{cfg['name']}工人角色")
    o.worker_id = w.id
    o.status = "assigned"
    await db.commit()
    p = o.project
    await push_message(db, to_user_id=w.id, kind="info",
                       text=f"【分派】{p.code} {p.name} 已分派给你，请到{cfg['name']}工作台接单填写时间。",
                       biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="assign", target_type="dept_order",
                      target_id=o.id, detail=f"→{_uname(w)}")
    return schemas.Msg(message=f"已分派给 {_uname(w)}")


@router.post("/{oid}/start", response_model=schemas.Msg)
async def start_order(
    oid: int, data: schemas.OrderStartIn,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """工人接单：填开始/预计完成时间并开始；回传一览（设计师/电工/制图开始）。
    B5：时间一旦填写本人不可改（重开始被拒），仅管理层可改。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可接单")
    if o.status != "assigned":
        raise HTTPException(400, "仅待接单任务可开始")
    if not _valid_date(data.start_date) or not _valid_date(data.due_date):
        raise HTTPException(400, "请填写有效日期 (YYYY-MM-DD)")
    if data.due_date < data.start_date:
        raise HTTPException(400, "预计完成不能早于开始")
    if (o.start_date or o.due_date) and not _is_mgr(current):
        raise HTTPException(403, "时间已填写，本人不可修改（仅管理层可改）")  # B5

    o.start_date, o.due_date = data.start_date, data.due_date
    o.status = "in_progress"

    # 回传一览：设计→设计师+制图开始；电工→电工（alias 双写 __h__ 电器）
    p = o.project
    wname = _uname(o.worker) or _uname(current)
    if cfg["writeback_worker"]:
        _writeback_overview(p, cfg["writeback_worker"], wname)
    if cfg["writeback_start"]:
        _writeback_overview(p, cfg["writeback_start"], data.start_date)
    await db.commit()
    await write_audit(db, user=current, action="start", target_type="dept_order",
                      target_id=o.id, detail=f"{data.start_date}~{data.due_date}")
    return schemas.Msg(message="已开始" + ("，已回传项目目录" if cfg["writeback_worker"] else ""))


@router.post("/{oid}/design_done", response_model=schemas.Msg)
async def mark_design_done(
    oid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """设计完成 = 直接完成(考核按设计完成日期)：置 status=done、计入考核。
    🆕 #4 放开文件必传校验：CAD激光图纸/外购附图/四表导入均改为可选，部分订单无需图纸表格也可完成流转。"""
    o = await _order_or_404(db, oid)
    if o.dept != "design":
        raise HTTPException(400, "仅设计部任务可用")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可操作")
    if o.status != "in_progress":
        raise HTTPException(400, "仅进行中任务可操作")

    # 🆕 #4 已放开：不再校验 CAD激光图纸/外购附图已上传、五表是否导入齐（图纸表格按需可选）

    # 🆕 点「设计完成」即完成：置 done + 写完成日期 + 计考核（发货准备的说明书/铭牌不在此校验/考核）
    cfg = DEPTS["design"]
    today_s = date.today().isoformat()
    o.design_done_flag = True
    o.status = "done"
    o.done_date = today_s
    p = o.project
    if cfg["writeback_done"]:
        _writeback_overview(p, cfg["writeback_done"], today_s)
    await db.commit()

    eff, on_time, overdue_days = compute_efficiency(o.start_date, o.due_date, o.done_date)
    if overdue_days > 0:
        text = (f"【逾期完成】设计 {p.code} {p.name}：预计 {o.due_date}，实际 {o.done_date}，"
                f"超 {overdue_days} 天，效率 {eff}%（负责人：{_uname(o.worker)}）")
        await push_message(db, to_role=cfg["lead_role"], kind="warn", text=text, biz_type="order", biz_id=o.id)
        await push_message(db, to_role="manager", kind="warn", text=text, biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="design_done", target_type="dept_order",
                      target_id=o.id, detail=f"done eff={eff}% on_time={on_time}")
    return schemas.Msg(message="设计完成，已计入考核。可继续上传产品说明书/铭牌（发货准备，不影响考核）"
                       + (f"；逾期 {overdue_days} 天已提醒主管" if overdue_days else ""))


@router.post("/{oid}/electric_done", response_model=schemas.Msg)
async def mark_electric_done(
    oid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """接线完成第一步：置 electric_done_flag。🆕 #4 放开采购清单必传校验（按需可选）。"""
    o = await _order_or_404(db, oid)
    if o.dept != "electric":
        raise HTTPException(400, "仅电工部任务可用")
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可操作")
    if o.status != "in_progress":
        raise HTTPException(400, "仅进行中任务可操作")

    # 🆕 #4 已放开：不再校验采购清单是否上传（按需可选）
    # 🆕 #197：接线完成 = 直接完成计考核（与设计部「设计完成」一致）；
    #   电路图上传/发货准备挪到完成后做，不再阻塞订单完成、不影响考核。
    cfg = DEPTS["electric"]
    today_s = date.today().isoformat()
    o.electric_done_flag = True
    o.status = "done"
    o.done_date = today_s
    await db.commit()
    p = o.project
    eff, on_time, overdue_days = compute_efficiency(o.start_date, o.due_date, o.done_date)
    if overdue_days > 0:
        text = (f"【逾期完成】电工 {p.code} {p.name}：预计 {o.due_date}，实际 {o.done_date}，"
                f"超 {overdue_days} 天，效率 {eff}%（负责人：{_uname(o.worker)}）")
        await push_message(db, to_role=cfg["lead_role"], kind="warn", text=text, biz_type="order", biz_id=o.id)
        await push_message(db, to_role="manager", kind="warn", text=text, biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="electric_done", target_type="dept_order",
                      target_id=o.id, detail=f"done eff={eff}% on_time={on_time}")
    return schemas.Msg(message="接线完成，已计入考核。可到「已完成」里上传电路图并做发货准备（不影响考核）"
                       + (f"；逾期 {overdue_days} 天已提醒主管" if overdue_days else ""))


@router.post("/{oid}/complete", response_model=schemas.Msg)
async def complete_order(
    oid: int, data: schemas.OrderCompleteIn,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """完成任务：设计部须先完成第一步（design_done_flag）→ 必传产物校验 → 通知人必选 → done；
    逾期则推部门主管+抄送管理层（含效率%）。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not (_is_mgr(current) or o.worker_id == current.id):
        raise HTTPException(403, "仅任务负责人可完成")
    if o.status != "in_progress":
        raise HTTPException(400, "仅进行中任务可完成")

    # 两步完成流第一步闸门
    if o.dept == "design" and not getattr(o, "design_done_flag", False):
        raise HTTPException(400, "请先点击「设计完成」完成第一步")
    if o.dept == "electric" and not getattr(o, "electric_done_flag", False):
        raise HTTPException(400, "请先点击「接线完成」完成第一步")

    # 🆕 #4 已放开：不再校验必传产物（上传资料按需可选，无需上传也可完成）

    # 通知人校验（必选，且属于通知池角色）
    res = await db.execute(select(models.User).join(models.Role).where(
        models.User.id == data.notify_user_id,
        models.Role.code == cfg["notify_pool"],
        models.User.is_active == True,  # noqa: E712
    ))
    notify_user = res.scalar_one_or_none()
    if not notify_user:
        raise HTTPException(400, f"请选择{cfg['notify_label']}")

    today_s = date.today().isoformat()
    o.status = "done"
    o.done_date = today_s
    o.notify_user_id = notify_user.id

    p = o.project
    if cfg["writeback_done"]:
        _writeback_overview(p, cfg["writeback_done"], today_s)
    await db.commit()

    eff, on_time, overdue_days = compute_efficiency(o.start_date, o.due_date, o.done_date)
    wname = _uname(o.worker)
    await push_message(db, to_user_id=notify_user.id, kind="wx",
                       text=f"【{cfg['name']}·{p.code}】{wname} 已完成{cfg['done_label']}。",
                       biz_type="order", biz_id=o.id)
    if overdue_days > 0:
        text = (f"【逾期完成】{cfg['name']} {p.code} {p.name}：预计 {o.due_date}，"
                f"实际 {o.done_date}，超 {overdue_days} 天，效率 {eff}%（负责人：{wname}）")
        await push_message(db, to_role=cfg["lead_role"], kind="warn",
                           text=text, biz_type="order", biz_id=o.id)
        await push_message(db, to_role="manager", kind="warn",
                           text=text, biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="complete", target_type="dept_order",
                      target_id=o.id, detail=f"eff={eff}% on_time={on_time}")
    return schemas.Msg(message="已完成" + (f"，逾期 {overdue_days} 天已提醒主管" if overdue_days else ""))


@router.post("/{oid}/reopen", response_model=schemas.Msg)
async def reopen_order(
    oid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """完成可逆（P-06）：done → in_progress，清完成日期；设计清一览制图结束。
    已上传的图纸包/采购清单(order_start_output)保留在下游收件箱（reopen=返工，资料仍有效）。
    但若项目已发货则禁止改回，避免"已发货却闸门未通过"的不一致（#51）。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not (_is_mgr(current) or o.worker_id == current.id or _is_lead(current, o.dept)):
        raise HTTPException(403, "仅任务负责人/部门负责人可改回")
    if o.status != "done":
        raise HTTPException(400, "仅已完成任务可改回进行中")
    # 🆕 #51 已发货项目不得改回（如需返工请先由物流撤回发货）
    sr = await db.execute(select(models.Shipment).where(
        models.Shipment.project_id == o.project_id,
        models.Shipment.status == "shipped",
    ))
    if sr.scalar_one_or_none() is not None:
        raise HTTPException(400, "该项目已发货，任务不可改回进行中（如需返工请先撤回发货）")
    o.status = "in_progress"
    o.done_date = None
    # 🆕 #197：改回=返工，设计/接线完成标记一并清掉，回到"未完成"状态机
    #   （也是 close_legacy_electric_done_orders 止损迁移不误关返工单的前提——
    #    迁移按 in_progress+electric_done_flag 识别"卡在旧二步流的存量单"）
    o.design_done_flag = False
    o.electric_done_flag = False
    p = o.project
    if cfg["writeback_done"]:
        _writeback_overview(p, cfg["writeback_done"], None)
    await db.commit()
    await write_audit(db, user=current, action="reopen", target_type="dept_order",
                      target_id=o.id)
    return schemas.Msg(message="已改回进行中")


@router.post("/{oid}/void", response_model=schemas.Msg)
async def void_order(
    oid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """作废单号（P-13 留痕；管理层收通知后可重新下单=新任务）。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not (_is_mgr(current) or _is_lead(current, o.dept)):
        raise HTTPException(403, f"仅{cfg['name']}负责人/管理层可作废")
    if o.status == "done":
        raise HTTPException(400, "已完成任务不可作废（请先改回进行中）")
    if o.status == "voided":
        raise HTTPException(400, "任务已作废")
    o.status = "voided"
    # 🆕 #30 作废设计任务 → 其负责的「待设计接收」反馈置空归属，转入设计负责人待指派列表
    if o.dept == "design" and o.worker_id:
        await db.execute(
            update(models.Feedback).where(
                models.Feedback.project_id == o.project_id,
                models.Feedback.status == "pending_design",
                models.Feedback.designer_uid == o.worker_id,
            ).values(designer_uid=None)
        )
    await db.commit()
    p = o.project
    await push_message(db, to_role="manager", kind="warn",
                       text=f"【作废】{cfg['name']} {p.code} 单号已作废，请管理层确认是否重新下单。",
                       biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="void", target_type="dept_order", target_id=o.id)
    return schemas.Msg(message="已作废单号，已通知管理层")


@router.delete("/{oid}", response_model=schemas.Msg)
async def delete_order(
    oid: int,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🆕 管理层直接删除任务单（彻底删除，区别于「作废」留痕）。仅 admin/manager。
    清理：① 生产分组任务 ProduceGroupTask（FK order_id）② 订单关联附件（DB 行 + 磁盘文件）
    ③ 设计任务的「待设计接收」反馈归属置空（与作废同口径）。删除后下游（采购/钣金/物流）
    因 JOIN 不到该单而自动消失。任意状态均可删（含已作废残留）。"""
    o = await _order_or_404(db, oid)
    if not _is_mgr(current):
        raise HTTPException(403, "仅管理层可删除任务单")
    cfg = DEPTS[o.dept]
    code = o.project.code if o.project else str(o.project_id)
    old_status = o.status
    # 1) 生产部：先删两组分组任务（FK order_id → dept_orders.id，否则删父单违反外键）
    if o.dept == "produce":
        await db.execute(delete(models.ProduceGroupTask).where(models.ProduceGroupTask.order_id == oid))
    # 2) 删除订单关联附件（DB 行 + 磁盘文件），避免孤儿文件
    ar = await db.execute(select(models.Attachment).where(
        models.Attachment.biz_id == oid,
        models.Attachment.biz_type.in_(("order_input", "order_start_output", "order_output")),
    ))
    for a in ar.scalars().all():
        await delete_attachment_file(db, a)
    # 3) 设计任务：其「待设计接收」反馈归属置空（避免悬挂到已删任务的负责人）
    if o.dept == "design" and o.worker_id:
        await db.execute(update(models.Feedback).where(
            models.Feedback.project_id == o.project_id,
            models.Feedback.status == "pending_design",
            models.Feedback.designer_uid == o.worker_id,
        ).values(designer_uid=None))
    # 4) 删除任务单本身
    await db.delete(o)
    await db.commit()
    await write_audit(db, user=current, action="delete", target_type="dept_order",
                      target_id=oid, detail=f"{cfg['name']} {code} 任务单已删除（原状态={old_status}）")
    return schemas.Msg(message="已删除任务单")


@router.post("/{oid}/reassign", response_model=schemas.Msg)
async def reassign_order(
    oid: int, data: schemas.OrderReassignIn,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🆕 M17 换人（防离职/请假）：待接单/进行中任务转交同部门他人；
    时间与已传产物保留；设计/电工回传项目目录。"""
    o = await _order_or_404(db, oid)
    cfg = DEPTS[o.dept]
    if not (_is_mgr(current) or _is_lead(current, o.dept)):
        raise HTTPException(403, f"仅{cfg['name']}负责人/管理层可换人")
    if o.status not in ("assigned", "in_progress"):
        raise HTTPException(400, "仅待接单/进行中任务可换人")
    if data.worker_id == o.worker_id:
        raise HTTPException(400, "转交对象不能是当前负责人")
    res = await db.execute(select(models.User).where(models.User.id == data.worker_id))
    w = res.scalar_one_or_none()
    if not w or not w.is_active or not _is_worker_role(w, o.dept):
        raise HTTPException(400, f"转交对象必须是{cfg['name']}工人角色")

    old_wid = o.worker_id
    old_name = _uname(o.worker) or "—"
    o.worker_id = w.id
    p = o.project
    new_name = _uname(w)
    if cfg["writeback_worker"] and o.status == "in_progress":
        _writeback_overview(p, cfg["writeback_worker"], new_name)
    # 🆕 #30 设计换人 → 该项目「待设计接收」反馈的归属同步到新设计师
    if o.dept == "design" and old_wid:
        await db.execute(
            update(models.Feedback).where(
                models.Feedback.project_id == o.project_id,
                models.Feedback.status == "pending_design",
                models.Feedback.designer_uid == old_wid,
            ).values(designer_uid=w.id)
        )
    await db.commit()

    await push_message(db, to_user_id=w.id, kind="info",
                       text=f"【任务转交】{p.code} {p.name} 由 {old_name} 转交给你，请接续{cfg['done_label']}。",
                       biz_type="order", biz_id=o.id)
    await push_message(db, to_role=cfg["lead_role"], kind="info",
                       text=f"【已换人】{p.code} {cfg['name']} 负责人由 {old_name} → {new_name}。",
                       biz_type="order", biz_id=o.id)
    await write_audit(db, user=current, action="reassign", target_type="dept_order",
                      target_id=o.id, detail=f"{old_name}→{new_name}")
    return schemas.Msg(message=f"已将任务由 {old_name} 转交给 {new_name}")


@router.post("/{oid}/edit-due", response_model=schemas.Msg)
async def edit_due(
    oid: int, data: schemas.OrderEditDueIn,
    current: models.User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """🆕 管理层修改任务「预计完成时间」(due_date)。不受「本人填写后不可改」限制，
    任意状态(待接单/进行中/已完成)均可改；仅 admin/manager。"""
    if not _is_mgr(current):
        raise HTTPException(403, "仅管理层可修改预计完成时间")
    if not _valid_date(data.due_date):
        raise HTTPException(400, "预计完成日期格式应为 YYYY-MM-DD")
    o = await _order_or_404(db, oid)
    if o.start_date and data.due_date < o.start_date:
        raise HTTPException(400, "预计完成不能早于开始日期")
    old = o.due_date
    o.due_date = data.due_date
    await db.commit()
    await write_audit(db, user=current, action="edit_due", target_type="dept_order",
                      target_id=o.id, detail=f"{old}→{data.due_date}")
    return schemas.Msg(message="已更新预计完成时间")
