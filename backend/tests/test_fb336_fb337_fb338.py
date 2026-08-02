"""反馈 #336/#337/#338（李新新，采购）：

#337 清单导入没有通知（真 bug）——图纸(附件)走 start-push 有推送，清单走 Excel 导入
     全程静默，采购员只能靠微信群里喊。现在按 BUYER_SHEET_MAP 分工域通知对应采购员，
     重导用「【更新】…请以最新为准」口径（与 #323 图纸二次推送一致）。
#338 采购部列表原来按项目编号倒序，看不出谁先谁后 → 附件补 pushed_at、清单出 sheet_times，
     前端按「本人可见列」的最大时间排序。
#336 采购明细加物料关键字（名称/规格/品牌/单号），空格分词逐词 AND —— 规格写法不统一
     （`6016.0` vs `GB／T276-94深沟球轴承6016-2R`），整串 LIKE 命中率太低。
"""
import asyncio, os, sys, tempfile, io

tmp = tempfile.mkdtemp(prefix="fb336")
os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{tmp}/test.db"
os.environ["FILES_DIR"] = f"{tmp}/files"
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.getcwd())

from httpx import AsyncClient, ASGITransport
from openpyxl import Workbook
from sqlalchemy import select
from app.main import app
from app.database import engine, SessionLocal, Base
from app import models
from app.seed import seed
from app.data_migration import run_all, ensure_schema_columns

FAIL = []
def chk(c, m):
    if not c: FAIL.append(m); print("FAIL:", m)
    else: print("  ok:", m)


def xlsx(sheet_name: str, rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["序号", "名称", "规格", "数量"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def main():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    await ensure_schema_columns(engine)
    async with SessionLocal() as db:
        await seed(db); await run_all(db)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test", timeout=60) as c:
        r = await c.post("/api/auth/login", json={"username": "admin", "password": "admin123"})
        H = {"Authorization": f"Bearer {r.json()['access_token']}"}
        rid = {x["code"]: x["id"] for x in (await c.get("/api/admin/roles", headers=H)).json()}

        async def mk(uname, rc, full):
            r = await c.post("/api/admin/users", headers=H, json={
                "username": uname, "password": "pass123", "full_name": full, "role_id": rid[rc]})
            assert r.status_code == 200, r.text
            return r.json()["id"]

        async def login(u):
            r = await c.post("/api/auth/login", json={"username": u, "password": "pass123"})
            return {"Authorization": f"Bearer {r.json()['access_token']}"}

        async def msgs(h):
            return [m["text"] for m in (await c.get("/api/messages", headers=h)).json()]

        ids = {}
        for u, rc, fn in [("s1", "sales", "赵仁辉"), ("dl", "design_lead", "陈工"),
                          ("d1", "designer", "张工"), ("sm", "sheetmetal", "何师傅"),
                          ("lixinxin", "buyer", "李新新"), ("wangqin", "buyer", "王芹"),
                          ("fangbusen", "buyer", "方步森")]:
            ids[u] = await mk(u, rc, fn)
        Hs1, Hdl, Hd1 = await login("s1"), await login("dl"), await login("d1")
        Hlxx, Hwq, Hfbs = await login("lixinxin"), await login("wangqin"), await login("fangbusen")

        r = await c.post("/api/sales/orders", headers=Hs1, json={
            "name": "清单通知机", "customer": "x", "cust_type": "经销商", "contract": "有",
            "amount": 100000, "tax_rate": "13%", "prepay": 0, "before_ship": 0,
            "ship_receivable": 0, "balance": 0, "balance_date": "", "depts": ["design"],
            "receiver": {"name": "a", "phone": "1", "addr": "b"}})
        assert r.status_code == 200, r.text
        pid, code = r.json()["project_id"], r.json()["code"]
        oid = [o for o in (await c.get("/api/orders?dept=design", headers=Hdl)).json()
               if o["project_id"] == pid][0]["id"]
        await c.post(f"/api/orders/{oid}/assign", headers=Hdl, json={"worker_id": ids["d1"]})
        await c.post(f"/api/orders/{oid}/start", headers=Hd1,
                     json={"start_date": "2026-07-01", "due_date": "2026-12-31"})

        sheets = {d["name"]: d["id"] for d in
                  (await c.get(f"/api/projects/{pid}/datasheets", headers=Hd1)).json()}

        # ================= #337 清单导入 → 通知对应分工域的采购员 =================
        print("\n===== #337 清单导入通知 =====")
        chk("标准件清单" in sheets, f"项目建好就有标准件清单模板表: {list(sheets)}")

        r = await c.post(f"/api/datasheets/{sheets['标准件清单']}/import-excel", headers=Hd1,
                         files=[("file", ("标准件.xlsx", xlsx("标准件清单", [[1, "轴承", "6016", 2]]),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
        chk(r.status_code == 200, f"标准件清单导入成功: {r.status_code} {r.text[:120]}")

        m_lxx = await msgs(Hlxx)
        chk(any("标准件清单" in t and code in t for t in m_lxx),
            f"#337 标准件清单导入 → 李新新收到通知（原来完全静默）: {[t for t in m_lxx if code in t]}")
        chk(all("【更新】" not in t for t in m_lxx if code in t),
            "#337 首次导入用「已上传」口径，不是【更新】")
        chk(not any(code in t for t in await msgs(Hwq)),
            "#337 标准件清单不推给王芹（standard 域只有李新新）")
        chk(not any(code in t for t in await msgs(Hfbs)),
            "#337 标准件清单不推给方步森")

        # 重导 → 【更新】口径
        r = await c.post(f"/api/datasheets/{sheets['标准件清单']}/import-excel", headers=Hd1,
                         files=[("file", ("标准件2.xlsx", xlsx("标准件清单", [[1, "轴承", "6016", 5]]),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
        chk(r.status_code == 200, f"标准件清单重导成功: {r.status_code}")
        upd = [t for t in await msgs(Hlxx) if code in t and "【更新】" in t]
        chk(any("标准件清单" in t and "请以最新为准" in t for t in upd),
            f"#337 重导 → 李新新收【更新】口径（与 #323 图纸二次推送一致）: {upd}")

        # 激光件清单 → laser 域 = 王芹，李新新不收（与 #324 图纸路由同口径）
        if "激光件清单" in sheets:
            await c.post(f"/api/datasheets/{sheets['激光件清单']}/import-excel", headers=Hd1,
                         files=[("file", ("激光.xlsx", xlsx("激光件清单", [[1, "侧板", "δ2", 4]]),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
            chk(any("激光件清单" in t and code in t for t in await msgs(Hwq)),
                "#337 激光件清单导入 → 王芹收到通知")
            chk(not any("激光件清单" in t for t in await msgs(Hlxx)),
                "#337 李新新不收激光件清单通知（分工域一致，不越界打扰）")

        # 整包导入（/projects/{pid}/import-excel）——设计部另一条上传路径，同样要通知。
        # 关键回归：新项目建好时模板表就已存在(imported_at=None)，不能因此把首次导入
        # 判成「【更新】」——判据必须是这张表自己导过没有。
        print("\n----- #337 整包导入（多表一次传）-----")
        r = await c.post("/api/sales/orders", headers=Hs1, json={
            "name": "整包导入机", "customer": "y", "cust_type": "经销商", "contract": "有",
            "amount": 50000, "tax_rate": "13%", "prepay": 0, "before_ship": 0,
            "ship_receivable": 0, "balance": 0, "balance_date": "", "depts": ["design"],
            "receiver": {"name": "a", "phone": "1", "addr": "b"}})
        pid2, code2 = r.json()["project_id"], r.json()["code"]

        wb = Workbook()
        wb.remove(wb.active)
        for sn, row in [("标准件清单", [1, "螺栓", "M8", 20]),
                        ("激光件清单", [1, "面板", "δ1.5", 3]),
                        ("外协加工", [1, "机架", "焊接", 1])]:
            ws = wb.create_sheet(sn)
            ws.append(["序号", "名称", "规格", "数量"])
            ws.append(row)
        buf = io.BytesIO(); wb.save(buf)
        r = await c.post(f"/api/projects/{pid2}/import-excel", headers=Hd1,
                         files=[("file", ("三表.xlsx", buf.getvalue(),
                                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
        chk(r.status_code == 200, f"整包导入成功: {r.status_code} {r.text[:120]}")

        t_lxx = [t for t in await msgs(Hlxx) if code2 in t]
        t_wq = [t for t in await msgs(Hwq) if code2 in t]
        t_fbs = [t for t in await msgs(Hfbs) if code2 in t]
        chk(len(t_lxx) == 1 and "标准件清单" in t_lxx[0],
            f"#337 整包导入：李新新只收 1 条（多表合并不刷屏）: {t_lxx}")
        chk(len(t_wq) == 1 and "激光件清单" in t_wq[0], f"#337 王芹收激光件清单: {t_wq}")
        chk(len(t_fbs) == 1 and "外协加工" in t_fbs[0], f"#337 方步森收外协加工: {t_fbs}")
        chk(all("【更新】" not in t for t in t_lxx + t_wq + t_fbs),
            f"#337 首次整包导入不能发成【更新】（模板表早就在，但从没导过）: {t_lxx + t_wq + t_fbs}")

        buf2 = io.BytesIO(); wb.save(buf2)
        await c.post(f"/api/projects/{pid2}/import-excel", headers=Hd1,
                     files=[("file", ("三表2.xlsx", buf2.getvalue(),
                                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))])
        chk(any("【更新】" in t for t in await msgs(Hlxx) if code2 in t),
            "#337 整包重导 → 才发【更新】")

        # ================= #338 推送时间 =================
        print("\n===== #338 推送时间 + 排序依据 =====")
        r = await c.post(f"/api/orders/{oid}/start-upload?kind=outsource_img", headers=Hd1,
                         files=[("files", ("外购附图.pdf", io.BytesIO(b"IMG"), "application/pdf"))])
        chk(r.status_code == 200, f"外购附图上传: {r.status_code}")
        r = await c.post(f"/api/orders/{oid}/start-push", headers=Hd1, json={"kind": "outsource_img"})
        chk(r.status_code == 200, f"外购附图推送: {r.status_code}")

        async with SessionLocal() as db:
            att = (await db.execute(select(models.Attachment).where(
                models.Attachment.kind == "outsource_img"))).scalars().first()
            chk(att is not None and att.pushed_at is not None,
                f"#338 推送动作写入 pushed_at: {att.pushed_at if att else None}")

        rows = (await c.get("/api/purchase/projects", headers=Hlxx)).json()
        row = next((x for x in rows if x["code"] == code), None)
        chk(row is not None, "采购部列表能取到该项目")
        chk(row and row.get("sheet_times", {}).get("standard"),
            f"#338 清单导入时间随行返回: {row.get('sheet_times') if row else None}")
        chk(row and row["outsource_img_files"] and row["outsource_img_files"][0].get("pushed_at"),
            "#338 附件带 pushed_at（前端按此排序 + 标「新」）")

        # 存量口径：上传即推送的老附件没有 pushed_at，读取时退回 created_at，不能是 null
        async with SessionLocal() as db:
            a2 = (await db.execute(select(models.Attachment).where(
                models.Attachment.kind == "outsource_img"))).scalars().first()
            a2.pushed_at = None
            await db.commit()
        rows = (await c.get("/api/purchase/projects", headers=Hlxx)).json()
        row = next(x for x in rows if x["code"] == code)
        chk(row["outsource_img_files"][0].get("pushed_at") is not None,
            "#338 存量附件(pushed_at 为空)读取时退回 created_at，前端不会显示 '—'")

        # 存量回填的对批逻辑：同一 kind 多次推送、且中间有文件被删过。
        # 生产实况(2026-059B/任务单160/sheetpkg)：审计声称 4+1=5 个，实际只剩 4 个。
        # 必须**从最新一批倒着对**——正着对会让幸存的都拿到较早那次的时间，
        # 而这列是给采购员「别漏单」用的，宁可偏新不可偏旧。
        print("\n----- #338 存量回填：审计条数 > 实际附件数 -----")
        from app.data_migration import backfill_attachment_pushed_at
        from datetime import datetime, timezone as _tz
        T1 = datetime(2026, 7, 31, 1, 31, tzinfo=_tz.utc)
        T2 = datetime(2026, 7, 31, 1, 34, tzinfo=_tz.utc)
        async with SessionLocal() as db:
            for i in range(4):   # 审计说 4+1=5 个，这里只建 4 个（第 4 个来自第二批）
                db.add(models.Attachment(
                    biz_type="order_start_output", biz_id=9999, kind="sheetpkg",
                    project_id=pid, name=f"图{i}.dwg", size=1, path=f"x/{i}",
                    pushed=True, pushed_at=None))
            for ts, n in [(T1, 4), (T2, 1)]:
                db.add(models.AuditLog(
                    user_id=ids["d1"], action="push", target_type="dept_order",
                    target_id=9999, detail=f"start-push:sheetpkg x{n}", created_at=ts))
            await db.commit()
            await backfill_attachment_pushed_at(db)
            rows4 = (await db.execute(select(models.Attachment)
                     .where(models.Attachment.biz_id == 9999)
                     .order_by(models.Attachment.id))).scalars().all()
            got = [a.pushed_at.replace(tzinfo=_tz.utc) if a.pushed_at
                   and a.pushed_at.tzinfo is None else a.pushed_at for a in rows4]
            chk(got[-1] == T2,
                f"#338 回填倒着对批：最后一个附件拿到最新那批的时间: {got[-1]}")
            chk(all(g == T1 for g in got[:3]),
                f"#338 前三个拿到第一批的时间: {got[:3]}")
            chk(all(g is not None for g in got), "#338 回填后没有留空")

        # ================= #336 物料关键字查询 =================
        print("\n===== #336 物料查询（历史价格 / 比价）=====")
        r = await c.post("/api/purchase-mgmt/suppliers", headers=Hlxx,
                         json={"name": "江苏鸿旭隆", "category": None})
        sup1 = r.json()["id"]
        r = await c.post("/api/purchase-mgmt/suppliers", headers=Hlxx,
                         json={"name": "无锡轴承商行", "category": None})
        sup2 = r.json()["id"]

        # 规格两种写法：Excel 转出来的 `6016.0` 和手填的国标全称——同一个轴承
        for sid, name, spec, price in [
            (sup1, "轴承", "6016.0", 62.0),
            (sup2, "轴承", "GB／T276-94深沟球轴承6016-2RS", 55.0),
            (sup1, "轴承", "6006.0", 6.0),
            (sup1, "链条", "10A-1", 25.0),
        ]:
            r = await c.post("/api/purchase-mgmt/items", headers=Hlxx, json={
                "supplier_id": sid, "item_name": name, "spec": spec,
                "qty": 1, "unit_price": price, "delivery_date": "2026-07-01"})
            assert r.status_code == 200, r.text

        async def items(**kw):
            return (await c.get("/api/purchase-mgmt/items", headers=Hlxx, params=kw)).json()

        async def count(**kw):
            return (await c.get("/api/purchase-mgmt/items/summary",
                                headers=Hlxx, params=kw)).json()["count"]

        got = await items(keyword="轴承")
        chk(len(got) == 3, f"#336 按名称查「轴承」得 3 条: {len(got)}")
        chk(await count(keyword="轴承") == 3, "#336 summary 计数随关键字联动（分页总数才对）")

        got = await items(keyword="6016")
        specs = sorted(x["spec"] for x in got)
        chk(len(got) == 2 and specs == ["6016.0", "GB／T276-94深沟球轴承6016-2RS"],
            f"#336 按规格查「6016」两种写法都命中（比价的关键）: {specs}")
        prices = sorted(x["unit_price"] for x in got)
        chk(prices == [55.0, 62.0], f"#336 同一规格两家供应商的价都查得到: {prices}")

        got = await items(keyword="轴承 6016")
        chk(len(got) == 2, f"#336 空格分词逐词 AND：「轴承 6016」得 2 条: {len(got)}")

        chk(len(await items(keyword="链条")) == 1, "#336 关键字不误伤其他物料")
        chk(len(await items(keyword="不存在的料")) == 0, "#336 查不到就是 0 条，不是全量")
        chk(len(await items()) == 4, "#336 不传关键字仍是全量（不影响原有行为）")

        # 关键字必须叠加行级权限，不能成为越权口子
        got_fbs = (await c.get("/api/purchase-mgmt/items", headers=Hfbs,
                               params={"keyword": "轴承"})).json()
        chk(got_fbs == [], f"#336 关键字不绕过采购员行级隔离（方步森查不到李新新的料）: {len(got_fbs)}")

    await engine.dispose()
    print("\nPASSED" if not FAIL else f"\n{len(FAIL)} FAILURES")
    sys.exit(1 if FAIL else 0)


asyncio.run(main())
