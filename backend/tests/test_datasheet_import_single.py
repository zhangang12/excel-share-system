"""🆕 2026-07-27：单表导入 POST /api/datasheets/{did}/import-excel（设计部五表逐张上传）。
1. 只替换目标表：字段/行重建，同项目其他表（钣金装配）字段 id/行数据完全不动；
2. datasheet 本身 id/name/sort_order 保留，header_lines 更新，imported_at 更新；
3. 进度列空白自动填「进行中」，Excel 已填值保留不覆盖；
4. 模板对齐按目标表名（文件内 sheet 名叫 Sheet1 也按「标准件清单」模板排）；
5. 权限：未登录 401 / 无项目编辑权 403 / datasheet 不存在 404。
"""
import asyncio, os, sys, tempfile
from io import BytesIO

tmp = tempfile.mkdtemp(prefix="impsingle")
os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{tmp}/test.db"
os.environ["FILES_DIR"] = f"{tmp}/files"
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.getcwd())

from openpyxl import Workbook
from sqlalchemy import select
from httpx import AsyncClient, ASGITransport
from app.main import app
from app.database import engine, SessionLocal, Base
from app.seed import seed
from app.data_migration import run_all, ensure_schema_columns
from app import models

FAIL = []
def chk(c, m):
    if not c: FAIL.append(m); print("FAIL:", m)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# 标准件清单模板列序（map_excel_to_template 按位置映射，进度在第 9 列）
STD_HEADERS = ['项目', '规格型号', '数量', '材质', '品牌', '采购负责人', '订购日期', '到货日期', '进度']
BJ_HEADERS = ['名称', '图纸名称', '编号', '工艺1', '工艺1发出日期', '工艺1完成日期',
              '工艺2', '工艺2发出日期', '工艺2完成日期', '进度', '负责人签字', '备注', '库位']


def make_full_xlsx() -> bytes:
    """整导用：钣金装配 2 行 + 标准件清单 2 行（均无项目头）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "钣金装配"
    ws.append(BJ_HEADERS)
    ws.append(['A件', 'B01', 'X1', '', '', '', '', '', '', '完成', '', '', ''])
    ws.append(['B件', 'B02', 'X2', '', '', '', '', '', '', '', '', '', ''])
    ws2 = wb.create_sheet("标准件清单")
    ws2.append(STD_HEADERS)
    ws2.append(['旧件A', 'X', 1, '', '', '', '', '', '完成'])
    ws2.append(['旧件B', 'Y', 2, '', '', '', '', '', ''])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def make_single_xlsx() -> bytes:
    """单表导入用：sheet 名故意叫 Sheet1（验证按目标表名对齐模板），
    带 2 行项目头（验证 header_lines 更新）；进度列 完成/空白/None 三态。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(['同辉机械项目管理系统'])
    ws.append(['项目编号：IMP-S01'])
    ws.append(STD_HEADERS)
    ws.append(['螺栓M8', '8.8级', 100, '', '', '', '', '', '完成'])
    ws.append(['螺母M8', '8.8级', 50, '', '', '', '', '', ''])
    ws.append(['垫片', '标准', 200, '', '', '', '', '', None])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def main():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    await ensure_schema_columns(engine)
    async with SessionLocal() as db:
        await seed(db); await run_all(db)
        p = models.Project(code="IMP-S01", name="单表导入测试", status="进行中")
        db.add(p)
        await db.commit()
        pid = p.id

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        H = {"Authorization": f"Bearer {(await c.post('/api/auth/login', json={'username':'admin','password':'admin123'})).json()['access_token']}"}

        # ===== 0. 先整导（钣金装配 + 标准件清单）=====
        r = await c.post(f"/api/projects/{pid}/import-excel", headers=H,
                         files={"file": ("五表.xlsx", make_full_xlsx(), XLSX_MIME)})
        chk(r.status_code == 200, f"整导成功: {r.status_code} {r.text[:200]}")

        async with SessionLocal() as db:
            bj = (await db.execute(select(models.Datasheet).where(
                models.Datasheet.project_id == pid, models.Datasheet.name == "钣金装配"))).scalar_one()
            std = (await db.execute(select(models.Datasheet).where(
                models.Datasheet.project_id == pid, models.Datasheet.name == "标准件清单"))).scalar_one()
            bj_id = bj.id
            bj_fields = {f.id: f.name for f in (await db.execute(select(models.Field).where(
                models.Field.datasheet_id == bj.id))).scalars().all()}
            bj_recs = sorted(
                [tuple(sorted((r.values or {}).items())) for r in (await db.execute(
                    select(models.Record).where(models.Record.datasheet_id == bj.id))).scalars().all()]
            )
            std_id = std.id
            std_sort = std.sort_order
            std_old_imported_at = std.imported_at
            chk(std_old_imported_at is not None, "整导后标准件清单已有 imported_at")
            chk(std.header_lines is None, f"整导(无项目头)后 header_lines 为空: {std.header_lines!r}")

        # ===== 1. 权限：未登录 401 / 无编辑权 403 / 不存在 404 =====
        r = await c.post(f"/api/datasheets/{std_id}/import-excel",
                         files={"file": ("x.xlsx", make_single_xlsx(), XLSX_MIME)})
        chk(r.status_code == 401, f"未登录 401: {r.status_code}")
        rid = {x["code"]: x["id"] for x in (await c.get("/api/admin/roles", headers=H)).json()}
        r = await c.post("/api/admin/users", headers=H,
                         json={"username": "w1", "password": "pass123", "full_name": "w1", "role_id": rid["warehouse"]})
        assert r.status_code == 200, r.text
        Hw1 = {"Authorization": f"Bearer {(await c.post('/api/auth/login', json={'username':'w1','password':'pass123'})).json()['access_token']}"}
        # 系统默认给所有员工补 edit 成员关系（backfill_members_all_users_all_projects），
        # 把 w1 降为 view 成员 → user_can_edit_project=False → 403
        async with SessionLocal() as db:
            w1 = (await db.execute(select(models.User).where(models.User.username == "w1"))).scalar_one()
            m = (await db.execute(select(models.ProjectMember).where(
                models.ProjectMember.project_id == pid,
                models.ProjectMember.user_id == w1.id))).scalar_one()
            m.permission = "view"
            await db.commit()
        r = await c.post(f"/api/datasheets/{std_id}/import-excel", headers=Hw1,
                         files={"file": ("x.xlsx", make_single_xlsx(), XLSX_MIME)})
        chk(r.status_code == 403, f"只读成员无编辑权 403: {r.status_code} {r.text[:120]}")
        r = await c.post("/api/datasheets/999999/import-excel", headers=H,
                         files={"file": ("x.xlsx", make_single_xlsx(), XLSX_MIME)})
        chk(r.status_code == 404, f"datasheet 不存在 404: {r.status_code}")

        # ===== 2. 单表导入标准件清单 =====
        r = await c.post(f"/api/datasheets/{std_id}/import-excel", headers=H,
                         files={"file": ("标准件清单.xlsx", make_single_xlsx(), XLSX_MIME)})
        chk(r.status_code == 200, f"单表导入成功: {r.status_code} {r.text[:200]}")
        chk("标准件清单" in r.json().get("message", "") and "3" in r.json().get("message", ""),
            f"返回 message 带表名+行数: {r.json()}")

        async with SessionLocal() as db:
            std = (await db.execute(select(models.Datasheet).where(
                models.Datasheet.id == std_id))).scalar_one()
            # 2a. datasheet 本身保留 + imported_at/header_lines 更新
            chk(std.name == "标准件清单" and std.sort_order == std_sort,
                f"id/name/sort_order 保留: {std.id}/{std.name}/{std.sort_order}")
            chk(std.imported_at is not None and std.imported_at != std_old_imported_at,
                f"imported_at 更新: {std_old_imported_at} → {std.imported_at}")
            chk(std.header_lines is not None and "IMP-S01" in std.header_lines,
                f"header_lines 更新为项目头: {std.header_lines!r}")
            # 2b. 字段/行重建（SQLite 会复用刚删除的 rowid，不能比字段 id——
            #     用内容判定：行数据整体换新、字段按模板列序）
            fields = list((await db.execute(select(models.Field).where(
                models.Field.datasheet_id == std_id).order_by(models.Field.sort_order))).scalars().all())
            names = [f.name for f in fields]
            chk(names[0] == "项目" and "进度" in names and "预计到货" in names,
                f"按目标表名(标准件清单)模板对齐: {names}")
            pf = next(f for f in fields if f.name == "进度")
            xf = next(f for f in fields if f.name == "项目")
            recs = list((await db.execute(select(models.Record).where(
                models.Record.datasheet_id == std_id).order_by(models.Record.sort_order))).scalars().all())
            chk(len(recs) == 3, f"3 行入库: {len(recs)}")
            by_item = {r.values.get(str(xf.id)): (r.values.get(str(pf.id)) or "") for r in recs}
            chk(set(by_item) == {"螺栓M8", "螺母M8", "垫片"},
                f"旧行(旧件A/B)已替换为新行: {sorted(by_item)}")
            chk(by_item.get("螺栓M8") == "完成", f"已填值保留不覆盖: {by_item.get('螺栓M8')!r}")
            chk(by_item.get("螺母M8") == "进行中", f"空白填进行中(螺母): {by_item.get('螺母M8')!r}")
            chk(by_item.get("垫片") == "进行中", f"None 填进行中(垫片): {by_item.get('垫片')!r}")

            # ===== 3. 其他表（钣金装配）完全不动 =====
            bj_fields2 = {f.id: f.name for f in (await db.execute(select(models.Field).where(
                models.Field.datasheet_id == bj_id))).scalars().all()}
            bj_recs2 = sorted(
                [tuple(sorted((r.values or {}).items())) for r in (await db.execute(
                    select(models.Record).where(models.Record.datasheet_id == bj_id))).scalars().all()]
            )
            chk(bj_fields2 == bj_fields, "钣金装配字段 id/名称完全不动")
            chk(bj_recs2 == bj_recs, "钣金装配行数据完全不动")

        # ===== 4. 列表接口带出 imported_at（前端显示已导入日期）=====
        r = await c.get(f"/api/projects/{pid}/datasheets", headers=H)
        chk(r.status_code == 200, f"数据表列表: {r.status_code}")
        it = next((x for x in r.json() if x["name"] == "标准件清单"), None)
        chk(it is not None and it.get("imported") and it.get("imported_at"),
            f"列表返回 imported_at: {it}")

    await engine.dispose()
    if FAIL:
        print(f"\n{len(FAIL)} 项失败"); sys.exit(1)
    print("PASSED")

asyncio.run(main())
