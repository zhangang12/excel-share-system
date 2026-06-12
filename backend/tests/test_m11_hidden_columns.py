"""M11 制图列逻辑删除：导出不含制图列、数据保留、模板带hidden标记"""
import asyncio, os, sys, tempfile, shutil
tmp = tempfile.mkdtemp(prefix="m11test")
os.environ["DATABASE_URL"] = f"sqlite+aiosqlite:///{tmp}/test.db"
os.environ["FILES_DIR"] = f"{tmp}/files"
os.chdir(r"D:/opencode-project/EXCEL共享维护系统1500/EXCEL共享维护系统/v2/backend"); sys.path.insert(0, os.getcwd())

from httpx import AsyncClient, ASGITransport
from app.main import app
from app.database import engine, SessionLocal, Base
from app.seed import seed
from app.data_migration import run_all, ensure_schema_columns

FAIL = []
def chk(c, m):
    if not c: FAIL.append(m); print("FAIL:", m)

async def main():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    await ensure_schema_columns(engine)
    async with SessionLocal() as db:
        await seed(db); await run_all(db)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        r = await c.post("/api/auth/login", json={"username":"admin","password":"admin123"})
        H = {"Authorization": f"Bearer {r.json()['access_token']}"}

        # 模板端点：制图列带 hidden=True（数据链路保留的证明：条目仍在）
        r = await c.get("/api/projects/_overview/template", headers=H)
        fields = r.json()["fields"]
        labels = [f["label"] for f in fields]
        chk("制图开始" in labels, "制图列定义仍在模板(数据链路保留)")
        hid = {f["label"] for f in fields if f.get("hidden")}
        chk(hid == {"制图开始","制图结束","制图用时"}, f"hidden 标记: {hid}")

        # 建项目并写制图数据（验证数据链路未断）
        r = await c.post("/api/projects", headers=H, json={"code":"T-011","name":"测试"})
        pid = r.json()["id"]
        r = await c.put(f"/api/projects/{pid}/header-cell", headers=H,
                        json={"key":"制图开始","value":"2026-06-01","is_overview": True})
        chk(r.status_code==200, f"制图数据仍可写入: {r.status_code} {r.text[:120]}")
        r = await c.get(f"/api/projects/{pid}", headers=H)
        chk(r.json()["overview_meta"].get("制图开始")=="2026-06-01", "制图数据已存(overview_meta.制图开始)")

        # 一览导出：不含制图列、其余列在
        r = await c.get("/api/overview/export", headers=H)
        chk(r.status_code==200, f"export 200: {r.status_code}")
        from openpyxl import load_workbook
        from io import BytesIO
        ws = load_workbook(BytesIO(r.content)).active
        headers = [c2.value for c2 in ws[1]]
        chk("制图开始" not in headers and "制图结束" not in headers and "制图用时" not in headers,
            f"导出不含制图列: {headers}")
        chk("设计师" in headers and "电工" in headers, f"其余列保留: {headers}")

    await engine.dispose()
    print("PASSED" if not FAIL else f"{len(FAIL)} FAILURES")
    shutil.rmtree(tmp, ignore_errors=True)
    sys.exit(1 if FAIL else 0)

asyncio.run(main())
