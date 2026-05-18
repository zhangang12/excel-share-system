"""Round-3: Excel 内容正确性 / 类型推断 / overview WS / 输入边界 / 并发 / 校验。

依赖: openpyxl, websockets (host 上)
"""
import urllib.request, urllib.error, json, time, asyncio, io
HOST = 'http://localhost:8000'
WS_HOST = 'ws://localhost:8000'

def call(method, url, data=None, tok=None, raw=None, ct=None):
    h = {}
    if tok: h['Authorization'] = f'Bearer {tok}'
    if data is not None:
        h['Content-Type'] = 'application/json'
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
    elif raw is not None:
        body = raw
        if ct: h['Content-Type'] = ct
    else:
        body = None
    req = urllib.request.Request(HOST + url, method=method, data=body, headers=h)
    try:
        resp = urllib.request.urlopen(req)
        b = resp.read()
        ctype = resp.headers.get('Content-Type', '')
        if 'json' in ctype:
            return resp.status, json.loads(b or b'null')
        return resp.status, b
    except urllib.error.HTTPError as e:
        b = e.read()
        try: return e.code, json.loads(b or b'null')
        except: return e.code, b.decode('utf-8', errors='replace')

def login(u, p):
    s, r = call('POST', '/api/auth/login', {'username': u, 'password': p})
    if s != 200: raise SystemExit(f'login {u}: {s} {r}')
    return r['access_token']

def upload(url, tok, filename, content):
    boundary = '----b' + str(int(time.time()*1000))
    body = (f'--{boundary}\r\nContent-Disposition: form-data; name="file"; '
            f'filename="{filename}"\r\nContent-Type: application/octet-stream\r\n\r\n'
            ).encode() + content + f'\r\n--{boundary}--\r\n'.encode()
    return call('POST', url, raw=body, tok=tok,
                ct=f'multipart/form-data; boundary={boundary}')

bugs = []
def expect(c, m, is_bug=True):
    print(f'  [{"PASS" if c else "BUG"}] {m}')
    if not c: bugs.append(m)
def header(t): print(f'\n{"="*60}\n {t}\n{"="*60}')

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    raise SystemExit('需要 openpyxl: pip install openpyxl')

atok = login('admin', 'admin123')
ts = str(int(time.time()))

# ============ A. Excel 导出内容正确性 ============
header('A. Excel 导出内容正确性')

_, r = call('POST', '/api/projects', {'code': f'EXP-{ts}', 'name': '导出测试'}, tok=atok)
pid_a = r['id']
_, r = call('POST', f'/api/projects/{pid_a}/datasheets', {'name': '表A'}, tok=atok)
did_a = r['id']
_, fA = call('POST', f'/api/datasheets/{did_a}/fields', {'name': '名称', 'type': 'text'}, tok=atok)
_, fB = call('POST', f'/api/datasheets/{did_a}/fields', {'name': '数量', 'type': 'text'}, tok=atok)
fid_A, fid_B = fA['id'], fB['id']
# 两行数据
for nm, qty in [('齿轮', '10'), ('轴承', '25')]:
    call('POST', f'/api/datasheets/{did_a}/records',
         {'values': {str(fid_A): nm, str(fid_B): qty}}, tok=atok)

# 导出 datasheet
s, body = call('GET', f'/api/datasheets/{did_a}/export', tok=atok)
expect(s == 200 and isinstance(body, bytes), f'导出 datasheet 返回二进制: {s}')
if isinstance(body, bytes):
    wb = load_workbook(io.BytesIO(body))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    expect(rows[0] == ('名称', '数量'), f'导出表头正确: {rows[0]}')
    expect(('齿轮', '10') in rows, f'导出数据行1正确: 齿轮/10 in {rows[1:]}')
    expect(('轴承', '25') in rows, f'导出数据行2正确: 轴承/25 in {rows[1:]}')

# 导出整个 project（多 sheet）
call('POST', f'/api/projects/{pid_a}/datasheets', {'name': '表B'}, tok=atok)
s, body = call('GET', f'/api/projects/{pid_a}/export', tok=atok)
expect(s == 200 and isinstance(body, bytes), f'导出 project 返回二进制: {s}')
if isinstance(body, bytes):
    wb = load_workbook(io.BytesIO(body))
    expect(set(wb.sheetnames) == {'表A', '表B'}, f'project 导出含两 sheet: {wb.sheetnames}')

# ============ B. 导入 → 导出 round-trip ============
header('B. 导入 → 导出 round-trip')

# 构造 xlsx：前 3 行 preamble + 第 4 行表头 + 数据
wb = Workbook(); ws = wb.active; ws.title = '进度'
ws.append(['某某公司进度表'])      # row1 公司标题
ws.append(['项目', '2026-RT'])     # row2
ws.append([])                      # row3
ws.append(['工序', '负责人', '完成日期'])   # row4 表头
ws.append(['下料', '张三', '2026-05-01'])
ws.append(['折弯', '李四', '2026-05-03'])
bio = io.BytesIO(); wb.save(bio)
rt_xlsx = bio.getvalue()

_, r = call('POST', '/api/projects', {'code': f'RT-{ts}', 'name': 'roundtrip'}, tok=atok)
pid_b = r['id']
s, r = upload(f'/api/projects/{pid_b}/import-excel', atok, 'rt.xlsx', rt_xlsx)
expect(s == 200, f'导入 round-trip xlsx: {s} {r}')

_, dlist = call('GET', f'/api/projects/{pid_b}/datasheets', tok=atok)
expect(len(dlist) == 1 and dlist[0]['name'] == '进度', f'导入产生 1 个数据表"进度": {[d["name"] for d in dlist]}')
did_b = dlist[0]['id']
_, recs = call('GET', f'/api/datasheets/{did_b}/records', tok=atok)
expect(len(recs) == 2, f'导入 2 行数据: {len(recs)}')

# 导出回来
s, body = call('GET', f'/api/datasheets/{did_b}/export', tok=atok)
if isinstance(body, bytes):
    wb = load_workbook(io.BytesIO(body))
    rows = list(wb.active.iter_rows(values_only=True))
    expect(rows[0] == ('工序', '负责人', '完成日期'), f'round-trip 表头: {rows[0]}')
    vals = [r for r in rows[1:]]
    expect(any('张三' in str(r) for r in vals), f'round-trip 含"张三": {vals}')

# ============ C. 字段类型推断 ============
header('C. 字段类型推断')

wb = Workbook(); ws = wb.active
ws.append(['标题行']); ws.append([]); ws.append([])
ws.append(['编号', '价格', '日期列', '状态'])    # row4 表头
# 价格全数字, 日期列全日期字符串, 状态少量重复值
import datetime as _dt
data = [
    ['A001', 100, '2026-01-01', '进行中'],
    ['A002', 200, '2026-01-02', '已完成'],
    ['A003', 300, '2026-01-03', '进行中'],
    ['A004', 400, '2026-01-04', '已完成'],
    ['A005', 500, '2026-01-05', '进行中'],
]
for row in data: ws.append(row)
bio = io.BytesIO(); wb.save(bio)
type_xlsx = bio.getvalue()

_, r = call('POST', '/api/projects', {'code': f'TYP-{ts}', 'name': 'typeinfer'}, tok=atok)
pid_c = r['id']
s, r = upload(f'/api/projects/{pid_c}/import-excel', atok, 'typ.xlsx', type_xlsx)
expect(s == 200, f'导入类型推断 xlsx: {s}')
_, dlist = call('GET', f'/api/projects/{pid_c}/datasheets', tok=atok)
_, flds = call('GET', f'/api/datasheets/{dlist[0]["id"]}/fields', tok=atok)
ftypes = {f['name']: f['type'] for f in flds}
print(f'  推断结果: {ftypes}')
expect(ftypes.get('价格') == 'number', f'"价格"列推断为 number: {ftypes.get("价格")}')
expect(ftypes.get('日期列') == 'date', f'"日期列"推断为 date: {ftypes.get("日期列")}')
# 系统设计：不再推断 select（所有重复值列也按 text，让用户自由输入）
expect(ftypes.get('状态') == 'text', f'"状态"推断为 text(不做 select 推断): {ftypes.get("状态")}')
expect(ftypes.get('编号') == 'text', f'"编号"推断为 text: {ftypes.get("编号")}')

# ============ D. overview WebSocket 协作 ============
header('D. overview WebSocket 协作广播')

mtok = login('manager', 'manager123')

async def ovw_ws_test():
    try:
        import websockets
    except ImportError:
        print('  (websockets 未安装，跳过)'); return []
    res = []
    async def recv_until(ws, pred, timeout=5):
        import time as _t
        dl = _t.monotonic() + timeout
        while _t.monotonic() < dl:
            try:
                raw = await asyncio.wait_for(ws.recv(), dl - _t.monotonic())
            except Exception:
                return None
            try: d = json.loads(raw)
            except: continue
            if pred(d): return d
        return None

    ws1 = await websockets.connect(f'{WS_HOST}/ws/overview?token={atok}', open_timeout=5)
    ws2 = await websockets.connect(f'{WS_HOST}/ws/overview?token={mtok}', open_timeout=5)
    try:
        # ws1 应收到 ws2 join
        d = await recv_until(ws1, lambda m: m.get('type') == 'presence' and m.get('action') == 'join')
        res.append(('overview presence join', d is not None, d))

        # 取一个 overview 字段和项目
        _, ov = call('GET', '/api/overview', tok=atok)
        if ov['fields'] and ov['rows']:
            ovf = ov['fields'][0]['id']
            ovp = ov['rows'][0]['id']
            # ws2(manager) 改 overview cell → ws1 应收到 cell_changed
            call('PUT', f'/api/overview/projects/{ovp}/cell',
                 {'field_id': ovf, 'value': f'WS测试{ts}'}, tok=mtok)
            d = await recv_until(ws1, lambda m: m.get('type') == 'cell_changed'
                                 and m.get('field_id') == ovf and m.get('project_id') == ovp)
            ok = d and d.get('value') == f'WS测试{ts}' and d.get('record_id') is None
            res.append(('overview cell_changed 广播(project_id 不为空)', ok, d))
        else:
            res.append(('overview cell_changed', False, 'no fields/rows'))
    finally:
        await ws1.close()
        try: await ws2.close()
        except: pass
    return res

for label, ok, info in asyncio.run(ovw_ws_test()):
    expect(ok, f'{label}: {info}')

# ============ E. 输入边界 / 特殊字符 / XSS ============
header('E. 输入边界 / 特殊字符 / XSS')

# XSS 内容应原样存储（前端 Vue 自动转义，不在此测渲染）
xss = '<script>alert(1)</script>'
s, r = call('POST', '/api/projects', {'code': f'XSS-{ts}', 'name': xss}, tok=atok)
expect(s == 200 and r['name'] == xss, f'XSS 内容原样存储: {r.get("name")==xss}')
xss_pid = r['id'] if s == 200 else None

# emoji + 中文 cell 值
if xss_pid:
    _, r = call('POST', f'/api/projects/{xss_pid}/datasheets', {'name': '表😀'}, tok=atok)
    emoji_did = r['id']
    _, r = call('POST', f'/api/datasheets/{emoji_did}/fields', {'name': '备注', 'type': 'text'}, tok=atok)
    emoji_fid = r['id']
    _, r = call('POST', f'/api/datasheets/{emoji_did}/records', {'values': {}}, tok=atok)
    emoji_rid = r['id']
    s, r = call('PUT', f'/api/records/{emoji_rid}/cell',
                {'field_id': emoji_fid, 'value': '测试😀✓特殊&<>"\''}, tok=atok)
    expect(s == 200 and r['values'][str(emoji_fid)] == '测试😀✓特殊&<>"\'',
           f'emoji+特殊字符 cell 正确往返')

# 超长字段名（>128）→ 422
longname = '字' * 200
s, r = call('POST', f'/api/datasheets/{emoji_did}/fields' if xss_pid else '/api/x',
            {'name': longname, 'type': 'text'}, tok=atok)
expect(s == 422, f'超长字段名(200字)被拒: {s}')

# 超长项目 code（>64）→ 422
s, r = call('POST', '/api/projects', {'code': 'C'*100, 'name': 'x'}, tok=atok)
expect(s == 422, f'超长项目编号(100字符)被拒: {s}')

# 超长 cell 值（应当能存，values 是 JSON 无长度限制）
if xss_pid:
    bigval = 'x' * 10000
    s, r = call('PUT', f'/api/records/{emoji_rid}/cell',
                {'field_id': emoji_fid, 'value': bigval}, tok=atok)
    expect(s == 200 and len(r['values'][str(emoji_fid)]) == 10000,
           f'超长 cell 值(1万字符)可存: {s}')

# ============ F. 并发编辑同一 cell ============
header('F. 并发编辑同一单元格 (last-write-wins)')

if xss_pid:
    import threading
    results_cc = []
    def writer(val):
        s, r = call('PUT', f'/api/records/{emoji_rid}/cell',
                    {'field_id': emoji_fid, 'value': val}, tok=atok)
        results_cc.append((val, s))
    threads = [threading.Thread(target=writer, args=(f'concurrent-{i}',)) for i in range(8)]
    for t in threads: t.start()
    for t in threads: t.join()
    all_200 = all(s == 200 for _, s in results_cc)
    expect(all_200, f'8 个并发写同一 cell 全部 200(无 500/死锁): {[s for _,s in results_cc]}')
    # 最终值应是 8 个写入之一
    s, recs = call('GET', f'/api/datasheets/{emoji_did}/records', tok=atok)
    final = next((x for x in recs if x['id'] == emoji_rid), {}).get('values', {}).get(str(emoji_fid))
    expect(str(final).startswith('concurrent-'), f'并发后最终值是某次写入: {final}')

# ============ G. 校验边界 / 重复数据 / logout ============
header('G. 数据校验 / 重复数据 / logout')

# 空 code
s, r = call('POST', '/api/projects', {'code': '', 'name': 'x'}, tok=atok)
expect(s == 422, f'空项目编号被拒: {s}')
# 用户名 < 2
s, r = call('POST', '/api/admin/users', {'username': 'a', 'password': 'abcdef', 'role_id': 3}, tok=atok)
expect(s == 422, f'用户名过短被拒: {s}')
# 密码 < 6
s, r = call('POST', '/api/admin/users',
            {'username': f'u{ts}', 'password': '123', 'role_id': 3}, tok=atok)
expect(s == 422, f'密码过短被拒: {s}')
# 非法字段类型
if xss_pid:
    s, r = call('POST', f'/api/datasheets/{emoji_did}/fields',
                {'name': 'badtype', 'type': 'nonsense'}, tok=atok)
    expect(s == 400, f'非法字段类型被拒: {s}')
# 重复项目 code
s, r = call('POST', '/api/projects', {'code': f'XSS-{ts}', 'name': 'dup'}, tok=atok)
expect(s == 400, f'重复项目编号被拒: {s}')
# 重复用户名
s, r = call('POST', '/api/admin/users',
            {'username': 'admin', 'password': 'abcdef', 'role_id': 3}, tok=atok)
expect(s == 400, f'重复用户名被拒: {s}')
# 重复 overview 字段名
s, r = call('POST', '/api/overview/fields', {'name': f'dupf{ts}', 'type': 'text'}, tok=atok)
dupf_id = r['id'] if s == 200 else None
s, r = call('POST', '/api/overview/fields', {'name': f'dupf{ts}', 'type': 'text'}, tok=atok)
expect(s == 400, f'重复一览字段名被拒: {s}')
if dupf_id: call('DELETE', f'/api/overview/fields/{dupf_id}', tok=atok)
# logout
s, r = call('POST', '/api/auth/logout', tok=atok)
expect(s == 200, f'logout: {s} {r}')
# logout 后 token 仍有效（JWT 无状态，这是预期行为）
s, r = call('GET', '/api/auth/me', tok=atok)
expect(s == 200, f'logout 后 JWT 仍有效(无状态,符合设计): {s}')

# ============ Cleanup ============
header('Cleanup')
for p in (pid_a, pid_b, pid_c, xss_pid):
    if p: call('DELETE', f'/api/projects/{p}', tok=atok)
print('  测试项目已清理')

# ============ Summary ============
header('Summary')
print(f'BUGs: {len(bugs)}')
for m in bugs: print(f'   - {m}')
if not bugs: print('  ✓ 全部通过')
